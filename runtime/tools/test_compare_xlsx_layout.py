import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from compare_xlsx_layout import compare_layout


class CompareXlsxLayoutTests(unittest.TestCase):
    def _workbook(self, path: Path, *, value: str, bold: bool = False) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "翻译清单"
        sheet["C68"] = value
        sheet["C68"].font = Font(bold=bold)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = "A1:C68"
        sheet.merge_cells("A1:C1")
        sheet.column_dimensions["C"].width = 24
        workbook.save(path)

    def test_value_only_change_keeps_layout_equal(self):
        with tempfile.TemporaryDirectory() as directory:
            before = Path(directory) / "before.xlsx"
            after = Path(directory) / "after.xlsx"
            self._workbook(before, value="京成杯秋季让赛")
            self._workbook(after, value="京成杯秋季赛")
            result = compare_layout(before, after)
            self.assertTrue(result["matched"])

    def test_style_change_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            before = Path(directory) / "before.xlsx"
            after = Path(directory) / "after.xlsx"
            self._workbook(before, value="京成杯秋季让赛")
            self._workbook(after, value="京成杯秋季赛", bold=True)
            with self.assertRaisesRegex(ValueError, "layout differs"):
                compare_layout(before, after)


if __name__ == "__main__":
    unittest.main()

class CompareXlsxLayoutNegativeTests(unittest.TestCase):
    def _workbook(self, path: Path, **overrides) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "翻译清单"
        sheet["C68"] = "京成杯秋季让赛"
        sheet.freeze_panes = overrides.get("freeze_panes", "A2")
        sheet.auto_filter.ref = overrides.get("auto_filter", "A1:C68")
        if overrides.get("merge", True):
            sheet.merge_cells("A1:C1")
        sheet.column_dimensions["C"].width = overrides.get("width", 24)
        workbook.save(path)

    def _pair(self, directory: str, **after_overrides):
        before = Path(directory) / "before.xlsx"
        after = Path(directory) / "after.xlsx"
        self._workbook(before)
        self._workbook(after, **after_overrides)
        return before, after

    def test_merge_change_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            before, after = self._pair(directory, merge=False)
            with self.assertRaisesRegex(ValueError, "layout differs"):
                compare_layout(before, after)

    def test_column_width_change_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            before, after = self._pair(directory, width=30)
            with self.assertRaisesRegex(ValueError, "layout differs"):
                compare_layout(before, after)

    def test_freeze_panes_change_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            before, after = self._pair(directory, freeze_panes="A3")
            with self.assertRaisesRegex(ValueError, "layout differs"):
                compare_layout(before, after)

    def test_auto_filter_change_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            before, after = self._pair(directory, auto_filter="A1:C10")
            with self.assertRaisesRegex(ValueError, "layout differs"):
                compare_layout(before, after)
