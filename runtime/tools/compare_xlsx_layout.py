#!/usr/bin/env python3
"""Fail-closed semantic layout comparison for two XLSX workbooks."""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.xml.functions import tostring


def _xml(value: Any) -> str:
    if value is None:
        return ""
    return tostring(value.to_tree()).decode("utf-8")


def _header_footer_snapshot(value: Any) -> dict[str, Any]:
    return {
        "left": value.left.text,
        "center": value.center.text,
        "right": value.right.text,
    }


def _dimension_snapshot(dimensions: Any) -> list[dict[str, Any]]:
    rows = []
    for key in sorted(dimensions):
        dimension = dimensions[key]
        rows.append(
            {
                "key": str(key),
                "hidden": bool(dimension.hidden),
                "outlineLevel": int(dimension.outlineLevel or 0),
                "collapsed": bool(dimension.collapsed),
                "style": int(dimension.style_id or 0),
                "height": getattr(dimension, "height", None),
                "width": getattr(dimension, "width", None),
                "min": getattr(dimension, "min", None),
                "max": getattr(dimension, "max", None),
                "bestFit": bool(getattr(dimension, "bestFit", False)),
                "customWidth": bool(getattr(dimension, "customWidth", False)),
                "customFormat": bool(getattr(dimension, "customFormat", False)),
            }
        )
    return rows


def _cell_style_snapshot(sheet: Any) -> list[dict[str, Any]]:
    rows = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.has_style:
                rows.append(
                    {
                        "coordinate": cell.coordinate,
                        "font": _xml(cell.font),
                        "fill": _xml(cell.fill),
                        "border": _xml(cell.border),
                        "alignment": _xml(cell.alignment),
                        "protection": _xml(cell.protection),
                        "numberFormat": cell.number_format,
                    }
                )
    return rows


def _conditional_formatting_snapshot(sheet: Any) -> list[dict[str, Any]]:
    rows = []
    for conditional_formatting in sheet.conditional_formatting:
        rows.append(
            {
                "sqref": str(conditional_formatting.sqref),
                "rules": [_xml(rule) for rule in conditional_formatting.rules],
            }
        )
    return rows


def _sheet_snapshot(sheet: Any) -> dict[str, Any]:
    return {
        "title": sheet.title,
        "state": sheet.sheet_state,
        "maxRow": sheet.max_row,
        "maxColumn": sheet.max_column,
        "mergedCells": sorted(str(value) for value in sheet.merged_cells.ranges),
        "rowDimensions": _dimension_snapshot(sheet.row_dimensions),
        "columnDimensions": _dimension_snapshot(sheet.column_dimensions),
        "cellStyles": _cell_style_snapshot(sheet),
        "freezePanes": str(sheet.freeze_panes or ""),
        "autoFilter": _xml(sheet.auto_filter),
        "dataValidations": _xml(sheet.data_validations),
        "conditionalFormatting": _conditional_formatting_snapshot(sheet),
        "tables": [
            {"name": name, "xml": _xml(sheet.tables[name])}
            for name in sorted(sheet.tables)
        ],
        "sheetProperties": _xml(sheet.sheet_properties),
        "sheetFormat": _xml(sheet.sheet_format),
        "sheetViews": _xml(sheet.views),
        "sheetProtection": _xml(sheet.protection),
        "printOptions": _xml(sheet.print_options),
        "pageMargins": _xml(sheet.page_margins),
        "pageSetup": _xml(sheet.page_setup),
        "headerFooter": {
            "oddHeader": _header_footer_snapshot(sheet.oddHeader),
            "oddFooter": _header_footer_snapshot(sheet.oddFooter),
            "evenHeader": _header_footer_snapshot(sheet.evenHeader),
            "evenFooter": _header_footer_snapshot(sheet.evenFooter),
            "firstHeader": _header_footer_snapshot(sheet.firstHeader),
            "firstFooter": _header_footer_snapshot(sheet.firstFooter),
        },
        "printArea": str(sheet.print_area or ""),
        "printTitleRows": str(sheet.print_title_rows or ""),
        "printTitleColumns": str(sheet.print_title_cols or ""),
        "images": len(sheet._images),
        "charts": len(sheet._charts),
    }


def workbook_layout_snapshot(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        return {
            "schemaVersion": "xlsx-layout-snapshot.v1",
            "sheetOrder": list(workbook.sheetnames),
            "activeSheetIndex": int(workbook.index(workbook.active)),
            "definedNames": [
                _xml(value)
                for _, value in sorted(workbook.defined_names.items())
            ],
            "workbookProtection": _xml(workbook.security),
            "sheets": [_sheet_snapshot(sheet) for sheet in workbook.worksheets],
        }
    finally:
        workbook.close()


def _sha256_json(value: Any) -> str:
    payload = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def compare_layout(before: Path, after: Path) -> dict[str, Any]:
    before_snapshot = workbook_layout_snapshot(before)
    after_snapshot = workbook_layout_snapshot(after)
    before_sha256 = _sha256_json(before_snapshot)
    after_sha256 = _sha256_json(after_snapshot)
    if before_snapshot != after_snapshot:
        raise ValueError(
            "XLSX structural layout differs (values/formulas are compared "
            "separately by the revision gate): "
            f"before={before_sha256}, after={after_sha256}"
        )
    return {
        "schemaVersion": "xlsx-layout-comparison.v1",
        "matched": True,
        "layoutSha256": before_sha256,
        "sheetCount": len(before_snapshot["sheets"]),
        "sheetNames": before_snapshot["sheetOrder"],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--before", required=True, type=Path)
    parser.add_argument("--after", required=True, type=Path)
    args = parser.parse_args()
    print(
        json.dumps(
            compare_layout(args.before.resolve(), args.after.resolve()),
            ensure_ascii=False,
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
