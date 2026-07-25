from __future__ import annotations

import json
from io import StringIO
import tempfile
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

from django.db import IntegrityError, connection, transaction
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase, SimpleTestCase, override_settings
from django.test.utils import CaptureQueriesContext
from django.utils import timezone

from stable.models import (
    HorseIdentityEvidenceCommitReceipt,
    HorseP0Source,
    HorseP0SourceStatus,
    HorseP0SourceType,
    HorseProfile,
    HorseProfileCompleteness,
    HorseProfileStatus,
    OperationLog,
    RaceEvent,
    RaceEventDataQuality,
    RaceEventRunner,
    RaceEventStatus,
    RaceEventSurface,
    RaceEventVisibility,
    RaceGrade,
    RacingRegion,
    SourceLanguage,
    TermEntry,
    TermType,
)
from stable.services import p0_horse_identity_bootstrap as identity_bootstrap
from stable.services.p0_horse_identity_bootstrap import (
    P0HorseIdentityBootstrapError,
    approve_identity_bootstrap_artifact,
    commit_identity_bootstrap_artifact,
    fetch_dual_source_identity,
    prepare_identity_bootstrap_batch,
    select_identity_bootstrap_batch,
)


@dataclass(frozen=True)
class StubResponse:
    text: str
    url: str
    status_code: int = 200
    headers: dict[str, str] = field(default_factory=dict)


class FixtureTransport:
    def __init__(self, pages: dict[str, str]):
        self.pages = pages
        self.calls: list[str] = []

    def get(self, url: str, **kwargs):
        self.calls.append(url)
        for marker, body in self.pages.items():
            if marker in url:
                return StubResponse(body, url)
        raise AssertionError(f"unexpected URL: {url}")


class ScriptedTransport:
    def __init__(self, responses: list[StubResponse]):
        self.responses = list(responses)
        self.calls: list[str] = []
        self.call_kwargs: list[dict[str, object]] = []

    def get(self, url: str, **kwargs):
        self.calls.append(url)
        self.call_kwargs.append(kwargs)
        if not self.responses:
            raise AssertionError(f"unexpected URL: {url}")
        response = self.responses.pop(0)
        return StubResponse(
            response.text,
            response.url or url,
            status_code=response.status_code,
            headers=response.headers,
        )


class FakeClock:
    def __init__(self):
        self.value = 0.0
        self.sleeps: list[float] = []

    def monotonic(self):
        return self.value

    def sleep(self, seconds: float):
        self.sleeps.append(seconds)
        self.value += seconds


NETKEIBA_PROFILE = """
<div class="horse_title"><h1>テストホース</h1><p class="txt_01">現役牡3歳鹿毛</p></div>
<table class="db_prof_table">
  <tr><th>生年月日</th><td>2022年3月15日</td></tr>
</table>
"""
NETKEIBA_PEDIGREE = """
<table class="blood_table">
  <tr><td data-role="sire">父馬</td></tr>
  <tr><td data-role="dam">母馬</td></tr>
</table>
"""
class DualSourceIdentityTests(SimpleTestCase):
    JRA_URL = (
        "https://www.jra.go.jp/JRADB/accessU.html"
        "?CNAME=pw01dud102022100001%2FAA"
    )
    NAR_URL = (
        "https://www.keiba.go.jp/KeibaWeb/DataRoom/RaceHorseInfo"
        "?k_lineageLoginCode=30034409046"
    )
    JRA_PROFILE = """
    <h1>テストホース</h1><table>
      <tr><th>生年月日</th><td>2022年3月15日</td></tr>
      <tr><th>父</th><td>父馬</td></tr><tr><th>母</th><td>母馬</td></tr>
      <tr><th>調教師</th><td>テスト調教師（栗東）</td></tr>
    </table>
    """
    NAR_PROFILE = """
    <h1>テストホース</h1><dl>
      <dt>生年月日</dt><dd>2022年3月15日</dd>
      <dt>父</dt><dd>父馬</dd><dt>母</dt><dd>母馬</dd>
      <dt>調教師</dt><dd>地方調教師</dd><dt>所属</dt><dd>大井</dd>
    </dl>
    """

    def _candidate(self, *, providers=("jra",)):
        urls = {"jra": self.JRA_URL, "nar": self.NAR_URL}
        return {
            "profile_id": 1,
            "candidate_key": "profile:1",
            "horse_name": "テストホース",
            "netkeiba_id": "2022100001",
            "qualification": [
                {
                    "race_event_id": 100 + index,
                    "grade": "G1",
                    "race_date": "2024-05-05",
                    "racecourse": "東京" if provider == "jra" else "大井",
                    "official_provider": provider,
                    "official_race_url": "",
                    "official_horse_url": urls[provider],
                    "official_source_horse_id": (
                        "pw01dud102022100001%2FAA"
                        if provider == "jra"
                        else "30034409046"
                    ),
                    "horse_number": "7",
                    "horse_name": "テストホース",
                }
                for index, provider in enumerate(providers)
            ],
        }

    def _fetch(self, candidate, pages):
        transport = FixtureTransport(
            {
                "/horse/ped/": NETKEIBA_PEDIGREE,
                "db.netkeiba.com/horse/": NETKEIBA_PROFILE,
                **pages,
            }
        )
        temporary = tempfile.TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        session = identity_bootstrap.IdentityRequestSession(
            transport=transport,
            allow_network=True,
            environment_network_enabled=True,
            cache_dir=temporary.name,
            parser_fingerprint="dual-source-v1",
            config_fingerprint="dual-source-config-v1",
            interval_seconds={"jra": 0, "nar": 0, "netkeiba": 0},
        )
        return fetch_dual_source_identity(candidate, request_session=session), transport

    def test_netkeiba_jra_consensus_is_grade_a(self):
        result, transport = self._fetch(
            self._candidate(), {"accessU.html": self.JRA_PROFILE}
        )
        self.assertEqual(result["status"], "candidate_pass")
        self.assertEqual(result["identity_mode"], "NETKEIBA_JRA_CONSENSUS")
        self.assertEqual(result["identity_evidence_grade"], "A")
        self.assertEqual(result["fields"]["birth_date"], "2022-03-15")
        self.assertEqual(result["official_providers"], ["jra"])
        self.assertEqual(len(transport.calls), 3)

    def test_jra_and_nar_consensus_is_grade_a_plus(self):
        result, transport = self._fetch(
            self._candidate(providers=("jra", "nar")),
            {
                "accessU.html": self.JRA_PROFILE,
                "RaceHorseInfo": self.NAR_PROFILE,
            },
        )
        self.assertEqual(result["status"], "candidate_pass")
        self.assertEqual(result["identity_mode"], "NETKEIBA_JRA_NAR_CONSENSUS")
        self.assertEqual(result["identity_evidence_grade"], "A+")
        self.assertEqual(result["official_providers"], ["jra", "nar"])
        self.assertEqual(len(transport.calls), 4)

    def test_any_official_conflict_blocks_entire_candidate(self):
        conflict = self.NAR_PROFILE.replace("2022年3月15日", "2022年3月16日")
        result, _ = self._fetch(
            self._candidate(providers=("jra", "nar")),
            {"accessU.html": self.JRA_PROFILE, "RaceHorseInfo": conflict},
        )
        self.assertEqual(result["status"], "blocker")
        self.assertEqual(result["reason"], "BIRTH_DATE_MISMATCH")

    def test_prepare_writes_complete_artifact_and_resumes_offline(self):
        candidate = self._candidate()
        candidate.update(
            {
                "highest_grade": "G1",
                "highest_grade_priority": 1,
                "has_official_identity_anchor": True,
                "has_complete_official_context": False,
                "training_scope_status": "provisional_japan",
                "training_evidence": [],
                "profile_snapshot": {},
                "qualification_sha256": identity_bootstrap._sha256(
                    candidate["qualification"]
                ),
            }
        )
        manifest = {
            "schema_version": identity_bootstrap.SCHEMA_VERSION,
            "status": "selected",
            "parser_version": identity_bootstrap.PARSER_VERSION,
            "config_fingerprint": identity_bootstrap._selection_config_fingerprint(),
            "horses": [candidate],
        }
        manifest["input_sha256"] = identity_bootstrap._sha256(manifest)
        transport = FixtureTransport(
            {
                "/horse/ped/": NETKEIBA_PEDIGREE,
                "db.netkeiba.com/horse/": NETKEIBA_PROFILE,
                "accessU.html": self.JRA_PROFILE,
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            first = prepare_identity_bootstrap_batch(
                manifest,
                output_dir=tmp,
                transport=transport,
                allow_network=True,
                environment_network_enabled=True,
                request_interval_seconds=0,
            )
            network_calls = len(transport.calls)
            resumed = prepare_identity_bootstrap_batch(
                manifest,
                output_dir=tmp,
                transport=FixtureTransport({}),
                allow_network=False,
                environment_network_enabled=False,
                request_interval_seconds=0,
            )
            artifact = json.loads(
                (Path(tmp) / "artifact.json").read_text(encoding="utf-8")
            )
            from openpyxl import load_workbook

            review_sheet = load_workbook(
                Path(tmp) / "review.xlsx", read_only=True
            )["身份审核"]
            headers = list(next(review_sheet.iter_rows(values_only=True)))
            review_row = dict(
                zip(headers, next(review_sheet.iter_rows(min_row=2, values_only=True)))
            )
            for name in (
                "qualification.jsonl",
                "candidates.jsonl",
                "blockers.jsonl",
                "summary.json",
                "source_evidence_manifest.json",
                "request_ledger.json",
                "state.json",
                "review.xlsx",
                "artifact.json",
            ):
                self.assertTrue((Path(tmp) / name).is_file(), name)
        self.assertEqual(first["summary"]["candidate_pass"], 1)
        self.assertEqual(resumed["summary"]["candidate_pass"], 1)
        self.assertEqual(network_calls, 3)
        self.assertIn('"race_event_id": 100', review_row["all_qualifications"])
        self.assertIn("accessU.html", review_row["official_anchors"])
        self.assertIn('"sire_name": "父馬"', review_row["netkeiba_raw_identity"])
        self.assertIn('"sire_name": "父馬"', review_row["official_raw_identity"])
        self.assertEqual(review_row["evidence_grade"], "A")
        prepared_candidate = artifact["candidates"][0]
        for field in (
            "highest_grade_priority",
            "has_official_identity_anchor",
            "has_complete_official_context",
            "training_scope_status",
            "training_evidence",
        ):
            self.assertEqual(prepared_candidate[field], candidate[field])
        self.assertEqual(set(artifact["artifact_paths"]), {
            "qualification",
            "candidates",
            "blockers",
            "summary",
            "source_evidence",
            "request_ledger",
            "workbook",
            "state",
        })


class JravanOfflinePackageTests(SimpleTestCase):
    def _package(self, root: Path, *, include_raw_record: bool = False):
        input_manifest = {
            "schema_version": identity_bootstrap.JRAVAN_INPUT_SCHEMA_VERSION,
            "records": [
                {
                    "profile_id": 1,
                    "candidate_key": "profile:1",
                    "netkeiba_id": "2022100001",
                }
            ],
        }
        input_manifest["manifest_sha256"] = identity_bootstrap._sha256(
            input_manifest
        )
        input_path = root / "jravan_input_manifest.json"
        identity_bootstrap._write_json(input_path, input_manifest)
        record = {
            "profile_id": 1,
            "candidate_key": "profile:1",
            "netkeiba_id": "2022100001",
            "record_type": "UM",
            "blood_registration_number": "2022100001",
            "registered_name": "テストホース",
            "sire_name": "父馬",
            "dam_name": "母馬",
            "birth_date": "2022-03-15",
            "data_spec_version": "4.9.0",
            "snapshot_at": "2026-07-25T12:00:00+09:00",
        }
        if include_raw_record:
            record["raw_record"] = "must-not-leave-windows-node"
        record["record_sha256"] = identity_bootstrap._sha256(record)
        identity_path = root / "horse_identity.jsonl"
        identity_bootstrap._write_jsonl(identity_path, [record])
        output_manifest = {
            "schema_version": identity_bootstrap.JRAVAN_OUTPUT_SCHEMA_VERSION,
            "input_file_sha256": identity_bootstrap._file_sha256(input_path),
            "identity_file_sha256": identity_bootstrap._file_sha256(identity_path),
            "record_type": "UM",
            "data_spec_version": "4.9.0",
            "snapshot_at": "2026-07-25T12:00:00+09:00",
            "record_count": 1,
        }
        output_manifest["manifest_sha256"] = identity_bootstrap._sha256(
            output_manifest
        )
        output_path = root / "jravan_output_manifest.json"
        identity_bootstrap._write_json(output_path, output_manifest)
        return input_path, identity_path, output_path

    def test_complete_um_exchange_validates_offline(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = self._package(Path(tmp))
            result = identity_bootstrap.validate_jravan_offline_package(
                input_manifest_path=paths[0],
                identity_jsonl_path=paths[1],
                output_manifest_path=paths[2],
            )
        self.assertEqual(result["data_spec_version"], "4.9.0")
        self.assertEqual(
            result["records"][0]["blood_registration_number"], "2022100001"
        )

    def test_manifest_drift_and_raw_um_record_fail_closed(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = self._package(Path(tmp), include_raw_record=True)
            with self.assertRaisesRegex(
                P0HorseIdentityBootstrapError, "outside manifest or incomplete"
            ):
                identity_bootstrap.validate_jravan_offline_package(
                    input_manifest_path=paths[0],
                    identity_jsonl_path=paths[1],
                    output_manifest_path=paths[2],
                )


class OfficialRaceContextResolverTests(SimpleTestCase):
    def _candidate(self, *, provider: str = "jra"):
        race_url = (
            "https://www.jra.go.jp/JRADB/accessS.html?fixture=index"
            if provider == "jra"
            else "https://www.keiba.go.jp/KeibaWeb/TodayRaceInfo/RaceMarkTable?fixture=index"
        )
        return {
            "profile_id": 1,
            "candidate_key": "profile:1",
            "horse_name": "テストホース",
            "netkeiba_id": "2022100001",
            "qualification": [
                {
                    "race_event_id": 101,
                    "grade": "G2",
                    "race_date": "2024-05-05",
                    "racecourse": "東京" if provider == "jra" else "大井",
                    "official_provider": provider,
                    "official_race_url": race_url,
                    "official_horse_url": "",
                    "official_source_horse_id": "",
                    "horse_number": "7",
                    "horse_name": "テストホース",
                }
            ],
        }

    def test_jra_index_to_unique_detail_row_returns_same_provider_anchor(self):
        index_html = """
        <a class="race-detail" href="/JRADB/accessS.html?fixture=detail"
           data-race-date="2024-05-05" data-racecourse="東京">重賞テスト</a>
        """
        detail_html = """
        <table class="race-entries">
          <tr data-horse-number="7">
            <td class="horse-number">7</td>
            <td class="horse-name">
              <a href="/JRADB/accessU.html?CNAME=pw01dud102022100001%2FAA">テストホース</a>
            </td>
          </tr>
        </table>
        """
        transport = FixtureTransport(
            {"fixture=index": index_html, "fixture=detail": detail_html}
        )

        result = identity_bootstrap.resolve_official_horse_anchor(
            self._candidate(),
            transport=transport,
            allow_network=True,
            environment_network_enabled=True,
            request_interval_seconds=0,
        )

        self.assertEqual(result["status"], "anchor_pass")
        self.assertEqual(result["provider"], "jra")
        self.assertIn("accessU.html", result["official_horse_url"])
        self.assertEqual(result["matched_row"]["horse_number"], "7")
        self.assertEqual(result["matched_row"]["horse_name"], "テストホース")
        self.assertEqual(len(result["hops"]), 2)
        self.assertTrue(all(len(hop["content_sha256"]) == 64 for hop in result["hops"]))
        self.assertEqual(len(transport.calls), 2)

    def test_context_resolution_uses_persistent_request_session_budget(self):
        index_html = """
        <a href="/JRADB/accessS.html?fixture=detail"
           data-race-date="2024-05-05" data-racecourse="東京">重賞テスト</a>
        """
        detail_html = """
        <tr data-horse-number="7"><td>7</td>
          <td><a href="/JRADB/accessU.html?CNAME=horse-1">テストホース</a></td>
        </tr>
        """
        transport = FixtureTransport(
            {"fixture=index": index_html, "fixture=detail": detail_html}
        )
        with tempfile.TemporaryDirectory() as tmp:
            session = identity_bootstrap.IdentityRequestSession(
                transport=transport,
                allow_network=True,
                environment_network_enabled=True,
                cache_dir=tmp,
                parser_fingerprint="resolver-v1",
                config_fingerprint="resolver-config-v1",
                interval_seconds={"jra": 0, "nar": 0, "netkeiba": 0},
            )
            result = identity_bootstrap.resolve_official_horse_anchor(
                self._candidate(),
                request_session=session,
            )
            ledger = session.ledger()

        self.assertEqual(result["status"], "anchor_pass")
        budget = ledger["budgets"]["profile:1"]
        self.assertEqual(budget["official_transfers"], 2)
        self.assertEqual(len(budget["official_urls"]), 2)
        self.assertEqual(len(ledger["events"]), 2)

    def test_direct_jra_and_nar_anchors_require_no_context_request(self):
        for provider, url in (
            (
                "jra",
                "https://www.jra.go.jp/JRADB/accessU.html?CNAME=pw01dud102022100001%2FAA",
            ),
            (
                "nar",
                "https://www.keiba.go.jp/KeibaWeb/DataRoom/RaceHorseInfo?k_lineageLoginCode=30034409046",
            ),
        ):
            with self.subTest(provider=provider):
                candidate = self._candidate(provider=provider)
                candidate["qualification"][0]["official_horse_url"] = url
                transport = FixtureTransport({})
                result = identity_bootstrap.resolve_official_horse_anchor(
                    candidate,
                    transport=transport,
                    allow_network=False,
                    environment_network_enabled=False,
                    request_interval_seconds=0,
                )
                self.assertEqual(result["status"], "anchor_pass")
                self.assertEqual(result["provider"], provider)
                self.assertEqual(result["official_horse_url"], url)
                self.assertEqual(transport.calls, [])

    def test_zero_or_multiple_participant_rows_fail_closed_without_search(self):
        zero_rows = "<table class='race-entries'></table>"
        duplicate_rows = """
        <table class="race-entries">
          <tr data-horse-number="7"><td>7</td><td><a href="/JRADB/accessU.html?CNAME=a">テストホース</a></td></tr>
          <tr data-horse-number="7"><td>7</td><td><a href="/JRADB/accessU.html?CNAME=b">テストホース</a></td></tr>
        </table>
        """
        for body, reason in (
            (zero_rows, "OFFICIAL_CONTEXT_NOT_FOUND"),
            (duplicate_rows, "OFFICIAL_CONTEXT_AMBIGUOUS"),
        ):
            with self.subTest(reason=reason):
                candidate = self._candidate()
                candidate["qualification"][0]["official_race_url"] = (
                    "https://www.jra.go.jp/JRADB/accessS.html?fixture=detail"
                )
                transport = FixtureTransport({"fixture=detail": body})
                result = identity_bootstrap.resolve_official_horse_anchor(
                    candidate,
                    transport=transport,
                    allow_network=True,
                    environment_network_enabled=True,
                    request_interval_seconds=0,
                )
                self.assertEqual(result["status"], "blocker")
                self.assertEqual(result["reason"], reason)
                self.assertEqual(len(transport.calls), 1)
                self.assertFalse(any("search" in url.casefold() for url in transport.calls))

    def test_cross_provider_horse_link_is_rejected(self):
        candidate = self._candidate(provider="nar")
        candidate["qualification"][0]["official_race_url"] = (
            "https://www.keiba.go.jp/KeibaWeb/TodayRaceInfo/RaceMarkTable?fixture=detail"
        )
        body = """
        <table class="race-entries">
          <tr data-horse-number="7"><td>7</td>
            <td><a href="https://www.jra.go.jp/JRADB/accessU.html?CNAME=wrong">テストホース</a></td>
          </tr>
        </table>
        """
        transport = FixtureTransport({"fixture=detail": body})

        result = identity_bootstrap.resolve_official_horse_anchor(
            candidate,
            transport=transport,
            allow_network=True,
            environment_network_enabled=True,
            request_interval_seconds=0,
        )

        self.assertEqual(result["status"], "blocker")
        self.assertEqual(result["reason"], "OFFICIAL_CONTEXT_NOT_FOUND")
        self.assertEqual(len(transport.calls), 1)

    def test_provider_neutral_four_field_consensus_and_conflict(self):
        netkeiba = {
            "horse_name": "テストホース",
            "sire_name": "父馬",
            "dam_name": "母馬",
            "birth_date": "2022-03-15",
        }
        jra = {**netkeiba, "provider": "jra", "source_id": "jra-1"}
        nar = {**netkeiba, "provider": "nar", "source_id": "nar-1"}

        a_plus = identity_bootstrap.compare_identity_sources(
            netkeiba=netkeiba, official=[jra, nar]
        )
        conflict = identity_bootstrap.compare_identity_sources(
            netkeiba=netkeiba,
            official=[{**jra, "birth_date": "2022-03-16"}],
        )

        self.assertEqual(a_plus["status"], "candidate_pass")
        self.assertEqual(a_plus["identity_mode"], "NETKEIBA_JRA_NAR_CONSENSUS")
        self.assertEqual(a_plus["identity_evidence_grade"], "A+")
        self.assertEqual(conflict["status"], "blocker")
        self.assertEqual(conflict["reason"], "BIRTH_DATE_MISMATCH")

    def test_partial_date_missing_field_country_suffix_and_script_alias_fail_closed(self):
        base = {
            "horse_name": "テストホース(USA)",
            "sire_name": "父馬",
            "dam_name": "母馬",
            "birth_date": "2022-03-15",
        }
        partial = identity_bootstrap.compare_identity_sources(
            netkeiba=base,
            official=[
                {
                    **base,
                    "horse_name": "テストホース（USA）",
                    "birth_date": "2022",
                    "birth_date_precision": "year",
                    "provider": "jra",
                }
            ],
        )
        missing = identity_bootstrap.compare_identity_sources(
            netkeiba=base,
            official=[{**base, "dam_name": "", "provider": "jra"}],
        )
        suffix_conflict = identity_bootstrap.compare_identity_sources(
            netkeiba=base,
            official=[
                {
                    **base,
                    "horse_name": "テストホース(GB)",
                    "provider": "jra",
                }
            ],
        )
        script_alias = identity_bootstrap.compare_identity_sources(
            netkeiba={**base, "sire_name": "American Pharoah"},
            official=[{**base, "sire_name": "アメリカンファラオ", "provider": "jra"}],
        )

        self.assertEqual(partial["status"], "candidate_partial")
        self.assertEqual(missing["reason"], "REQUIRED_FIELD_MISSING")
        self.assertEqual(suffix_conflict["reason"], "NAME_MISMATCH")
        self.assertEqual(script_alias["reason"], "SCRIPT_ALIAS_UNRESOLVED")


class NetworkAccessContractTests(SimpleTestCase):
    def _session(
        self,
        *,
        transport,
        cache_dir,
        allow_network=True,
        environment_network_enabled=True,
        parser_fingerprint="parser-v1",
        config_fingerprint="config-v1",
        clock=None,
    ):
        clock = clock or FakeClock()
        return identity_bootstrap.IdentityRequestSession(
            transport=transport,
            allow_network=allow_network,
            environment_network_enabled=environment_network_enabled,
            cache_dir=cache_dir,
            parser_fingerprint=parser_fingerprint,
            config_fingerprint=config_fingerprint,
            interval_seconds={"jra": 2.0, "nar": 3.0, "netkeiba": 4.0},
            monotonic=clock.monotonic,
            sleep=clock.sleep,
        )

    def test_network_requires_both_gates_before_first_transport(self):
        for allow_network, environment_enabled in ((False, True), (True, False)):
            with self.subTest(
                allow_network=allow_network,
                environment_enabled=environment_enabled,
            ), tempfile.TemporaryDirectory() as tmp:
                transport = FixtureTransport({})
                session = self._session(
                    transport=transport,
                    cache_dir=tmp,
                    allow_network=allow_network,
                    environment_network_enabled=environment_enabled,
                )
                with self.assertRaisesRegex(
                    P0HorseIdentityBootstrapError, "network requires"
                ):
                    session.get(
                        candidate_key="profile:1",
                        provider="jra",
                        url="https://www.jra.go.jp/JRADB/accessS.html?fixture=1",
                        official_chain=True,
                    )
                self.assertEqual(transport.calls, [])

    def test_source_requests_require_https_for_input_and_redirects(self):
        with tempfile.TemporaryDirectory() as tmp:
            transport = FixtureTransport({})
            session = self._session(transport=transport, cache_dir=tmp)
            with self.assertRaisesRegex(
                P0HorseIdentityBootstrapError, "SOURCE_ACCESS_DENIED"
            ):
                session.get(
                    candidate_key="profile:1",
                    provider="jra",
                    url="http://www.jra.go.jp/JRADB/accessS.html?fixture=1",
                    official_chain=True,
                )
            self.assertEqual(transport.calls, [])

        redirect = StubResponse(
            "",
            "",
            status_code=302,
            headers={
                "Location": "http://www.jra.go.jp/JRADB/accessS.html?fixture=2"
            },
        )
        with tempfile.TemporaryDirectory() as tmp:
            transport = ScriptedTransport([redirect])
            session = self._session(transport=transport, cache_dir=tmp)
            with self.assertRaisesRegex(
                P0HorseIdentityBootstrapError, "SOURCE_ACCESS_DENIED"
            ):
                session.get(
                    candidate_key="profile:1",
                    provider="jra",
                    url="https://www.jra.go.jp/JRADB/accessS.html?fixture=1",
                    official_chain=True,
                )
            self.assertEqual(len(transport.calls), 1)

    def test_every_transport_request_has_connect_and_read_timeout(self):
        with tempfile.TemporaryDirectory() as tmp:
            transport = ScriptedTransport([StubResponse("ok", "")])
            session = self._session(transport=transport, cache_dir=tmp)
            session.get(
                candidate_key="profile:1",
                provider="jra",
                url="https://www.jra.go.jp/JRADB/accessS.html?fixture=1",
                official_chain=True,
            )
        self.assertEqual(
            transport.call_kwargs,
            [{"allow_redirects": False, "timeout": (5.0, 20.0)}],
        )

    def test_provider_rate_limits_are_independent(self):
        clock = FakeClock()
        transport = ScriptedTransport(
            [
                StubResponse("jra-1", ""),
                StubResponse("nar-1", ""),
                StubResponse("jra-2", ""),
            ]
        )
        with tempfile.TemporaryDirectory() as tmp:
            session = self._session(
                transport=transport, cache_dir=tmp, clock=clock
            )
            session.get(
                candidate_key="profile:1",
                provider="jra",
                url="https://www.jra.go.jp/JRADB/accessS.html?fixture=1",
                official_chain=True,
            )
            session.get(
                candidate_key="profile:1",
                provider="nar",
                url="https://www.keiba.go.jp/KeibaWeb/TodayRaceInfo/DebaTable?fixture=1",
                official_chain=True,
            )
            session.get(
                candidate_key="profile:1",
                provider="jra",
                url="https://www.jra.go.jp/JRADB/accessS.html?fixture=2",
                official_chain=True,
            )
        self.assertEqual(clock.sleeps, [2.0])

    def test_official_and_total_url_budgets_fail_before_extra_transport(self):
        with tempfile.TemporaryDirectory() as tmp:
            transport = ScriptedTransport([StubResponse("ok", "") for _ in range(6)])
            session = self._session(transport=transport, cache_dir=tmp)
            for index in range(3):
                session.get(
                    candidate_key="profile:1",
                    provider="jra",
                    url=f"https://www.jra.go.jp/JRADB/accessS.html?fixture={index}",
                    official_chain=True,
                )
            with self.assertRaisesRegex(
                P0HorseIdentityBootstrapError, "REQUEST_BUDGET_EXHAUSTED"
            ):
                session.get(
                    candidate_key="profile:1",
                    provider="jra",
                    url="https://www.jra.go.jp/JRADB/accessS.html?fixture=overflow",
                    official_chain=True,
                )
            self.assertEqual(len(transport.calls), 3)

        with tempfile.TemporaryDirectory() as tmp:
            transport = ScriptedTransport([StubResponse("ok", "") for _ in range(7)])
            session = self._session(transport=transport, cache_dir=tmp)
            for index in range(6):
                session.get(
                    candidate_key="profile:1",
                    provider="netkeiba",
                    url=f"https://db.netkeiba.com/horse/fixture-{index}/",
                    official_chain=False,
                )
            with self.assertRaisesRegex(
                P0HorseIdentityBootstrapError, "REQUEST_BUDGET_EXHAUSTED"
            ):
                session.get(
                    candidate_key="profile:1",
                    provider="netkeiba",
                    url="https://db.netkeiba.com/horse/fixture-overflow/",
                    official_chain=False,
                )
            self.assertEqual(len(transport.calls), 6)

    def test_cache_resume_works_offline_and_fingerprint_drift_rejects_cache(self):
        url = "https://www.jra.go.jp/JRADB/accessS.html?fixture=cache"
        with tempfile.TemporaryDirectory() as tmp:
            first_transport = ScriptedTransport(
                [StubResponse("<html>cached secret page</html>", "")]
            )
            first = self._session(transport=first_transport, cache_dir=tmp)
            first_response = first.get(
                candidate_key="profile:1",
                provider="jra",
                url=url,
                official_chain=True,
            )
            self.assertEqual(first_response.text, "<html>cached secret page</html>")
            self.assertTrue((Path(tmp) / "state.json").exists())

            offline_transport = FixtureTransport({})
            resumed = self._session(
                transport=offline_transport,
                cache_dir=tmp,
                allow_network=False,
                environment_network_enabled=False,
            )
            resumed_response = resumed.get(
                candidate_key="profile:1",
                provider="jra",
                url=url,
                official_chain=True,
            )
            self.assertEqual(resumed_response.text, "<html>cached secret page</html>")
            self.assertEqual(offline_transport.calls, [])
            self.assertTrue(resumed.ledger()["events"][-1]["cache_hit"])

            drifted = self._session(
                transport=FixtureTransport({}),
                cache_dir=tmp,
                allow_network=False,
                environment_network_enabled=False,
                parser_fingerprint="parser-v2",
            )
            with self.assertRaisesRegex(
                P0HorseIdentityBootstrapError, "fingerprint"
            ):
                drifted.get(
                    candidate_key="profile:1",
                    provider="jra",
                    url=url,
                    official_chain=True,
                )

    def test_429_opens_provider_circuit_and_ledger_never_contains_page_body(self):
        denied = StubResponse(
            "<html>too many requests secret body</html>",
            "",
            status_code=429,
        )
        transport = ScriptedTransport([denied, StubResponse("must-not-fetch", "")])
        with tempfile.TemporaryDirectory() as tmp:
            session = self._session(transport=transport, cache_dir=tmp)
            with self.assertRaisesRegex(
                P0HorseIdentityBootstrapError, "SOURCE_ACCESS_DENIED"
            ):
                session.get(
                    candidate_key="profile:1",
                    provider="jra",
                    url="https://www.jra.go.jp/JRADB/accessS.html?fixture=denied",
                    official_chain=True,
                )
            with self.assertRaisesRegex(
                P0HorseIdentityBootstrapError, "SOURCE_ACCESS_DENIED"
            ):
                session.get(
                    candidate_key="profile:1",
                    provider="jra",
                    url="https://www.jra.go.jp/JRADB/accessS.html?fixture=after-denied",
                    official_chain=True,
                )
            ledger_text = json.dumps(session.ledger(), ensure_ascii=False)
        self.assertEqual(len(transport.calls), 1)
        self.assertNotIn("secret body", ledger_text)
        self.assertNotIn("<html>", ledger_text)


class HorseIdentityProviderTests(SimpleTestCase):
    def _session(self, transport, cache_dir):
        clock = FakeClock()
        return identity_bootstrap.IdentityRequestSession(
            transport=transport,
            allow_network=True,
            environment_network_enabled=True,
            cache_dir=cache_dir,
            parser_fingerprint="provider-fixture-v1",
            config_fingerprint="provider-config-v1",
            interval_seconds={"jra": 0, "nar": 0, "netkeiba": 0},
            monotonic=clock.monotonic,
            sleep=clock.sleep,
        )

    @staticmethod
    def _candidate():
        return {
            "profile_id": 1,
            "candidate_key": "profile:1",
            "horse_name": "テストホース",
            "netkeiba_id": "2022100001",
        }

    def test_netkeiba_provider_extracts_only_identity_fields(self):
        transport = FixtureTransport(
            {
                "/horse/ped/": NETKEIBA_PEDIGREE,
                "db.netkeiba.com/horse/": NETKEIBA_PROFILE,
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            provider = identity_bootstrap.NetkeibaHorseIdentityProvider(
                self._session(transport, tmp)
            )
            result = provider.fetch(self._candidate())

        self.assertEqual(result["status"], "source_pass")
        self.assertEqual(
            result["identity"],
            {
                "horse_name": "テストホース",
                "sire_name": "父馬",
                "dam_name": "母馬",
                "birth_date": "2022-03-15",
                "birth_date_precision": "day",
            },
        )
        self.assertEqual(result["provider"], "netkeiba")
        self.assertEqual(len(transport.calls), 2)
        self.assertNotIn("<html", json.dumps(result, ensure_ascii=False).casefold())

    def test_direct_official_anchor_requires_provider_identifier(self):
        for provider, url in (
            ("jra", "https://www.jra.go.jp/JRADB/accessU.html"),
            (
                "nar",
                "https://www.keiba.go.jp/KeibaWeb/DataRoom/RaceHorseInfo",
            ),
        ):
            candidate = {
                **self._candidate(),
                "qualification": [
                    {
                        "official_provider": provider,
                        "official_horse_url": url,
                        "official_source_horse_id": "",
                    }
                ],
            }
            with self.subTest(provider=provider):
                result = identity_bootstrap.resolve_official_horse_anchor(
                    candidate,
                    allow_network=False,
                    environment_network_enabled=False,
                )
                self.assertEqual(result["status"], "blocker")
                self.assertEqual(result["reason"], "OFFICIAL_ANCHOR_MISSING")

    def test_jra_provider_saves_cname_and_confirmed_training_evidence(self):
        url = (
            "https://www.jra.go.jp/JRADB/accessU.html"
            "?CNAME=pw01dud102022100001%2FAA"
        )
        body = """
        <h1>テストホース</h1>
        <table class="horse-profile">
          <tr><th>生年月日</th><td>2022年3月15日</td></tr>
          <tr><th>父</th><td>父馬</td></tr>
          <tr><th>母</th><td>母馬</td></tr>
          <tr><th>調教師</th><td>テスト調教師（栗東）</td></tr>
        </table>
        """
        transport = FixtureTransport({"accessU.html": body})
        anchor = {
            "provider": "jra",
            "official_horse_url": url,
            "official_source_horse_id": "pw01dud102022100001%2FAA",
            "qualification": {"race_date": "2024-05-05", "race_event_id": 101},
        }
        with tempfile.TemporaryDirectory() as tmp:
            provider = identity_bootstrap.JraHorseIdentityProvider(
                self._session(transport, tmp)
            )
            result = provider.fetch(self._candidate(), anchor)

        self.assertEqual(result["status"], "source_pass")
        self.assertEqual(result["identity"]["sire_name"], "父馬")
        self.assertEqual(result["identity"]["birth_date"], "2022-03-15")
        self.assertEqual(result["source_id_raw"], "pw01dud102022100001%2FAA")
        self.assertEqual(result["training_scope_status"], "confirmed_japan")
        self.assertEqual(result["training_evidence"][0]["affiliation"], "栗東")
        self.assertEqual(len(transport.calls), 1)

    def test_nar_provider_saves_lineage_code_and_local_training_evidence(self):
        url = (
            "https://www.keiba.go.jp/KeibaWeb/DataRoom/RaceHorseInfo"
            "?k_lineageLoginCode=30034409046"
        )
        body = """
        <h1>テストホース</h1>
        <dl class="horse-profile">
          <dt>生年月日</dt><dd>2022年3月15日</dd>
          <dt>父</dt><dd>父馬</dd>
          <dt>母</dt><dd>母馬</dd>
          <dt>調教師</dt><dd>テスト調教師</dd>
          <dt>所属</dt><dd>大井</dd>
        </dl>
        """
        transport = FixtureTransport({"RaceHorseInfo": body})
        anchor = {
            "provider": "nar",
            "official_horse_url": url,
            "official_source_horse_id": "30034409046",
            "qualification": {"race_date": "2024-05-05", "race_event_id": 102},
        }
        with tempfile.TemporaryDirectory() as tmp:
            provider = identity_bootstrap.NarHorseIdentityProvider(
                self._session(transport, tmp)
            )
            result = provider.fetch(self._candidate(), anchor)

        self.assertEqual(result["status"], "source_pass")
        self.assertEqual(result["identity"]["dam_name"], "母馬")
        self.assertEqual(result["source_id_raw"], "30034409046")
        self.assertEqual(result["training_scope_status"], "confirmed_japan")
        self.assertEqual(result["training_evidence"][0]["affiliation"], "大井")
        self.assertEqual(len(transport.calls), 1)

    def test_provider_layout_change_fails_closed(self):
        for provider_name, url in (
            (
                "jra",
                "https://www.jra.go.jp/JRADB/accessU.html?CNAME=broken",
            ),
            (
                "nar",
                "https://www.keiba.go.jp/KeibaWeb/DataRoom/RaceHorseInfo"
                "?k_lineageLoginCode=broken",
            ),
        ):
            with self.subTest(provider=provider_name), tempfile.TemporaryDirectory() as tmp:
                transport = FixtureTransport({url: "<h1>テストホース</h1>"})
                session = self._session(transport, tmp)
                provider = (
                    identity_bootstrap.JraHorseIdentityProvider(session)
                    if provider_name == "jra"
                    else identity_bootstrap.NarHorseIdentityProvider(session)
                )
                result = provider.fetch(
                    self._candidate(),
                    {
                        "provider": provider_name,
                        "official_horse_url": url,
                        "official_source_horse_id": "broken",
                        "qualification": {
                            "race_date": "2024-05-05",
                            "race_event_id": 103,
                        },
                    },
                )
                self.assertEqual(result["status"], "blocker")
                self.assertEqual(result["reason"], "SOURCE_LAYOUT_CHANGED")


class IdentityBootstrapDatabaseTests(TestCase):
    def _profile(
        self,
        index: int,
        *,
        netkeiba_id: str | None = None,
        published: bool = False,
    ):
        term = TermEntry.objects.create(
            term_type=TermType.HORSE,
            source_language=SourceLanguage.JAPANESE,
            source_ja=f"テスト馬{index:03d}",
            racing_region=RacingRegion.JAPAN,
            is_active=True,
        )
        profile = HorseProfile.objects.create(
            primary_term=term,
            original_name=term.source_ja,
            racing_region=RacingRegion.JAPAN,
            review_status=(
                HorseProfileStatus.PUBLISHED if published else HorseProfileStatus.DRAFT
            ),
            completeness_status=HorseProfileCompleteness.EMPTY,
            published_at=timezone.now() if published else None,
            source_refs={
                "horse_identity_keys": [f"netkeiba:{netkeiba_id or 2022000000 + index}"]
            },
        )
        event = RaceEvent.objects.create(
            year=2024,
            slug=f"identity-commit-{index}",
            original_name=f"提交测试重賞{index}",
            chinese_name=f"提交测试重赏{index}",
            country_region=RacingRegion.JAPAN,
            racecourse="東京",
            grade_text=RaceGrade.G2,
            normalized_grade=RaceGrade.G2,
            surface=RaceEventSurface.TURF,
            local_date=date(2024, 6, min(index + 1, 28)),
            status=RaceEventStatus.FINISHED,
            visibility_status=RaceEventVisibility.PUBLISHED,
            data_quality_status=RaceEventDataQuality.COMPLETE,
            source_refs={
                "official": {
                    "provider": "jra",
                    "url": f"https://www.jra.go.jp/JRADB/accessS.html?commit={index}",
                }
            },
        )
        runner = RaceEventRunner.objects.create(
            event=event,
            external_runner_id=f"commit-runner-{index}",
            sort_order=1,
            horse_number=str(index + 1),
            horse_name=profile.original_name,
            source_refs=event.source_refs,
        )
        HorseP0Source.objects.create(
            profile=profile,
            race_event=event,
            race_runner=runner,
            source_type=HorseP0SourceType.MAJOR_RACE_PARTICIPANT,
            status=HorseP0SourceStatus.ACTIVE,
            racing_region=RacingRegion.JAPAN,
            race_grade=RaceGrade.G2,
            horse_name=profile.original_name,
            participant_key=f"test:{index}",
            source_url=f"https://www.jra.go.jp/JRADB/accessS.html?commit={index}",
            evidence_payload={
                "training_scope": {
                    "status": "confirmed_japan",
                    "evidence": [
                        {
                            "source": "reviewed_manual",
                            "source_id": f"jra-training-{index}",
                            "source_url": (
                                "https://www.jra.go.jp/JRADB/accessU.html"
                                f"?CNAME=training-{index}"
                            ),
                            "race_date": event.local_date.isoformat(),
                            "affiliation": "栗東",
                            "trainer_name": "テスト調教師",
                            "reviewed": True,
                        }
                    ],
                }
            },
        )
        return profile

    def test_selection_is_bounded_unique_and_excludes_old_blockers(self):
        profiles = [self._profile(i) for i in range(105)]
        excluded = {profile.pk for profile in profiles[:5]}
        with self.assertNumQueries(2):
            manifest = select_identity_bootstrap_batch(
                target_count=100,
                excluded_profile_ids=excluded,
                excluded_batch_id="p0batch-old",
                exclusion_reason="previous_batch_blockers",
                scan_limit=500,
            )
        self.assertEqual(len(manifest["horses"]), 100)
        self.assertFalse(excluded & {row["profile_id"] for row in manifest["horses"]})
        self.assertEqual(len({row["netkeiba_id"] for row in manifest["horses"]}), 100)

    def test_duplicate_netkeiba_id_fails_closed(self):
        self._profile(1, netkeiba_id="2022000001")
        self._profile(2, netkeiba_id="2022000001")
        with self.assertRaises(P0HorseIdentityBootstrapError):
            select_identity_bootstrap_batch(target_count=2)

    def _prepared_artifact(
        self,
        profiles: list[HorseProfile],
        root: Path,
        *,
        statuses: dict[int, str] | None = None,
    ):
        selected = select_identity_bootstrap_batch(target_count=len(profiles))
        selected_by_id = {
            int(row["profile_id"]): row for row in selected["horses"]
        }
        statuses = statuses or {}
        candidates = []
        for index, profile in enumerate(profiles, start=1):
            selected_row = selected_by_id[profile.pk]
            qualification = selected_row["qualification"][0]
            official_url = qualification["official_horse_url"]
            official_source_id = qualification["official_source_horse_id"]
            candidates.append(
                {
                    **selected_row,
                    "fields": {
                        "sire_text": f"父{index}",
                        "dam_text": f"母{index}",
                        "birth_date": f"2022-03-{index:02d}",
                    },
                    "evidence": {
                        "netkeiba": {
                            "url": (
                                "https://db.netkeiba.com/horse/"
                                f"{selected_row['netkeiba_id']}/"
                            ),
                            "content_sha256": "a" * 64,
                            "identity": {
                                "horse_name": selected_row["horse_name"],
                                "sire_name": f"父{index}",
                                "dam_name": f"母{index}",
                                "birth_date": f"2022-03-{index:02d}",
                                "birth_date_precision": "day",
                            },
                        },
                        "official": [
                            {
                                "provider": "jra",
                                "source_id_raw": official_source_id,
                                "url": official_url,
                                "content_sha256": "b" * 64,
                                "identity": {
                                    "horse_name": selected_row["horse_name"],
                                    "sire_name": f"父{index}",
                                    "dam_name": f"母{index}",
                                    "birth_date": f"2022-03-{index:02d}",
                                    "birth_date_precision": "day",
                                },
                            }
                        ],
                    },
                    "official_providers": ["jra"],
                    "anchors": [
                        {
                            "status": "anchor_pass",
                            "provider": "jra",
                            "official_horse_url": official_url,
                            "official_source_horse_id": official_source_id,
                            "qualification": qualification,
                            "matched_row": {},
                            "hops": [],
                        }
                    ],
                    "status": statuses.get(profile.pk, "candidate_pass"),
                    "identity_mode": "NETKEIBA_JRA_CONSENSUS",
                    "identity_evidence_grade": "A",
                }
            )
        artifact_paths = {}
        artifact_hashes = {}
        for key, filename in (
            ("qualification", "qualification.jsonl"),
            ("candidates", "candidates.jsonl"),
            ("blockers", "blockers.jsonl"),
            ("summary", "summary.json"),
            ("source_evidence", "source_evidence_manifest.json"),
            ("request_ledger", "request_ledger.json"),
            ("workbook", "review.xlsx"),
            ("state", "state.json"),
        ):
            artifact_file = root / filename
            if key in {"candidates", "blockers"}:
                identity_bootstrap._write_jsonl(
                    artifact_file,
                    candidates if key == "candidates" else [],
                )
            else:
                artifact_file.write_bytes(f"{key}\n".encode())
            artifact_paths[key] = filename
            artifact_hashes[key] = __import__("hashlib").sha256(
                artifact_file.read_bytes()
            ).hexdigest()
        artifact = {
            "schema_version": "p0-horse-identity-bootstrap.v1",
            "status": "prepared",
            "input_sha256": selected["input_sha256"],
            "parser_version": "p0-horse-identity-bootstrap-parser.v1",
            "config_fingerprint": identity_bootstrap._selection_config_fingerprint(),
            "candidates": candidates,
            "blockers": [],
            "artifact_paths": artifact_paths,
            "artifact_hashes": artifact_hashes,
        }
        path = root / "artifact.json"
        path.write_text(json.dumps(artifact, ensure_ascii=False), encoding="utf-8")
        return path

    def _approved_artifact(self, profiles: list[HorseProfile], root: Path):
        path = self._prepared_artifact(profiles, root)
        approved = approve_identity_bootstrap_artifact(
            path,
            reviewer="reviewer-a",
            approved_profile_ids=[profile.pk for profile in profiles],
        )
        return path, approved["approved_sha256"]

    def test_approval_rejects_partial_or_blocker_candidate(self):
        for index, status in enumerate(
            ("candidate_partial", "blocker"), start=1
        ):
            with self.subTest(status=status):
                profile = self._profile(index)
                with tempfile.TemporaryDirectory() as tmp:
                    path = self._prepared_artifact(
                        [profile],
                        Path(tmp),
                        statuses={profile.pk: status},
                    )
                    with self.assertRaises(P0HorseIdentityBootstrapError):
                        approve_identity_bootstrap_artifact(
                            path,
                            reviewer="reviewer-a",
                            approved_profile_ids=[profile.pk],
                        )

    def test_approval_recomputes_consensus_and_rejects_tampered_fields(self):
        profile = self._profile(1)
        with tempfile.TemporaryDirectory() as tmp:
            path = self._prepared_artifact([profile], Path(tmp))
            artifact = json.loads(path.read_text(encoding="utf-8"))
            artifact["candidates"][0]["fields"]["sire_text"] = "伪造父马"
            candidate_path = Path(tmp) / artifact["artifact_paths"]["candidates"]
            artifact["artifact_hashes"]["candidates"] = (
                identity_bootstrap._write_jsonl(
                    candidate_path,
                    artifact["candidates"],
                )
            )
            path.write_text(
                json.dumps(artifact, ensure_ascii=False),
                encoding="utf-8",
            )
            with self.assertRaisesRegex(
                P0HorseIdentityBootstrapError, "consensus"
            ):
                approve_identity_bootstrap_artifact(
                    path,
                    reviewer="reviewer-a",
                    approved_profile_ids=[profile.pk],
                )

    def test_approval_rejects_embedded_candidates_not_in_reviewed_sidecar(self):
        profile = self._profile(1)
        with tempfile.TemporaryDirectory() as tmp:
            path = self._prepared_artifact([profile], Path(tmp))
            artifact = json.loads(path.read_text(encoding="utf-8"))
            candidate = artifact["candidates"][0]
            candidate["fields"]["sire_text"] = "伪造父马"
            candidate["evidence"]["netkeiba"]["identity"]["sire_name"] = "伪造父马"
            candidate["evidence"]["official"][0]["identity"]["sire_name"] = "伪造父马"
            path.write_text(
                json.dumps(artifact, ensure_ascii=False),
                encoding="utf-8",
            )
            with self.assertRaisesRegex(
                P0HorseIdentityBootstrapError, "review package"
            ):
                approve_identity_bootstrap_artifact(
                    path,
                    reviewer="reviewer-a",
                    approved_profile_ids=[profile.pk],
                )

    def test_commit_requires_exact_approved_sha(self):
        profile = self._profile(1)
        with tempfile.TemporaryDirectory() as tmp:
            path, _ = self._approved_artifact([profile], Path(tmp))
            with self.assertRaises(P0HorseIdentityBootstrapError):
                commit_identity_bootstrap_artifact(
                    path,
                    approved_sha256="0" * 64,
                    approved_by="operator-b",
                )

    def test_approved_artifact_rejects_review_event_or_anchor_tampering(self):
        for index, drift_kind in enumerate(("review_event", "anchor"), start=1):
            with self.subTest(drift_kind=drift_kind):
                profile = self._profile(index)
                with tempfile.TemporaryDirectory() as tmp:
                    path, _ = self._approved_artifact([profile], Path(tmp))
                    artifact = json.loads(path.read_text(encoding="utf-8"))
                    if drift_kind == "review_event":
                        artifact["approval"]["reviewer"] = "tampered-reviewer"
                    else:
                        artifact["approved_candidates"][0]["evidence"]["official"][0][
                            "url"
                        ] = (
                            "https://www.jra.go.jp/JRADB/accessU.html"
                            "?CNAME=tampered"
                        )
                    artifact["approved_sha256"] = identity_bootstrap._sha256(
                        {
                            key: value
                            for key, value in artifact.items()
                            if key != "approved_sha256"
                        }
                    )
                    identity_bootstrap._write_json(path, artifact)
                    with self.assertRaises(P0HorseIdentityBootstrapError):
                        commit_identity_bootstrap_artifact(
                            path,
                            approved_sha256=artifact["approved_sha256"],
                            approved_by="operator-b",
                        )

    def test_commit_is_atomic_and_exact_replay_is_zero_write(self):
        profiles = [self._profile(1), self._profile(2)]
        with tempfile.TemporaryDirectory() as tmp:
            path, sha = self._approved_artifact(profiles, Path(tmp))
            first = commit_identity_bootstrap_artifact(
                path, approved_sha256=sha, approved_by="operator-b"
            )
            replay = commit_identity_bootstrap_artifact(
                path, approved_sha256=sha, approved_by="operator-b"
            )
            approved_artifact_sha = identity_bootstrap._file_sha256(path)
        self.assertEqual(first["profiles_written"], 2)
        self.assertEqual(replay["profiles_written"], 0)
        self.assertTrue(replay["replay"])
        self.assertEqual(HorseIdentityEvidenceCommitReceipt.objects.count(), 1)
        receipt = HorseIdentityEvidenceCommitReceipt.objects.select_related(
            "operation_log"
        ).get()
        self.assertEqual(receipt.artifact_sha256, approved_artifact_sha)
        self.assertEqual(
            receipt.evidence_summary[str(profiles[0].pk)]["netkeiba"],
            "a" * 64,
        )
        self.assertEqual(
            receipt.operation_log.action_type,
            "p0_horse_identity_evidence_commit",
        )
        self.assertEqual(
            receipt.operation_log.detail,
            json.dumps(
                receipt.result_payload,
                ensure_ascii=False,
                sort_keys=True,
            ),
        )

    def test_replay_rejects_receipt_or_audit_content_drift(self):
        for index, drift_kind in enumerate(
            ("evidence_summary", "operation_log"), start=1
        ):
            with self.subTest(drift_kind=drift_kind):
                profile = self._profile(index)
                with tempfile.TemporaryDirectory() as tmp:
                    path, sha = self._approved_artifact([profile], Path(tmp))
                    commit_identity_bootstrap_artifact(
                        path,
                        approved_sha256=sha,
                        approved_by="operator-b",
                    )
                    receipt = HorseIdentityEvidenceCommitReceipt.objects.get(
                        approved_sha256=sha
                    )
                    if drift_kind == "evidence_summary":
                        receipt.evidence_summary = {}
                        receipt.save(update_fields=["evidence_summary", "updated_at"])
                    else:
                        operation = receipt.operation_log
                        operation.detail = "{}"
                        operation.save(update_fields=["detail"])
                    with self.assertRaises(P0HorseIdentityBootstrapError):
                        commit_identity_bootstrap_artifact(
                            path,
                            approved_sha256=sha,
                            approved_by="operator-b",
                        )

    def test_commit_and_verify_commands_require_explicit_confirmation(self):
        profile = self._profile(1)
        with tempfile.TemporaryDirectory() as tmp:
            path, sha = self._approved_artifact([profile], Path(tmp))
            with self.assertRaises(CommandError):
                call_command(
                    "bootstrap_p0_horse_identity_evidence",
                    "--commit",
                    "--manifest",
                    str(path),
                    "--approved-sha256",
                    sha,
                    "--approved-by",
                    "operator-b",
                    stdout=StringIO(),
                )
            call_command(
                "bootstrap_p0_horse_identity_evidence",
                "--commit",
                "--manifest",
                str(path),
                "--approved-sha256",
                sha,
                "--approved-by",
                "operator-b",
                "--confirm-approved-artifact",
                stdout=StringIO(),
            )
            with self.assertRaises(CommandError):
                call_command(
                    "bootstrap_p0_horse_identity_evidence",
                    "--verify",
                    "--manifest",
                    str(path),
                    "--approved-sha256",
                    sha,
                    stdout=StringIO(),
                )
            output = StringIO()
            call_command(
                "bootstrap_p0_horse_identity_evidence",
                "--verify",
                "--manifest",
                str(path),
                "--approved-sha256",
                sha,
                "--confirm-approved-artifact",
                "--json",
                stdout=output,
            )
            self.assertTrue(json.loads(output.getvalue())["verified"])

    def test_database_uniqueness_prevents_second_same_sha_receipt(self):
        profile = self._profile(1)
        with tempfile.TemporaryDirectory() as tmp:
            path, sha = self._approved_artifact([profile], Path(tmp))
            commit_identity_bootstrap_artifact(
                path, approved_sha256=sha, approved_by="operator-b"
            )
        other_log = OperationLog.objects.create(
            action_type="duplicate_identity_receipt_test",
            target_type="approved_sha256",
            target_id=sha,
        )
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                HorseIdentityEvidenceCommitReceipt.objects.create(
                    approved_sha256=sha,
                    artifact_sha256="f" * 64,
                    approved_by="other",
                    approved_profile_ids=[profile.pk],
                    before_after={},
                    evidence_summary={},
                    result_payload={},
                    operation_log=other_log,
                )

    def test_any_drift_rolls_back_entire_approved_set(self):
        profiles = [self._profile(1), self._profile(2)]
        with tempfile.TemporaryDirectory() as tmp:
            path, sha = self._approved_artifact(profiles, Path(tmp))
            profiles[1].sire_text = "外部写入"
            profiles[1].save(update_fields=["sire_text", "updated_at"])
            with self.assertRaises(P0HorseIdentityBootstrapError):
                commit_identity_bootstrap_artifact(
                    path, approved_sha256=sha, approved_by="operator-b"
                )
        profiles[0].refresh_from_db()
        self.assertEqual(profiles[0].sire_text, "")
        self.assertEqual(HorseIdentityEvidenceCommitReceipt.objects.count(), 0)

    def test_same_values_without_receipt_are_not_replay(self):
        profile = self._profile(1)
        with tempfile.TemporaryDirectory() as tmp:
            path, sha = self._approved_artifact([profile], Path(tmp))
            profile.sire_text = "父1"
            profile.dam_text = "母1"
            profile.birth_date = date(2022, 3, 1)
            profile.save(update_fields=["sire_text", "dam_text", "birth_date", "updated_at"])
            with self.assertRaises(P0HorseIdentityBootstrapError):
                commit_identity_bootstrap_artifact(
                    path, approved_sha256=sha, approved_by="operator-b"
                )

    def test_manual_lock_or_qualification_drift_blocks_commit(self):
        for index, drift_kind in enumerate(
            ("manual_lock", "qualification"), start=1
        ):
            with self.subTest(drift_kind=drift_kind):
                profile = self._profile(index)
                with tempfile.TemporaryDirectory() as tmp:
                    path, sha = self._approved_artifact([profile], Path(tmp))
                    if drift_kind == "manual_lock":
                        profile.manual_lock_flags = {"sire_text": True}
                        profile.save(
                            update_fields=["manual_lock_flags", "updated_at"]
                        )
                    else:
                        runner = profile.p0_sources.get().race_runner
                        runner.horse_number = "99"
                        runner.save(update_fields=["horse_number", "updated_at"])
                    with self.assertRaises(P0HorseIdentityBootstrapError):
                        commit_identity_bootstrap_artifact(
                            path,
                            approved_sha256=sha,
                            approved_by="operator-b",
                        )

    def test_commit_preserves_public_and_completeness_state(self):
        profile = self._profile(1, published=True)
        before = {
            "review_status": profile.review_status,
            "published_at": profile.published_at,
            "completeness_status": profile.completeness_status,
            "p0_source_count": profile.p0_sources.count(),
        }
        with tempfile.TemporaryDirectory() as tmp:
            path, sha = self._approved_artifact([profile], Path(tmp))
            commit_identity_bootstrap_artifact(
                path, approved_sha256=sha, approved_by="operator-b"
            )
        profile.refresh_from_db()
        self.assertEqual(profile.review_status, before["review_status"])
        self.assertEqual(profile.published_at, before["published_at"])
        self.assertEqual(profile.completeness_status, before["completeness_status"])
        self.assertEqual(profile.p0_sources.count(), before["p0_source_count"])


class PhaseOneQualificationSelectionTests(TestCase):
    def _event(
        self,
        index: int,
        *,
        grade: str,
        year: int = 2024,
        provider: str = "jra",
        country_region: str = RacingRegion.JAPAN,
    ) -> RaceEvent:
        if provider == "jra":
            source_url = f"https://www.jra.go.jp/JRADB/accessS.html?race={index}"
        elif provider == "nar":
            source_url = (
                "https://www.keiba.go.jp/KeibaWeb/TodayRaceInfo/"
                f"RaceMarkTable?race={index}"
            )
        else:
            source_url = f"https://www.france-galop.com/race/{index}"
        return RaceEvent.objects.create(
            year=year,
            slug=f"phase-one-{index}-{grade.lower()}",
            original_name=f"重賞{index}",
            chinese_name=f"重赏{index}",
            country_region=country_region,
            racecourse="東京" if provider == "jra" else "大井",
            grade_text=grade,
            normalized_grade=grade,
            surface=RaceEventSurface.TURF,
            local_date=date(year, 5, min(index, 28)),
            status=RaceEventStatus.FINISHED,
            visibility_status=RaceEventVisibility.PUBLISHED,
            data_quality_status=RaceEventDataQuality.COMPLETE,
            source_refs={"official": {"url": source_url, "provider": provider}},
        )

    def _profile(
        self,
        index: int,
        *,
        events: list[RaceEvent],
        horse_number: str = "7",
        training_scope: dict | None = None,
        published: bool = False,
    ) -> HorseProfile:
        term = TermEntry.objects.create(
            term_type=TermType.HORSE,
            source_language=SourceLanguage.JAPANESE,
            source_ja=f"資格馬{index:03d}",
            racing_region=RacingRegion.JAPAN,
            is_active=True,
        )
        profile = HorseProfile.objects.create(
            primary_term=term,
            original_name=term.source_ja,
            racing_region=RacingRegion.JAPAN,
            review_status=(
                HorseProfileStatus.PUBLISHED if published else HorseProfileStatus.DRAFT
            ),
            published_at=timezone.now() if published else None,
            source_refs={"horse_identity_keys": [f"netkeiba:{2023000000 + index}"]},
        )
        for sequence, event in enumerate(events, start=1):
            runner = RaceEventRunner.objects.create(
                event=event,
                external_runner_id=f"runner-{index}-{sequence}",
                sort_order=sequence,
                horse_number=horse_number,
                horse_name=profile.original_name,
                source_refs=event.source_refs,
            )
            provider = "nar" if "keiba.go.jp" in str(event.source_refs) else "jra"
            if "keiba.go.jp" in str(event.source_refs):
                source_url = (
                    "https://www.keiba.go.jp/KeibaWeb/TodayRaceInfo/"
                    f"RaceMarkTable?race={event.pk}"
                )
            elif "jra.go.jp" in str(event.source_refs):
                source_url = (
                    f"https://www.jra.go.jp/JRADB/accessS.html?race={event.pk}"
                )
            else:
                source_url = str(event.source_refs["official"]["url"])
            evidence_payload = {}
            if training_scope is not None:
                evidence_payload["training_scope"] = training_scope
            HorseP0Source.objects.create(
                profile=profile,
                race_event=event,
                race_runner=runner,
                source_type=HorseP0SourceType.MAJOR_RACE_PARTICIPANT,
                status=HorseP0SourceStatus.ACTIVE,
                racing_region=RacingRegion.JAPAN,
                race_grade=event.normalized_grade,
                horse_name=profile.original_name,
                participant_key=f"phase-one:{event.pk}:{horse_number}:{index}",
                source_url=source_url,
                evidence_payload=evidence_payload,
            )
        return profile

    @staticmethod
    def _reviewed_japan_training_scope() -> dict:
        return {
            "status": "confirmed_japan",
            "evidence": [
                {
                    "source": "reviewed_manual",
                    "source_id": "jra-training-fixture",
                    "source_url": "https://www.jra.go.jp/JRADB/accessU.html?fixture=1",
                    "race_date": "2024-05-01",
                    "affiliation": "栗東",
                    "trainer_name": "テスト調教師",
                    "reviewed": True,
                }
            ],
        }

    def test_scope_deduplicates_profile_and_freezes_all_graded_qualifications(self):
        g3 = self._event(1, grade=RaceGrade.G3)
        g1 = self._event(2, grade=RaceGrade.G1)
        listed = self._event(3, grade=RaceGrade.LISTED)
        profile = self._profile(1, events=[g3, g1, listed])

        manifest = select_identity_bootstrap_batch(target_count=1)

        self.assertEqual([row["profile_id"] for row in manifest["horses"]], [profile.pk])
        horse = manifest["horses"][0]
        self.assertEqual(horse["highest_grade"], RaceGrade.G1)
        self.assertEqual(horse["highest_grade_priority"], 1)
        self.assertEqual(horse["graded_start_count"], 2)
        self.assertEqual(
            {item["race_event_id"] for item in horse["qualification"]},
            {g1.pk, g3.pk},
        )
        self.assertEqual(
            {
                (
                    item["official_provider"],
                    item["official_race_url"],
                    item["race_date"],
                    item["racecourse"],
                    item["horse_number"],
                    item["horse_name"],
                )
                for item in horse["qualification"]
            },
            {
                (
                    "jra",
                    f"https://www.jra.go.jp/JRADB/accessS.html?race={g1.pk}",
                    "2024-05-02",
                    "東京",
                    "7",
                    profile.original_name,
                ),
                (
                    "jra",
                    f"https://www.jra.go.jp/JRADB/accessS.html?race={g3.pk}",
                    "2024-05-01",
                    "東京",
                    "7",
                    profile.original_name,
                ),
            },
        )
        self.assertRegex(manifest["config_fingerprint"], r"^[0-9a-f]{64}$")

    def test_scope_excludes_out_of_range_and_stale_overseas_training_evidence(self):
        valid = self._event(9, grade=RaceGrade.G3)
        too_early = self._event(10, grade=RaceGrade.G1, year=1997)
        too_late = self._event(11, grade=RaceGrade.G1, year=2027)
        profile = self._profile(1, events=[valid, too_early, too_late])
        manifest = select_identity_bootstrap_batch(target_count=1)
        self.assertEqual(manifest["horses"][0]["profile_id"], profile.pk)
        self.assertEqual(
            {row["race_event_id"] for row in manifest["horses"][0]["qualification"]},
            {valid.pk},
        )

        overseas = self._event(
            12,
            grade=RaceGrade.G1,
            provider="france_galop",
            country_region=RacingRegion.FRANCE,
        )
        self._profile(
            2,
            events=[overseas],
            training_scope={
                "status": "confirmed_japan",
                "evidence": [
                    {
                        "source": "reviewed_manual",
                        "source_id": "stale-current-membership",
                        "source_url": "https://www.jra.go.jp/JRADB/accessU.html?fixture=stale",
                        "race_date": "2025-05-12",
                        "affiliation": "栗東",
                        "trainer_name": "テスト調教師",
                        "reviewed": True,
                    }
                ],
            },
        )
        manifest = select_identity_bootstrap_batch(target_count=1)
        self.assertEqual(manifest["horses"][0]["profile_id"], profile.pk)
        with self.assertRaises(P0HorseIdentityBootstrapError):
            select_identity_bootstrap_batch(target_count=2)

    def test_complete_second_layer_context_wins_stable_same_grade_priority(self):
        g2_missing = self._event(4, grade=RaceGrade.G2)
        g2_complete = self._event(5, grade=RaceGrade.G2)
        incomplete = self._profile(1, events=[g2_missing], horse_number="")
        complete = self._profile(2, events=[g2_complete], horse_number="9")

        manifest = select_identity_bootstrap_batch(target_count=1)

        self.assertEqual(manifest["horses"][0]["profile_id"], complete.pk)
        self.assertTrue(manifest["horses"][0]["has_complete_official_context"])
        self.assertNotEqual(incomplete.pk, complete.pk)

    def test_training_scope_gate_excludes_foreign_visitor_and_preserves_evidence(self):
        event = self._event(6, grade=RaceGrade.G2)
        foreign = self._profile(
            1,
            events=[event],
            training_scope={
                "status": "foreign_visitor",
                "evidence": [{"source": "reviewed_manual", "reviewed": True}],
            },
        )
        confirmed = self._profile(
            2,
            events=[event],
            horse_number="8",
            training_scope=self._reviewed_japan_training_scope(),
        )
        provisional = self._profile(3, events=[event], horse_number="9")

        manifest = select_identity_bootstrap_batch(target_count=2)
        horses = {row["profile_id"]: row for row in manifest["horses"]}

        self.assertNotIn(foreign.pk, horses)
        self.assertEqual(horses[confirmed.pk]["training_scope_status"], "confirmed_japan")
        self.assertEqual(
            horses[confirmed.pk]["training_evidence"][0]["affiliation"], "栗東"
        )
        self.assertEqual(
            horses[provisional.pk]["training_scope_status"], "provisional_japan"
        )
        self.assertEqual(horses[provisional.pk]["training_evidence"], [])

    def test_selection_is_bounded_prefetched_and_excludes_frozen_old_blockers(self):
        event = self._event(7, grade=RaceGrade.G3, provider="nar")
        profiles = [self._profile(index, events=[event]) for index in range(1, 8)]
        excluded = {profiles[0].pk, profiles[1].pk}

        with CaptureQueriesContext(connection) as queries:
            manifest = select_identity_bootstrap_batch(
                target_count=5,
                excluded_profile_ids=excluded,
                excluded_batch_id="old-39",
                exclusion_reason="previous_batch_blockers",
            )

        self.assertLessEqual(len(queries), 6)
        self.assertEqual(len(manifest["horses"]), 5)
        self.assertFalse(excluded & {row["profile_id"] for row in manifest["horses"]})
        self.assertTrue(
            all(
                row["qualification"][0]["official_provider"] == "nar"
                for row in manifest["horses"]
            )
        )

    def test_qualification_snapshot_drift_fails_closed(self):
        event = self._event(8, grade=RaceGrade.JG1)
        profile = self._profile(1, events=[event], horse_number="3")
        manifest = select_identity_bootstrap_batch(target_count=1)
        runner = profile.p0_sources.get().race_runner
        runner.horse_number = "4"
        runner.save(update_fields=["horse_number", "updated_at"])

        with self.assertRaises(P0HorseIdentityBootstrapError):
            identity_bootstrap.validate_identity_bootstrap_snapshot(manifest)
