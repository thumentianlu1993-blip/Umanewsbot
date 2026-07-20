"""Per-batch human review workbook for rolling P0 horse completion.

Builds one standalone xlsx per batch: a summary sheet, one sheet per region
with one summary row per horse, and an exception sampling sheet collecting
blocked payloads, career gaps, unknown results and identity-enrichment
candidates. The workbook is a reading interface for sampling / spot checks
/ AI-assisted review; the JSONL artifact remains the only commit credential.
"""

from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

from stable.services.p0_horse_completion_batch import P0HorseBatchError

SUMMARY_SHEET = "汇总"
EXCEPTION_SHEET = "异常抽样"
REGION_SHEET_NAMES = {
    "japan": "日本",
    "hong_kong": "中国香港",
    "united_kingdom": "英国",
    "france": "法国",
    "united_states": "美国",
}

HORSE_COLUMNS = [
    ("profile_id", "profile_id"),
    ("candidate_key", "candidate_key"),
    ("horse_name", "马名"),
    ("external_horse_id", "来源马ID"),
    ("basic_profile_complete", "基础资料完整"),
    ("pedigree_complete", "血统完整"),
    ("career_history_complete", "履历完整"),
    ("source_evidence_complete", "来源证据完整"),
    ("official_or_source_start_count", "来源总出赛"),
    ("collected_start_count", "采集实际出赛"),
    ("career_history_gap_count", "生涯缺口"),
    ("unknown_result_count", "未知结果数"),
    ("failure_reason", "失败原因"),
    ("queue_reasons", "队列排序原因"),
    ("source_urls", "来源URL"),
]


def _payload_row(payload: dict[str, Any], queue_reasons: str) -> dict[str, Any]:
    coverage = payload.get("coverage") if isinstance(payload.get("coverage"), dict) else {}
    career = (
        payload.get("career_history")
        if isinstance(payload.get("career_history"), dict)
        else {}
    )
    records = payload.get("race_records") or []
    unknown_results = sum(
        1 for record in records if record.get("result_status") == "unknown"
    )
    evidence = [
        str(item.get("source_url"))
        for item in payload.get("source_evidence") or []
        if isinstance(item, dict) and item.get("source_url")
    ]
    return {
        "candidate_key": payload.get("candidate_key", ""),
        "horse_name": payload.get("horse_name", ""),
        "external_horse_id": payload.get("external_horse_id", ""),
        "basic_profile_complete": bool(coverage.get("basic_profile", {}).get("complete")),
        "pedigree_complete": bool(coverage.get("pedigree", {}).get("complete")),
        "career_history_complete": bool(coverage.get("career_history", {}).get("complete")),
        "source_evidence_complete": bool(coverage.get("source_evidence", {}).get("complete")),
        "official_or_source_start_count": career.get("official_or_source_start_count", ""),
        "collected_start_count": career.get("collected_start_count", ""),
        "career_history_gap_count": career.get("gap_count", ""),
        "unknown_result_count": unknown_results,
        "failure_reason": ";".join(payload.get("failure_reason") or []),
        "queue_reasons": queue_reasons,
        "source_urls": ";".join(evidence),
    }


def _anomaly_flags(row: dict[str, Any]) -> list[str]:
    flags: list[str] = []
    if row["failure_reason"]:
        flags.append("blocked")
    gap = row["career_history_gap_count"]
    if isinstance(gap, int) and gap > 0:
        flags.append("career_gap")
    if row["unknown_result_count"]:
        flags.append("unknown_results")
    if not row["career_history_complete"]:
        flags.append("career_incomplete")
    if not row["basic_profile_complete"]:
        flags.append("basic_incomplete")
    if not row["pedigree_complete"]:
        flags.append("pedigree_incomplete")
    return flags


def build_batch_review_workbook(
    *,
    manifest: dict[str, Any],
    artifact_dir: str | Path,
    output_path: str | Path,
) -> Path:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise P0HorseBatchError(
            "openpyxl is required to build the batch review workbook"
        ) from exc

    combined_path = Path(artifact_dir) / "combined_candidates.jsonl"
    try:
        payloads = [
            json.loads(line)
            for line in combined_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
    except (OSError, ValueError) as exc:
        raise P0HorseBatchError(
            f"batch combined candidates are unreadable: {combined_path}"
        ) from exc

    queue_reasons_by_key = {
        horse["candidate_key"]: ";".join(horse.get("queue_reasons") or [])
        for horse in manifest.get("horses") or []
    }
    profile_id_by_key = {
        horse["candidate_key"]: horse.get("profile_id")
        for horse in manifest.get("horses") or []
    }

    rows_by_region: dict[str, list[dict[str, Any]]] = {}
    exception_rows: list[dict[str, Any]] = []
    for payload in payloads:
        key = payload.get("candidate_key", "")
        row = _payload_row(payload, queue_reasons_by_key.get(key, ""))
        row["profile_id"] = profile_id_by_key.get(key, "")
        region = str(payload.get("region") or "")
        rows_by_region.setdefault(region, []).append(row)
        flags = _anomaly_flags(row)
        if flags:
            exception_rows.append({**row, "region": region, "anomaly_flags": ";".join(flags)})

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = SUMMARY_SHEET
    summary_rows = [
        ("batch_id", manifest.get("batch_id", "")),
        ("batch_sha256", manifest.get("batch_sha256", "")),
        ("status", manifest.get("status", "")),
        ("horses", len(payloads)),
        ("exceptions", len(exception_rows)),
    ]
    for region, rows in sorted(rows_by_region.items()):
        summary_rows.append((f"region:{region}", len(rows)))
    for label, value in summary_rows:
        summary_sheet.append([label, value])

    for region, sheet_name in REGION_SHEET_NAMES.items():
        rows = rows_by_region.get(region)
        if not rows:
            continue
        sheet = workbook.create_sheet(sheet_name)
        sheet.append([label for _, label in HORSE_COLUMNS])
        for row in rows:
            sheet.append([row.get(field, "") for field, _ in HORSE_COLUMNS])

    sheet = workbook.create_sheet(EXCEPTION_SHEET)
    exception_columns = ["region", "anomaly_flags"] + [field for field, _ in HORSE_COLUMNS]
    sheet.append(exception_columns)
    for row in exception_rows:
        sheet.append([row.get(field, "") for field in exception_columns])

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    workbook.save(tmp_path)
    os.replace(tmp_path, path)
    return path
