from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import stat
from collections import Counter, defaultdict
from copy import deepcopy
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qsl, urlsplit

from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from stable.models import (
    HistoricalRaceEventTarget,
    RaceEvent,
    RaceEventAlias,
    RaceSeries,
    RaceSeriesName,
    RaceSeriesRelation,
)
from stable.services.race_event_reconciliation import (
    _canonical_bytes,
    _set_repeatable_read_snapshot,
    _status_is_compatible,
    classify_historical_race_event_targets,
    event_identity,
)
from stable.services.race_series_identity_review import (
    is_identity_pair_do_not_merge,
    race_series_identity,
)


SCHEMA_VERSION = "1.0"
YEAR = 2026
SHEET_ORDER = (
    "审核说明",
    "唯一名称匹配",
    "同名多候选",
    "无名称匹配",
    "未举办",
    "异常清单",
)
DATA_SHEETS = SHEET_ORDER[1:]
EDITABLE_COLUMNS = {"decision", "review_note"}
SUPPORTED_DECISIONS = {
    "defer",
    "merge_and_link",
    "keep_independent",
    "ignore_false_match",
}
PUBLIC_URL_KEYS = {"official", "result", "source_url", "url", "result_url"}
SENSITIVE_QUERY_KEYS = {"token", "key", "secret", "signature"}

ROW_COLUMNS = (
    "decision",
    "review_note",
    "engine_compatible",
    "compatibility_failures",
    "target_chinese_name",
    "target_original_name",
    "event_chinese_name",
    "event_original_name",
    "country_region",
    "year",
    "target_local_date",
    "event_local_date",
    "target_racecourse",
    "event_racecourse",
    "target_grade_text",
    "event_grade_text",
    "target_normalized_grade",
    "event_normalized_grade",
    "target_surface",
    "event_surface",
    "target_distance_text",
    "event_distance_text",
    "event_status",
    "event_visibility_status",
    "target_series_chinese_name",
    "target_series_original_name",
    "source_series_chinese_name",
    "source_series_original_name",
    "classification",
    "reason",
    "review_bucket",
    "public_source_urls",
    "supplemental_suggestions",
    "sequence",
    "target_id",
    "event_id",
    "target_series_id",
    "destination_series_id",
    "source_series_id",
    "target_series_key",
    "source_series_key",
    "target_identity_sha256",
    "event_identity_sha256",
    "source_series_identity_sha256",
    "destination_series_identity_sha256",
    "candidate_event_ids",
    "dependency_checks",
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
EDITABLE_HEADER_FILL = PatternFill("solid", fgColor="C65911")
EDITABLE_CELL_FILL = PatternFill("solid", fgColor="FFF2CC")
INFO_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="B7C9D6")
CELL_BORDER = Border(bottom=THIN_GRAY)
HEADER_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)

COLUMN_WIDTHS = {
    "decision": 24,
    "review_note": 38,
    "engine_compatible": 18,
    "compatibility_failures": 32,
    "target_chinese_name": 24,
    "target_original_name": 30,
    "event_chinese_name": 24,
    "event_original_name": 30,
    "country_region": 16,
    "year": 10,
    "target_local_date": 16,
    "event_local_date": 16,
    "target_racecourse": 22,
    "event_racecourse": 22,
    "target_grade_text": 16,
    "event_grade_text": 16,
    "target_normalized_grade": 20,
    "event_normalized_grade": 20,
    "target_surface": 16,
    "event_surface": 16,
    "target_distance_text": 20,
    "event_distance_text": 20,
    "event_status": 16,
    "event_visibility_status": 20,
    "target_series_chinese_name": 26,
    "target_series_original_name": 32,
    "source_series_chinese_name": 26,
    "source_series_original_name": 32,
    "classification": 22,
    "reason": 30,
    "review_bucket": 24,
    "public_source_urls": 42,
    "supplemental_suggestions": 38,
    "sequence": 12,
    "target_id": 14,
    "event_id": 14,
    "target_series_id": 18,
    "destination_series_id": 20,
    "source_series_id": 18,
    "target_series_key": 28,
    "source_series_key": 28,
    "target_identity_sha256": 24,
    "event_identity_sha256": 24,
    "source_series_identity_sha256": 24,
    "destination_series_identity_sha256": 24,
    "candidate_event_ids": 24,
    "dependency_checks": 42,
}


class RaceSeriesIdentity2026ReviewError(ValueError):
    pass


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _safe_read(path: Path, *, label: str) -> bytes:
    flags = os.O_RDONLY
    if hasattr(os, "O_NOFOLLOW"):
        flags |= os.O_NOFOLLOW
    try:
        descriptor = os.open(path, flags)
    except OSError as exc:
        raise RaceSeriesIdentity2026ReviewError(f"cannot safely read {label}: {path}") from exc
    try:
        metadata = os.fstat(descriptor)
        if not stat.S_ISREG(metadata.st_mode):
            raise RaceSeriesIdentity2026ReviewError(f"{label} must be a regular file")
        chunks: list[bytes] = []
        while chunk := os.read(descriptor, 1024 * 1024):
            chunks.append(chunk)
        return b"".join(chunks)
    finally:
        os.close(descriptor)


def _json_object(payload: bytes, *, label: str) -> dict[str, Any]:
    try:
        value = json.loads(payload)
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise RaceSeriesIdentity2026ReviewError(f"invalid JSON: {label}") from exc
    if not isinstance(value, dict):
        raise RaceSeriesIdentity2026ReviewError(f"{label} must be a JSON object")
    return value


def _normalized(value: Any) -> str:
    return " ".join(str(value or "").casefold().split())


def _public_url(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    value = value.strip()
    try:
        parsed = urlsplit(value)
    except ValueError:
        return None
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return None
    if parsed.username or parsed.password:
        return None
    if any(
        any(marker in key.casefold() for marker in SENSITIVE_QUERY_KEYS)
        for key, _ in parse_qsl(parsed.query)
    ):
        return None
    return value


def _urls_from_refs(*refs_values: Any) -> list[str]:
    urls: set[str] = set()
    for refs in refs_values:
        if not isinstance(refs, dict):
            continue
        for key, value in refs.items():
            if str(key).casefold() not in PUBLIC_URL_KEYS:
                continue
            values = value if isinstance(value, list) else [value]
            for item in values:
                url = _public_url(item)
                if url:
                    urls.add(url)
    return sorted(urls)


def _identity_sha(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    sha = value.get("sha256")
    return str(sha or "")


def _base_row(classification: dict[str, Any], *, bucket: str) -> dict[str, Any]:
    candidate_ids = sorted({int(value) for value in classification.get("candidate_event_ids") or []})
    event_id = classification.get("event_id")
    if event_id is None and len(candidate_ids) == 1:
        event_id = candidate_ids[0]
    target_series_id = int(classification.get("series_id") or 0)
    candidate_identity = classification.get("candidate_event_identity") or {}
    candidate_payload = candidate_identity.get("payload") if isinstance(candidate_identity, dict) else {}
    source_series_id = (
        int(candidate_payload.get("race_series_id") or 0)
        if isinstance(candidate_payload, dict)
        else 0
    )
    return {
        "sequence": 0,
        "target_id": int(classification["target_id"]),
        "target_identity_sha256": _identity_sha(classification.get("target_identity")),
        "target_series_id": target_series_id,
        "target_series_key": str(classification.get("series_key") or ""),
        "target_series_original_name": str(
            classification.get("target_series_original_name") or ""
        ),
        "target_series_chinese_name": str(
            classification.get("target_series_chinese_name") or ""
        ),
        "destination_series_id": target_series_id,
        "source_series_id": source_series_id,
        "source_series_key": str(
            candidate_payload.get("series_key") or ""
            if isinstance(candidate_payload, dict)
            else ""
        ),
        "source_series_original_name": str(
            classification.get("source_series_original_name") or ""
        ),
        "source_series_chinese_name": str(
            classification.get("source_series_chinese_name") or ""
        ),
        "event_id": int(event_id) if event_id else None,
        "event_identity_sha256": _identity_sha(candidate_identity),
        "source_series_identity_sha256": str(
            classification.get("source_series_identity_sha256") or ""
        ),
        "destination_series_identity_sha256": str(
            classification.get("destination_series_identity_sha256") or ""
        ),
        "candidate_event_ids": candidate_ids,
        "country_region": str(classification.get("country_region") or ""),
        "year": int(classification.get("year") or YEAR),
        "target_original_name": str(classification.get("target_original_name") or ""),
        "target_chinese_name": str(classification.get("target_chinese_name") or ""),
        "event_original_name": str(
            candidate_payload.get("original_name") or ""
            if isinstance(candidate_payload, dict)
            else ""
        ),
        "event_chinese_name": str(
            candidate_payload.get("chinese_name") or ""
            if isinstance(candidate_payload, dict)
            else ""
        ),
        "target_local_date": str(classification.get("local_date") or ""),
        "event_local_date": str(
            candidate_payload.get("local_date") or ""
            if isinstance(candidate_payload, dict)
            else ""
        ),
        "target_racecourse": str(classification.get("racecourse") or ""),
        "event_racecourse": str(
            candidate_payload.get("racecourse") or ""
            if isinstance(candidate_payload, dict)
            else ""
        ),
        "target_grade_text": str(classification.get("grade_text") or ""),
        "event_grade_text": str(
            candidate_payload.get("grade_text") or ""
            if isinstance(candidate_payload, dict)
            else ""
        ),
        "target_normalized_grade": str(classification.get("normalized_grade") or ""),
        "event_normalized_grade": str(
            candidate_payload.get("normalized_grade") or ""
            if isinstance(candidate_payload, dict)
            else ""
        ),
        "target_surface": str(classification.get("surface") or ""),
        "event_surface": str(
            candidate_payload.get("surface") or ""
            if isinstance(candidate_payload, dict)
            else ""
        ),
        "target_distance_text": str(classification.get("distance_text") or ""),
        "event_distance_text": str(
            candidate_payload.get("distance_text") or ""
            if isinstance(candidate_payload, dict)
            else ""
        ),
        "event_status": str(
            candidate_payload.get("status") or ""
            if isinstance(candidate_payload, dict)
            else ""
        ),
        "event_visibility_status": str(
            candidate_payload.get("visibility_status") or ""
            if isinstance(candidate_payload, dict)
            else ""
        ),
        "classification": str(classification.get("classification") or ""),
        "reason": str(classification.get("reason") or ""),
        "review_bucket": bucket,
        "engine_compatible": False,
        "compatibility_failures": [],
        "dependency_checks": {},
        "supplemental_suggestions": [],
        "public_source_urls": sorted(
            filter(None, (_public_url(url) for url in classification.get("public_source_urls") or []))
        ),
        "decision": "defer",
        "review_note": "",
    }


def _compatibility_failures(facts: dict[str, Any] | None) -> list[str]:
    if not facts:
        return ["missing_dependency_facts"]
    failures: list[str] = []
    event_id = facts.get("event_id")
    annual_ids = list(facts.get("source_annual_event_ids") or [])
    if annual_ids != [event_id]:
        failures.append("source_has_other_events")
    if facts.get("source_target_ids"):
        failures.append("source_has_targets")
    if facts.get("source_name_ids"):
        failures.append("source_has_names")
    if facts.get("source_relation_ids"):
        failures.append("source_has_relations")
    if facts.get("destination_year_event_ids"):
        failures.append("destination_has_year_event")
    if facts.get("event_owner_target_id") not in (None, facts.get("target_id")):
        failures.append("event_already_owned")
    if facts.get("do_not_merge"):
        failures.append("do_not_merge")
    for key in ("region_matches", "year_matches", "status_compatible", "detail_consistent"):
        if facts.get(key) is not True:
            failures.append(key)
    return failures


def build_review_snapshot(
    *,
    classifications: Iterable[dict[str, Any]],
    alias_suggestions_by_target: dict[int, list[dict[str, Any]]],
    dependency_facts: dict[int, dict[str, Any]],
) -> dict[str, Any]:
    sheets: dict[str, list[dict[str, Any]]] = {name: [] for name in DATA_SHEETS}
    all_rows: list[dict[str, Any]] = []
    counts: Counter[str] = Counter()
    seen: set[int] = set()

    bucket_map = {
        ("identity_conflict", "series_mismatch"): ("unique_series_mismatch", "唯一名称匹配"),
        ("identity_conflict", "ambiguous_name_match"): ("ambiguous_name_match", "同名多候选"),
        ("missing_event", "no_series_year_event"): ("no_name_match", "无名称匹配"),
        ("status_conflict", "not_held_target"): ("not_held", "未举办"),
    }
    ordered = sorted(classifications, key=lambda row: int(row["target_id"]))
    for raw in ordered:
        target_id = int(raw["target_id"])
        if target_id in seen:
            raise RaceSeriesIdentity2026ReviewError(f"duplicate target_id: {target_id}")
        seen.add(target_id)
        classification = str(raw.get("classification") or "")
        reason = str(raw.get("reason") or "")
        if classification == "already_linked" and not reason:
            bucket = "already_linked"
            sheet_name = None
        else:
            bucket, sheet_name = bucket_map.get(
                (classification, reason), ("anomaly", "异常清单")
            )
        row = _base_row(raw, bucket=bucket)
        suggestions = (
            alias_suggestions_by_target.get(target_id, [])
            if bucket in {"ambiguous_name_match", "no_name_match"}
            else []
        )
        row["supplemental_suggestions"] = [
            {
                key: suggestion[key]
                for key in ("event_id", "event_series_id", "country_region", "source")
                if key in suggestion
            }
            for suggestion in suggestions
            if isinstance(suggestion, dict)
        ]
        if bucket == "unique_series_mismatch":
            facts = deepcopy(dependency_facts.get(target_id) or {})
            facts.setdefault("target_id", target_id)
            row["source_series_id"] = int(facts.get("source_series_id") or row["source_series_id"] or 0)
            row["destination_series_id"] = int(
                facts.get("destination_series_id") or row["destination_series_id"] or 0
            )
            row["event_id"] = int(facts.get("event_id") or row["event_id"] or 0) or None
            row["source_series_identity_sha256"] = str(
                facts.get("source_series_identity_sha256")
                or row["source_series_identity_sha256"]
                or ""
            )
            row["destination_series_identity_sha256"] = str(
                facts.get("destination_series_identity_sha256")
                or row["destination_series_identity_sha256"]
                or ""
            )
            row["dependency_checks"] = {
                "source_single_event": list(facts.get("source_annual_event_ids") or [])
                == [facts.get("event_id")],
                "source_has_targets": bool(facts.get("source_target_ids")),
                "source_has_names": bool(facts.get("source_name_ids")),
                "source_has_relations": bool(facts.get("source_relation_ids")),
                "destination_has_year_event": bool(
                    facts.get("destination_year_event_ids")
                ),
                "event_already_owned": facts.get("event_owner_target_id")
                not in (None, facts.get("target_id")),
                "do_not_merge": bool(facts.get("do_not_merge")),
                "region_matches": facts.get("region_matches") is True,
                "year_matches": facts.get("year_matches") is True,
                "status_compatible": facts.get("status_compatible") is True,
                "detail_consistent": facts.get("detail_consistent") is True,
            }
            row["compatibility_failures"] = _compatibility_failures(facts)
        counts[bucket] += 1
        all_rows.append(row)
        if sheet_name:
            sheets[sheet_name].append(row)

    unique_rows = sheets["唯一名称匹配"]
    for fact_key, failure in (
        ("source_series_id", "duplicate_source_series"),
        ("destination_series_id", "duplicate_destination_series"),
        ("event_id", "duplicate_event"),
        ("target_id", "duplicate_target"),
    ):
        value_counts = Counter(row[fact_key] for row in unique_rows if row.get(fact_key))
        duplicates = {value for value, count in value_counts.items() if count > 1}
        for row in unique_rows:
            if row.get(fact_key) in duplicates:
                row["compatibility_failures"].append(failure)
    for sheet_name in DATA_SHEETS:
        for sequence, row in enumerate(sheets[sheet_name], start=1):
            row["sequence"] = sequence
            row["compatibility_failures"] = sorted(set(row["compatibility_failures"]))
            row["engine_compatible"] = (
                row["review_bucket"] == "unique_series_mismatch"
                and not row["compatibility_failures"]
            )

    counts["total_targets"] = len(all_rows)
    return {
        "schema_version": SCHEMA_VERSION,
        "year": YEAR,
        "counts": dict(sorted(counts.items())),
        "all_rows": all_rows,
        "sheets": sheets,
        "anomalies": sheets["异常清单"],
        "blocks_decisions": bool(sheets["异常清单"]),
    }


def _cell_value(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return value


def _write_workbook(snapshot: dict[str, Any], path: Path) -> None:
    workbook = Workbook()
    info = workbook.active
    info.title = "审核说明"
    info.sheet_view.zoomScale = 100
    info.append(["字段", "说明"])
    info.append(["规则", "仅唯一名称匹配表的 decision 与 review_note 可编辑"])
    info.append(["允许动作", "defer / merge_and_link / keep_independent / ignore_false_match"])
    info.freeze_panes = "A2"
    info.row_dimensions[1].height = 28
    info.column_dimensions["A"].width = 20
    info.column_dimensions["B"].width = 88
    for cell in info[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in info.iter_rows(min_row=2):
        row[0].fill = INFO_FILL
        row[0].font = Font(bold=True, color="1F1F1F")
        for cell in row:
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    info.row_dimensions[2].height = 34
    info.row_dimensions[3].height = 34
    for sheet_name in DATA_SHEETS:
        sheet = workbook.create_sheet(sheet_name)
        sheet.sheet_view.zoomScale = 85
        sheet.append(list(ROW_COLUMNS))
        for row in snapshot["sheets"][sheet_name]:
            sheet.append([_cell_value(row.get(column)) for column in ROW_COLUMNS])
        sheet.freeze_panes = "C2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 42
        for index, column in enumerate(ROW_COLUMNS, start=1):
            letter = get_column_letter(index)
            sheet.column_dimensions[letter].width = COLUMN_WIDTHS[column]
            header = sheet.cell(1, index)
            header.fill = (
                EDITABLE_HEADER_FILL
                if sheet_name == "唯一名称匹配" and column in EDITABLE_COLUMNS
                else HEADER_FILL
            )
            header.font = HEADER_FONT
            header.border = HEADER_BORDER
            header.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        for row in sheet.iter_rows(min_row=2):
            sheet.row_dimensions[row[0].row].height = 34
            for cell in row:
                cell.border = CELL_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if sheet_name == "唯一名称匹配":
                row[0].fill = EDITABLE_CELL_FILL
                row[1].fill = EDITABLE_CELL_FILL
        if sheet_name == "唯一名称匹配" and sheet.max_row >= 2:
            validation = DataValidation(
                type="list",
                formula1='"defer,merge_and_link,keep_independent,ignore_false_match"',
                allow_blank=False,
                error="请选择允许的审核动作。",
                errorTitle="无效审核动作",
                prompt="请选择 defer、merge_and_link、keep_independent 或 ignore_false_match。",
                promptTitle="审核动作",
                showErrorMessage=True,
                showInputMessage=True,
            )
            sheet.add_data_validation(validation)
            validation.add(f"A2:A{sheet.max_row}")
    workbook.save(path)


def _flat_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet_name in DATA_SHEETS:
        for row in snapshot["sheets"][sheet_name]:
            rows.append({"sheet": sheet_name, **row})
    return rows


def write_review_package(
    *, snapshot: dict[str, Any], output_dir: Path | str, production_head: str, as_of: str
) -> dict[str, Any]:
    output = Path(output_dir)
    if output.exists():
        raise RaceSeriesIdentity2026ReviewError(f"output directory already exists: {output}")
    output.mkdir(parents=True)
    try:
        snapshot_payload = {
            "schema_version": SCHEMA_VERSION,
            "as_of": as_of,
            "production_head": production_head,
            **snapshot,
        }
        review_payload = {
            "schema_version": SCHEMA_VERSION,
            "year": snapshot["year"],
            "counts": snapshot["counts"],
            "sheets": snapshot["sheets"],
            "blocks_decisions": snapshot["blocks_decisions"],
        }
        (output / "snapshot.json").write_bytes(_canonical_bytes(snapshot_payload))
        (output / "review.json").write_bytes(_canonical_bytes(review_payload))
        with (output / "review.csv").open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=("sheet", *ROW_COLUMNS))
            writer.writeheader()
            for row in _flat_rows(snapshot):
                writer.writerow({key: _cell_value(row.get(key)) for key in writer.fieldnames})
        _write_workbook(snapshot, output / "review.xlsx")
        artifact_names = ("snapshot.json", "review.json", "review.csv", "review.xlsx")
        artifacts = {}
        for name in artifact_names:
            payload = _safe_read(output / name, label=name)
            artifacts[name] = {"path": name, "size": len(payload), "sha256": _sha256_bytes(payload)}
        manifest = {
            "schema_version": SCHEMA_VERSION,
            "generator": "race_series_identity_2026_review",
            "as_of": as_of,
            "production_head": production_head,
            "year": snapshot["year"],
            "counts": snapshot["counts"],
            "blocks_decisions": snapshot["blocks_decisions"],
            "canonical_rows_sha256": _sha256_bytes(_canonical_bytes(_flat_rows(snapshot))),
            "artifacts": artifacts,
        }
        manifest_bytes = _canonical_bytes(manifest)
        (output / "manifest.json").write_bytes(manifest_bytes)
        return {
            "output_dir": str(output),
            "manifest_sha256": _sha256_bytes(manifest_bytes),
            "counts": snapshot["counts"],
            "blocks_decisions": snapshot["blocks_decisions"],
        }
    except Exception:
        for path in sorted(output.glob("*")):
            if path.is_file() or path.is_symlink():
                path.unlink()
        output.rmdir()
        raise


def _verified_package(root: Path, expected_manifest_sha256: str) -> tuple[dict[str, Any], bytes]:
    expected_names = {"snapshot.json", "review.json", "review.csv", "review.xlsx", "manifest.json"}
    try:
        if root.is_symlink() or not root.is_dir():
            raise RaceSeriesIdentity2026ReviewError("original package must be a real directory")
        actual_names = {path.name for path in root.iterdir()}
    except OSError as exc:
        raise RaceSeriesIdentity2026ReviewError("cannot inspect original package") from exc
    if actual_names != expected_names:
        raise RaceSeriesIdentity2026ReviewError("original package file set mismatch")
    manifest_bytes = _safe_read(root / "manifest.json", label="manifest")
    if _sha256_bytes(manifest_bytes) != expected_manifest_sha256:
        raise RaceSeriesIdentity2026ReviewError("manifest SHA-256 mismatch")
    manifest = _json_object(manifest_bytes, label="manifest")
    if manifest.get("schema_version") != SCHEMA_VERSION:
        raise RaceSeriesIdentity2026ReviewError("unsupported manifest schema")
    artifacts = manifest.get("artifacts")
    if not isinstance(artifacts, dict) or set(artifacts) != expected_names - {"manifest.json"}:
        raise RaceSeriesIdentity2026ReviewError("manifest artifact set mismatch")
    for name, identity in artifacts.items():
        payload = _safe_read(root / name, label=name)
        if (
            not isinstance(identity, dict)
            or identity.get("path") != name
            or identity.get("size") != len(payload)
            or identity.get("sha256") != _sha256_bytes(payload)
        ):
            raise RaceSeriesIdentity2026ReviewError(f"manifest artifact mismatch: {name}")
    return manifest, _safe_read(root / "review.xlsx", label="original workbook")


def _workbook_rows(payload: bytes, *, label: str) -> dict[str, list[dict[str, Any]]]:
    try:
        workbook = load_workbook(io.BytesIO(payload), data_only=False, read_only=False)
    except Exception as exc:
        raise RaceSeriesIdentity2026ReviewError(f"invalid {label}") from exc
    if workbook.sheetnames != list(SHEET_ORDER):
        raise RaceSeriesIdentity2026ReviewError(f"{label} sheet set/order mismatch")
    rows_by_sheet: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in SHEET_ORDER:
        sheet = workbook[sheet_name]
        if sheet.sheet_state != "visible":
            raise RaceSeriesIdentity2026ReviewError(f"{label} contains hidden sheet")
        if any(dimension.hidden for dimension in sheet.row_dimensions.values()):
            raise RaceSeriesIdentity2026ReviewError(f"{label} contains hidden rows")
        if any(dimension.hidden for dimension in sheet.column_dimensions.values()):
            raise RaceSeriesIdentity2026ReviewError(f"{label} contains hidden columns")
        for row in sheet.iter_rows():
            if any(cell.data_type == "f" for cell in row):
                raise RaceSeriesIdentity2026ReviewError(f"{label} contains formula cells")
        if sheet_name == "审核说明":
            rows_by_sheet[sheet_name] = [
                {str(row[0].value): row[1].value if len(row) > 1 else None}
                for row in sheet.iter_rows(min_row=2)
                if row and row[0].value is not None
            ]
            continue
        headers = [cell.value for cell in sheet[1]]
        if headers != list(ROW_COLUMNS):
            raise RaceSeriesIdentity2026ReviewError(f"{label} header mismatch: {sheet_name}")
        data_rows: list[dict[str, Any]] = []
        for cells in sheet.iter_rows(min_row=2):
            if all(cell.value is None for cell in cells):
                continue
            data_rows.append(dict(zip(headers, (cell.value for cell in cells), strict=True)))
        rows_by_sheet[sheet_name] = data_rows
    return rows_by_sheet


def _normalized_cell(value: Any) -> Any:
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return value
        if isinstance(parsed, (list, dict)):
            return parsed
    return value


def build_decisions_from_reviewed_workbook(
    *,
    original_package_dir: Path | str,
    expected_manifest_sha256: str,
    reviewed_workbook: Path | str,
    expected_workbook_sha256: str,
) -> dict[str, Any]:
    root = Path(original_package_dir)
    manifest, original_bytes = _verified_package(root, expected_manifest_sha256)
    if manifest.get("blocks_decisions"):
        raise RaceSeriesIdentity2026ReviewError("manifest contains blocking anomalies")
    reviewed_bytes = _safe_read(Path(reviewed_workbook), label="reviewed workbook")
    if _sha256_bytes(reviewed_bytes) != expected_workbook_sha256:
        raise RaceSeriesIdentity2026ReviewError("reviewed workbook SHA-256 mismatch")
    original = _workbook_rows(original_bytes, label="original workbook")
    reviewed = _workbook_rows(reviewed_bytes, label="reviewed workbook")
    if original["审核说明"] != reviewed["审核说明"]:
        raise RaceSeriesIdentity2026ReviewError("reviewed workbook instructions changed")

    decisions: list[dict[str, Any]] = []
    for sheet_name in DATA_SHEETS:
        original_rows = original[sheet_name]
        reviewed_rows = reviewed[sheet_name]
        if len(original_rows) != len(reviewed_rows):
            raise RaceSeriesIdentity2026ReviewError(f"row set mismatch: {sheet_name}")
        for index, (before, after) in enumerate(zip(original_rows, reviewed_rows, strict=True), start=1):
            for column in ROW_COLUMNS:
                if column in EDITABLE_COLUMNS and sheet_name == "唯一名称匹配":
                    continue
                if _normalized_cell(before[column]) != _normalized_cell(after[column]):
                    raise RaceSeriesIdentity2026ReviewError(
                        f"machine/read-only column changed: {sheet_name}!{column}:{index}"
                    )
            decision = str(after.get("decision") or "defer").strip()
            note = str(after.get("review_note") or "").strip()
            if decision not in SUPPORTED_DECISIONS:
                raise RaceSeriesIdentity2026ReviewError(f"unsupported decision: {decision}")
            if decision == "defer":
                continue
            if sheet_name != "唯一名称匹配":
                raise RaceSeriesIdentity2026ReviewError("only unique match sheet can produce decisions")
            if not note:
                raise RaceSeriesIdentity2026ReviewError("non-defer decision requires review_note")
            if decision == "merge_and_link" and before.get("engine_compatible") is not True:
                raise RaceSeriesIdentity2026ReviewError("decision is not engine compatible")
            urls = _normalized_cell(before.get("public_source_urls"))
            if not isinstance(urls, list) or not urls:
                raise RaceSeriesIdentity2026ReviewError("non-defer decision requires public source URL")
            target_id = int(before["target_id"])
            target_identity_sha256 = str(before.get("target_identity_sha256") or "")
            event_identity_sha256 = str(before.get("event_identity_sha256") or "")
            source_series_identity_sha256 = str(
                before.get("source_series_identity_sha256") or ""
            )
            destination_series_identity_sha256 = str(
                before.get("destination_series_identity_sha256") or ""
            )
            if (
                re.fullmatch(r"[0-9a-f]{64}", target_identity_sha256) is None
                or re.fullmatch(r"[0-9a-f]{64}", event_identity_sha256) is None
                or re.fullmatch(r"[0-9a-f]{64}", source_series_identity_sha256)
                is None
                or re.fullmatch(
                    r"[0-9a-f]{64}", destination_series_identity_sha256
                )
                is None
            ):
                raise RaceSeriesIdentity2026ReviewError(
                    "non-defer decision requires locked target/event/series identity SHA"
                )
            decisions.append(
                {
                    "decision_id": f"2026-series-identity:{target_id}",
                    "sheet": "2026-unique-series-mismatch",
                    "sequence": int(before["sequence"]),
                    "decision": decision,
                    "target_id": target_id,
                    "target_series_id": int(before["target_series_id"]),
                    "event_id": int(before["event_id"]),
                    "event_series_id": int(before["source_series_id"]),
                    "year": int(before["year"]),
                    "country_region": str(before["country_region"]),
                    "confidence": "high",
                    "target_identity_sha256": target_identity_sha256,
                    "event_identity_sha256": event_identity_sha256,
                    "source_series_identity_sha256": source_series_identity_sha256,
                    "destination_series_identity_sha256": (
                        destination_series_identity_sha256
                    ),
                    "evidence": {"summary": note, "source_urls": sorted(set(urls))},
                }
            )
    return {
        "decisions": {
            "schema_version": SCHEMA_VERSION,
            "source": "review_2026_race_series_identities",
            "source_sha256": expected_manifest_sha256,
            "decisions": decisions,
        },
        "field_repairs": {"schema_version": SCHEMA_VERSION, "repairs": []},
        "decision_count": len(decisions),
        "reviewed_workbook_sha256": expected_workbook_sha256,
    }


def _orm_supplemental_suggestions(
    targets: list[HistoricalRaceEventTarget], events: list[RaceEvent]
) -> dict[int, list[dict[str, Any]]]:
    target_series_ids = {target.race_series_id for target in targets}
    names = list(
        RaceSeriesName.objects.filter(series_id__in=target_series_ids, is_active=True)
        .order_by("id")
        .values("series_id", "text")
    )
    aliases = list(
        RaceEventAlias.objects.filter(event__in=events, is_active=True)
        .order_by("id")
        .values("event_id", "text")
    )
    target_aliases: dict[int, set[str]] = defaultdict(set)
    targets_by_series: dict[int, list[int]] = defaultdict(list)
    for target in targets:
        targets_by_series[target.race_series_id].append(target.pk)
        target_aliases[target.pk].update(
            filter(
                None,
                (
                    _normalized(target.original_name),
                    _normalized(target.chinese_name),
                    _normalized(target.race_series.canonical_name_original),
                    _normalized(target.race_series.chinese_name),
                ),
            )
        )
    for name in names:
        for target_id in targets_by_series[name["series_id"]]:
            target_aliases[target_id].add(_normalized(name["text"]))
    targets_by_name: dict[str, set[int]] = defaultdict(set)
    for target_id, values in target_aliases.items():
        for value in values:
            targets_by_name[value].add(target_id)
    targets_by_id = {target.pk: target for target in targets}
    events_by_id = {event.pk: event for event in events}
    result: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for alias in aliases:
        event = events_by_id[alias["event_id"]]
        alias_text = _normalized(alias["text"])
        for target_id in targets_by_name.get(alias_text, set()):
            target = targets_by_id[target_id]
            if event.year == target.year:
                result[target_id].append(
                    {
                        "event_id": event.pk,
                        "event_series_id": event.race_series_id,
                        "country_region": event.country_region,
                        "source": "series_name_x_event_alias",
                    }
                )
    return {key: sorted(value, key=lambda row: row["event_id"]) for key, value in result.items()}


def export_2026_review_snapshot(*, year: int = YEAR) -> dict[str, Any]:
    with transaction.atomic():
        _set_repeatable_read_snapshot()
        targets = list(
            HistoricalRaceEventTarget.objects.filter(year=year)
            .select_related("race_series", "event", "event__race_series")
            .order_by("id")
        )
        classifications = classify_historical_race_event_targets(targets)
        events = list(RaceEvent.objects.filter(year=year).select_related("race_series").order_by("id"))
        events_by_id = {event.pk: event for event in events}
        targets_by_event = dict(
            HistoricalRaceEventTarget.objects.exclude(event_id=None).values_list("event_id", "id")
        )
        all_series_ids = {target.race_series_id for target in targets}
        all_series_ids.update(event.race_series_id for event in events if event.race_series_id)
        series = {row.pk: row for row in RaceSeries.objects.filter(pk__in=all_series_ids)}
        annual_by_series: dict[int, list[int]] = defaultdict(list)
        for event_id, series_id in RaceEvent.objects.filter(race_series_id__in=all_series_ids).values_list(
            "id", "race_series_id"
        ):
            annual_by_series[series_id].append(event_id)
        target_ids_by_series: dict[int, list[int]] = defaultdict(list)
        for target_id, series_id in HistoricalRaceEventTarget.objects.filter(
            race_series_id__in=all_series_ids
        ).values_list("id", "race_series_id"):
            target_ids_by_series[series_id].append(target_id)
        names_by_series: dict[int, list[int]] = defaultdict(list)
        for name_id, series_id in RaceSeriesName.objects.filter(series_id__in=all_series_ids).values_list(
            "id", "series_id"
        ):
            names_by_series[series_id].append(name_id)
        relations_by_series: dict[int, list[int]] = defaultdict(list)
        for relation_id, from_id, to_id in RaceSeriesRelation.objects.filter(
            Q(from_series_id__in=all_series_ids) | Q(to_series_id__in=all_series_ids)
        ).values_list("id", "from_series_id", "to_series_id"):
            relations_by_series[from_id].append(relation_id)
            relations_by_series[to_id].append(relation_id)
        destination_year: dict[int, list[int]] = defaultdict(list)
        for event in events:
            if event.race_series_id:
                destination_year[event.race_series_id].append(event.pk)

        dependencies: dict[int, dict[str, Any]] = {}
        for classification, target in zip(classifications, targets, strict=True):
            classification["target_original_name"] = target.original_name
            classification["target_chinese_name"] = target.chinese_name
            classification["target_series_original_name"] = (
                target.race_series.canonical_name_original
            )
            classification["target_series_chinese_name"] = target.race_series.chinese_name
            classification["local_date"] = target.local_date
            classification["racecourse"] = target.racecourse
            classification["grade_text"] = target.grade_text
            classification["normalized_grade"] = target.normalized_grade
            classification["surface"] = target.surface
            classification["distance_text"] = target.distance_text
            candidate_ids = classification.get("candidate_event_ids") or []
            candidate = events_by_id.get(candidate_ids[0]) if len(candidate_ids) == 1 else None
            if candidate:
                classification["candidate_event_identity"] = event_identity(candidate)
                classification["source_series_original_name"] = (
                    candidate.race_series.canonical_name_original
                    if candidate.race_series
                    else ""
                )
                classification["source_series_chinese_name"] = (
                    candidate.race_series.chinese_name if candidate.race_series else ""
                )
            classification["public_source_urls"] = _urls_from_refs(
                target.source_refs,
                target.race_series.source_refs,
                candidate.source_refs if candidate else None,
                candidate.race_series.source_refs if candidate and candidate.race_series else None,
            )
            if classification.get("reason") != "series_mismatch" or not candidate or not candidate.race_series_id:
                continue
            source_id = candidate.race_series_id
            destination_id = target.race_series_id
            source = series[source_id]
            destination = series[destination_id]
            dependencies[target.pk] = {
                "target_id": target.pk,
                "source_series_id": source_id,
                "destination_series_id": destination_id,
                "event_id": candidate.pk,
                "source_annual_event_ids": sorted(annual_by_series[source_id]),
                "source_target_ids": sorted(target_ids_by_series[source_id]),
                "source_name_ids": sorted(names_by_series[source_id]),
                "source_relation_ids": sorted(set(relations_by_series[source_id])),
                "destination_year_event_ids": sorted(destination_year[destination_id]),
                "event_owner_target_id": targets_by_event.get(candidate.pk),
                "do_not_merge": is_identity_pair_do_not_merge(source, destination),
                "region_matches": (
                    source.country_region
                    == destination.country_region
                    == candidate.country_region
                    == target.country_region
                ),
                "year_matches": candidate.year == target.year == year,
                "status_compatible": _status_is_compatible(target, candidate),
                "detail_consistent": all(
                    not left or not right or _normalized(left) == _normalized(right)
                    for left, right in (
                        (target.local_date, candidate.local_date),
                        (target.racecourse, candidate.racecourse),
                        (
                            target.normalized_grade or target.grade_text,
                            candidate.normalized_grade or candidate.grade_text,
                        ),
                        (target.surface, candidate.surface),
                        (target.distance_text, candidate.distance_text),
                    )
                ),
                "source_series_identity_sha256": race_series_identity(source)[
                    "sha256"
                ],
                "destination_series_identity_sha256": race_series_identity(
                    destination
                )["sha256"],
            }
        snapshot = build_review_snapshot(
            classifications=classifications,
            alias_suggestions_by_target=_orm_supplemental_suggestions(targets, events),
            dependency_facts=dependencies,
        )
        snapshot["year"] = year
        snapshot["as_of"] = timezone.now().isoformat()
        expected = {
            "total_targets": 1085,
            "already_linked": 684,
            "unique_series_mismatch": 226,
            "ambiguous_name_match": 11,
            "no_name_match": 162,
            "not_held": 2,
        }
        snapshot["exploration_baseline"] = expected
        snapshot["baseline_drift"] = {
            key: {"expected": value, "actual": snapshot["counts"].get(key, 0)}
            for key, value in expected.items()
            if snapshot["counts"].get(key, 0) != value
        }
        if snapshot["baseline_drift"]:
            snapshot["blocks_decisions"] = True
        return snapshot
