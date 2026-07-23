"""Tests for rolling P0 horse completion batch selection and manifest gates."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

from django.test import TestCase, override_settings

from stable.models import (
    HorseP0Source,
    HorseP0SourceStatus,
    HorseP0SourceType,
    HorseProfile,
    HorseProfileCompleteness,
    HorseProfileCompletionRun,
    RacingRegion,
    SourceLanguage,
    TermEntry,
    TermType,
)
from stable.services.p0_horse_completion_batch import (
    P0HorseBatchError,
    approve_batch_manifest,
    load_batch_manifest,
    select_p0_horse_batch,
    validate_approved_batch_manifest,
    write_batch_manifest,
)


class P0HorseBatchTestBase(TestCase):
    def _term(self, name: str, region: str = RacingRegion.JAPAN, **overrides):
        defaults = {
            "term_type": TermType.HORSE,
            "source_language": SourceLanguage.JAPANESE,
            "source_ja": name,
            "target_zh": "",
            "racing_region": region,
            "is_active": True,
        }
        defaults.update(overrides)
        return TermEntry.objects.create(**defaults)

    def _profile(self, name: str, region: str = RacingRegion.JAPAN, **overrides):
        term = overrides.pop("primary_term", None) or self._term(name, region)
        defaults = {
            "primary_term": term,
            "original_name": name,
            "racing_region": region,
        }
        defaults.update(overrides)
        return HorseProfile.objects.create(**defaults)

    def _p0_source(self, profile, region: str = RacingRegion.JAPAN, **overrides):
        defaults = {
            "profile": profile,
            "source_type": HorseP0SourceType.MAJOR_RACE_PARTICIPANT,
            "status": HorseP0SourceStatus.ACTIVE,
            "racing_region": region,
            "horse_name": profile.original_name,
            "participant_key": f"test:{profile.pk}",
            "source_url": "https://example.test/race/1",
        }
        defaults.update(overrides)
        return HorseP0Source.objects.create(**defaults)


class P0HorseBatchSelectionTests(P0HorseBatchTestBase):
    def test_select_fail_closed_without_scope_and_limit(self):
        with self.assertRaises(P0HorseBatchError):
            select_p0_horse_batch()

    def test_select_fail_closed_total_limit_exceeded(self):
        with self.assertRaises(P0HorseBatchError):
            select_p0_horse_batch(
                regions=[RacingRegion.JAPAN],
                limit_per_region=99999,
            )

    def test_select_defaults_to_region_batch_limit(self):
        for index in range(120):
            profile = self._profile(f"テスト馬{index:03d}")
            self._p0_source(profile)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        self.assertEqual(len(manifest["horses"]), 100)
        self.assertEqual(manifest["region_counts"][RacingRegion.JAPAN], 100)

    def test_select_excludes_complete_profile_full(self):
        complete = self._profile(
            "完成馬",
            completeness_status=HorseProfileCompleteness.COMPLETE_PROFILE_FULL,
        )
        self._p0_source(complete)
        pending = self._profile("未完成馬")
        self._p0_source(pending)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        names = {horse["horse_name"] for horse in manifest["horses"]}
        self.assertIn("未完成馬", names)
        self.assertNotIn("完成馬", names)

    def test_select_include_complete_with_reason(self):
        complete = self._profile(
            "完成馬",
            completeness_status=HorseProfileCompleteness.COMPLETE_PROFILE_FULL,
        )
        self._p0_source(complete)
        manifest = select_p0_horse_batch(
            regions=[RacingRegion.JAPAN],
            include_complete=True,
        )
        horse = next(h for h in manifest["horses"] if h["horse_name"] == "完成馬")
        self.assertIn("include_complete_override", horse["queue_reasons"])

    def test_select_excludes_in_flight_profiles(self):
        busy = self._profile("在批馬")
        self._p0_source(busy)
        HorseProfileCompletionRun.objects.create(
            status="running",
            parameters={"p0_batch": {"profile_ids": [busy.pk]}},
        )
        free = self._profile("空闲馬")
        self._p0_source(free)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        names = {horse["horse_name"] for horse in manifest["horses"]}
        self.assertIn("空闲馬", names)
        self.assertNotIn("在批馬", names)

    def test_select_allow_in_flight_records_reason(self):
        busy = self._profile("在批馬")
        self._p0_source(busy)
        HorseProfileCompletionRun.objects.create(
            status="running",
            parameters={"p0_batch": {"profile_ids": [busy.pk]}},
        )
        manifest = select_p0_horse_batch(
            regions=[RacingRegion.JAPAN],
            allow_in_flight=True,
        )
        horse = next(h for h in manifest["horses"] if h["horse_name"] == "在批馬")
        self.assertIn("allow_in_flight_override", horse["queue_reasons"])

    def test_candidate_conversion_identity_and_enrichment(self):
        identified = self._profile(
            "有身份马",
            source_refs={"horse_identity_keys": ["jbis:0001234567"]},
        )
        self._p0_source(
            identified,
            evidence_payload={"horse_identity_keys": ["jbis:0001234567", "netkeiba:2010100001"]},
        )
        unidentified = self._profile("无身份马")
        self._p0_source(unidentified)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        by_name = {horse["horse_name"]: horse for horse in manifest["horses"]}
        identified_row = by_name["有身份马"]
        self.assertEqual(identified_row["candidate_key"], f"profile:{identified.pk}")
        self.assertIn("jbis:0001234567", identified_row["identity_keys"])
        self.assertIn("netkeiba:2010100001", identified_row["identity_keys"])
        # japan candidates prefer netkeiba when both keys exist (ID-direct
        # fetch path; JBIS name search fails closed on same-name horses)
        self.assertEqual(identified_row["source_namespace"], "netkeiba")
        self.assertNotIn("identity_status", identified_row)
        for url in identified_row["source_urls"]:
            self.assertNotIn("example.test/race", url)
        enrichment_row = by_name["无身份马"]
        self.assertEqual(enrichment_row["identity_status"], "needs_identity_enrichment")

    def test_candidate_conversion_birth_year_and_pedigree(self):
        import datetime

        profile = self._profile(
            "谱系马",
            birth_date=datetime.date(2020, 3, 15),
            sire_text="父马",
            dam_text="母马",
        )
        self._p0_source(profile)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        horse = manifest["horses"][0]
        self.assertEqual(horse["expected_birth_year"], 2020)
        self.assertEqual(horse["expected_sire_name"], "父马")
        self.assertEqual(horse["expected_dam_name"], "母马")

    def test_manifest_pending_shape(self):
        profile = self._profile("形状马")
        self._p0_source(profile)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN], operator="tester")
        self.assertEqual(manifest["schema_version"], "p0-horse-completion-batch.v1")
        self.assertEqual(manifest["status"], "pending")
        self.assertIsNone(manifest["approval"])
        self.assertEqual(manifest["created_by"], "tester")
        self.assertTrue(manifest["batch_id"].startswith("p0batch-"))
        horse = manifest["horses"][0]
        self.assertEqual(horse["queue_reasons"][0], "p0_source")
        self.assertEqual(horse["profile_id"], profile.pk)


@override_settings(HORSE_PROFILE_COMPLETION_BATCH_STATE_DIR=None)
class P0HorseBatchManifestTests(P0HorseBatchTestBase):
    def setUp(self):
        import tempfile

        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        self.state_dir = Path(self._tmp.name)
        profile = self._profile("清单马")
        self._p0_source(profile)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        self.manifest_path = write_batch_manifest(manifest, state_dir=self.state_dir)

    def test_write_and_load_roundtrip(self):
        loaded = load_batch_manifest(self.manifest_path)
        self.assertEqual(loaded["status"], "pending")
        self.assertEqual(
            loaded["batch_sha256"],
            json.loads(self.manifest_path.read_text(encoding="utf-8"))["batch_sha256"],
        )
        self.assertEqual(self.manifest_path.parent.name, loaded["batch_id"])

    def test_load_rejects_tampered_manifest(self):
        data = json.loads(self.manifest_path.read_text(encoding="utf-8"))
        data["horses"][0]["horse_name"] = "被篡改马"
        self.manifest_path.write_text(json.dumps(data), encoding="utf-8")
        with self.assertRaises(P0HorseBatchError):
            load_batch_manifest(self.manifest_path)

    def test_validate_requires_approved(self):
        with self.assertRaises(P0HorseBatchError):
            validate_approved_batch_manifest(self.manifest_path)

    def test_approve_writes_approval_and_ledger(self):
        approved = approve_batch_manifest(
            self.manifest_path,
            reviewer="reviewer-a",
            note="抽样通过",
        )
        self.assertEqual(approved["status"], "approved")
        self.assertEqual(approved["approval"]["reviewer"], "reviewer-a")
        self.assertEqual(approved["approval"]["note"], "抽样通过")
        self.assertTrue(approved["approval"]["approved_at"])
        ledger_path = self.manifest_path.parent / "approvals_ledger.jsonl"
        ledger_entries = [
            json.loads(line)
            for line in ledger_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        self.assertEqual(len(ledger_entries), 1)
        self.assertEqual(ledger_entries[0]["event"], "batch_approved")
        self.assertEqual(ledger_entries[0]["reviewer"], "reviewer-a")
        self.assertEqual(ledger_entries[0]["batch_sha256"], approved["batch_sha256"])
        validated = validate_approved_batch_manifest(
            self.manifest_path,
            expected_sha256=approved["batch_sha256"],
        )
        self.assertEqual(validated["batch_sha256"], approved["batch_sha256"])

    def test_validate_requires_explicit_expected_sha(self):
        approve_batch_manifest(self.manifest_path, reviewer="reviewer-a")
        with self.assertRaises(P0HorseBatchError):
            validate_approved_batch_manifest(self.manifest_path)

    def test_select_rejects_oversized_region_limit(self):
        with self.assertRaises(P0HorseBatchError):
            select_p0_horse_batch(
                regions=[RacingRegion.JAPAN],
                limit_per_region=101,
            )

    def test_approve_exclusion_writes_blocker_pool(self):
        excluded_id = self.manifest_horse_profile_id()
        approve_batch_manifest(
            self.manifest_path,
            reviewer="reviewer-a",
            excluded_profile_ids=[excluded_id],
            note="字段存疑",
        )
        pool_path = self.manifest_path.parent / "blocker_pool.jsonl"
        entries = [
            json.loads(line)
            for line in pool_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0]["profile_id"], excluded_id)
        self.assertEqual(entries[0]["reason"], "excluded_at_batch_approval")

    def test_validate_expected_sha_mismatch(self):
        approved = approve_batch_manifest(self.manifest_path, reviewer="reviewer-a")
        self.assertTrue(approved["batch_sha256"])
        with self.assertRaises(P0HorseBatchError):
            validate_approved_batch_manifest(
                self.manifest_path,
                expected_sha256="0" * 64,
            )

    def test_approve_requires_reviewer(self):
        with self.assertRaises(P0HorseBatchError):
            approve_batch_manifest(self.manifest_path, reviewer="")

    def test_approve_excluded_profiles_removed_from_horses(self):
        excluded_id = self.manifest_horse_profile_id()
        approved = approve_batch_manifest(
            self.manifest_path,
            reviewer="reviewer-a",
            excluded_profile_ids=[excluded_id],
            note="整匹排除",
        )
        self.assertEqual(approved["horses"], [])
        self.assertEqual(
            approved["approval"]["excluded_profile_ids"],
            [excluded_id],
        )

    def manifest_horse_profile_id(self) -> int:
        return json.loads(self.manifest_path.read_text(encoding="utf-8"))["horses"][0][
            "profile_id"
        ]


class P0HorseBatchCommandTests(P0HorseBatchTestBase):
    def setUp(self):
        import tempfile

        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        self.state_dir = Path(self._tmp.name)
        profile = self._profile("命令马")
        self._p0_source(profile)

    def _call(self, *command_args) -> dict:
        from io import StringIO

        from django.core.management import call_command

        out = StringIO()
        with self._state_dir_override():
            call_command("p0_horse_completion_batch", *command_args, "--json", stdout=out)
        text = out.getvalue()
        marker = text.rfind("\n{")
        payload = text[marker + 1 :] if marker != -1 else text
        return json.loads(payload)

    def _state_dir_override(self):
        return override_settings(
            HORSE_PROFILE_COMPLETION_BATCH_STATE_DIR=str(self.state_dir)
        )

    def test_select_approve_validate_roundtrip(self):
        selected = self._call("--select", "--regions", "japan")
        self.assertEqual(selected["status"], "pending")
        self.assertEqual(selected["horse_count"], 1)
        manifest_path = selected["manifest_path"]
        approved = self._call("--approve", manifest_path, "--reviewer", "reviewer-a")
        self.assertEqual(approved["status"], "approved")
        validated = self._call(
            "--validate",
            manifest_path,
            "--expected-sha256",
            approved["batch_sha256"],
        )
        self.assertEqual(validated["batch_sha256"], approved["batch_sha256"])

    def test_select_unbounded_rejected(self):
        from django.core.management.base import CommandError

        with self.assertRaises(CommandError):
            self._call("--select")

    def test_approve_requires_reviewer(self):
        from django.core.management.base import CommandError

        selected = self._call("--select", "--regions", "japan")
        with self.assertRaises(CommandError):
            self._call("--approve", selected["manifest_path"])


class BatchRunStateTests(TestCase):
    def setUp(self):
        import tempfile

        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        self.run_dir = Path(self._tmp.name) / "p0batch-abc123"

    def _candidate(self, **overrides):
        candidate = {
            "candidate_key": "profile:1",
            "region": "japan",
            "horse_name": "テスト馬",
            "identity_keys": ["jbis:0001"],
            "source_namespace": "jbis",
            "source_urls": ["https://www.jbis.or.jp/horse/0001/"],
            "expected_sire_name": "父",
            "expected_dam_name": "母",
            "expected_birth_year": 2020,
        }
        candidate.update(overrides)
        return candidate

    def test_write_read_roundtrip(self):
        from stable.services.p0_horse_completion_batch import BatchRunState

        state = BatchRunState.create(batch_id="p0batch-abc123", run_dir=self.run_dir)
        state.stage = "preparing"
        state.write()
        loaded = BatchRunState.read(self.run_dir)
        self.assertEqual(loaded.batch_id, "p0batch-abc123")
        self.assertEqual(loaded.stage, "preparing")
        self.assertEqual(loaded.candidate_states, {})
        self.assertEqual(loaded.resume_history, [])

    def test_input_fingerprint_stability(self):
        from stable.services.p0_horse_completion_batch import (
            candidate_input_fingerprint,
        )

        base = candidate_input_fingerprint(self._candidate())
        self.assertEqual(base, candidate_input_fingerprint(self._candidate()))
        changed = candidate_input_fingerprint(self._candidate(expected_sire_name="别的父"))
        self.assertNotEqual(base, changed)
        reordered = candidate_input_fingerprint(
            self._candidate(identity_keys=["jbis:0001", "netkeiba:2010"])
        )
        self.assertNotEqual(base, reordered)

    def test_netkeiba_parser_version_changes_adapter_and_candidate_fingerprints(self):
        from unittest.mock import patch

        from stable.services.p0_horse_completion_batch import (
            adapter_config_fingerprint,
            candidate_input_fingerprint,
        )

        candidate = self._candidate(
            identity_keys=["netkeiba:2022110137"],
            source_namespace="netkeiba",
        )
        base_adapter = adapter_config_fingerprint()
        base_candidate = candidate_input_fingerprint(candidate)
        self.assertEqual(base_adapter, adapter_config_fingerprint())
        self.assertEqual(base_candidate, candidate_input_fingerprint(candidate))
        with patch(
            "stable.services.p0_horse_completion_source_clients.NETKEIBA_PARSER_VERSION",
            "netkeiba-parser.test-next",
        ):
            self.assertNotEqual(base_adapter, adapter_config_fingerprint())
            self.assertNotEqual(base_candidate, candidate_input_fingerprint(candidate))

    def test_resume_decision_matrix(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            candidate_input_fingerprint,
            plan_candidate_resume,
            record_candidate_success,
        )

        state = BatchRunState.create(batch_id="p0batch-abc123", run_dir=self.run_dir)
        candidate = self._candidate()
        output_file = self.run_dir / "staging" / "profile_1.json"
        output_file.parent.mkdir(parents=True)
        output_file.write_text('{"ok": true}', encoding="utf-8")
        record_candidate_success(state, candidate, outputs={"payload": output_file})
        decisions = plan_candidate_resume(state, [candidate])
        self.assertEqual(decisions[candidate["candidate_key"]]["action"], "skipped_unchanged")

        state.candidate_states[candidate["candidate_key"]]["status"] = "failed"
        decisions = plan_candidate_resume(state, [candidate])
        self.assertEqual(decisions[candidate["candidate_key"]]["action"], "retry_failed")

        state.candidate_states[candidate["candidate_key"]]["status"] = "succeeded"
        changed = self._candidate(expected_birth_year=2021)
        decisions = plan_candidate_resume(state, [changed])
        self.assertEqual(
            decisions[changed["candidate_key"]]["action"], "rerun_input_changed"
        )

        decisions = plan_candidate_resume(state, [self._candidate(candidate_key="profile:2")])
        self.assertEqual(decisions["profile:2"]["action"], "executed")

    def test_resume_output_drift_codes(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            candidate_input_fingerprint,
            plan_candidate_resume,
        )

        state = BatchRunState.create(batch_id="p0batch-abc123", run_dir=self.run_dir)
        candidate = self._candidate()
        missing = self.run_dir / "staging" / "gone.json"
        state.candidate_states[candidate["candidate_key"]] = {
            "status": "succeeded",
            "input_fingerprint": candidate_input_fingerprint(candidate),
            "outputs": {"payload": str(missing)},
        }
        decisions = plan_candidate_resume(state, [candidate])
        self.assertEqual(
            decisions[candidate["candidate_key"]]["action"], "rerun_output_missing"
        )

        missing.parent.mkdir(parents=True, exist_ok=True)
        missing.write_text('{"ok": true}', encoding="utf-8")
        decisions = plan_candidate_resume(state, [candidate])
        self.assertEqual(
            decisions[candidate["candidate_key"]]["action"], "rerun_output_unverified"
        )

        from stable.services.p0_horse_completion_batch import record_candidate_success

        record_candidate_success(state, candidate, outputs={"payload": missing})
        missing.write_text('{"ok": false, "changed": true}', encoding="utf-8")
        decisions = plan_candidate_resume(state, [candidate])
        self.assertEqual(
            decisions[candidate["candidate_key"]]["action"], "rerun_output_changed"
        )

    def test_record_success_and_failure(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            candidate_input_fingerprint,
            record_candidate_failure,
            record_candidate_success,
        )

        state = BatchRunState.create(batch_id="p0batch-abc123", run_dir=self.run_dir)
        candidate = self._candidate()
        output_file = self.run_dir / "staging" / "profile_1.json"
        output_file.parent.mkdir(parents=True)
        output_file.write_text('{"ok": true}', encoding="utf-8")
        record_candidate_success(state, candidate, outputs={"payload": output_file})
        entry = state.candidate_states[candidate["candidate_key"]]
        self.assertEqual(entry["status"], "succeeded")
        self.assertEqual(
            entry["input_fingerprint"], candidate_input_fingerprint(candidate)
        )
        self.assertEqual(len(entry["outputs"]["payload"]["sha256"]), 64)

        record_candidate_failure(state, candidate, error="http_error: HTTP 500")
        entry = state.candidate_states[candidate["candidate_key"]]
        self.assertEqual(entry["status"], "failed")
        self.assertEqual(entry["error"], "http_error: HTTP 500")

    def test_downstream_stages_invalidated_on_rerun(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            invalidate_downstream_stages,
        )

        state = BatchRunState.create(batch_id="p0batch-abc123", run_dir=self.run_dir)
        state.completed_stages = [
            "prepare",
            "artifact",
            "review:japan",
            "commit:japan",
        ]
        state.artifacts = {
            "artifact_dir": "/tmp/x",
            "bundle:japan": {"research_path": "/tmp/r.json"},
            "commit:japan": {"artifact_sha256": "abc"},
        }
        invalidate_downstream_stages(state, reran=True)
        self.assertEqual(state.completed_stages, ["prepare"])
        self.assertEqual(state.artifacts, {"artifact_dir": "/tmp/x"})
        invalidate_downstream_stages(state, reran=False)
        self.assertEqual(state.completed_stages, ["prepare"])

    def test_resume_history_appended(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            append_resume_history,
        )

        state = BatchRunState.create(batch_id="p0batch-abc123", run_dir=self.run_dir)
        append_resume_history(
            state,
            from_stage="preparing",
            decisions={"skipped_unchanged": 3, "retry_failed": 1},
        )
        self.assertEqual(len(state.resume_history), 1)
        entry = state.resume_history[0]
        self.assertEqual(entry["from_stage"], "preparing")
        self.assertEqual(entry["decisions"]["retry_failed"], 1)
        self.assertTrue(entry["started_at"])
        loaded = BatchRunState.read(self.run_dir)
        self.assertEqual(len(loaded.resume_history), 1)

    def test_abandon_requires_reason_and_preserves_evidence(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
            abandon_batch_run,
        )

        state = BatchRunState.create(batch_id="p0batch-abc123", run_dir=self.run_dir)
        staging = self.run_dir / "staging" / "profile_1.json"
        staging.parent.mkdir(parents=True)
        staging.write_text('{"ok": true}', encoding="utf-8")
        with self.assertRaises(P0HorseBatchError):
            abandon_batch_run(state, reason="")
        abandon_batch_run(state, reason="批次构成有误，另起新批")
        self.assertEqual(state.stage, "abandoned")
        self.assertTrue(staging.exists())
        loaded = BatchRunState.read(self.run_dir)
        self.assertEqual(loaded.stage, "abandoned")
        self.assertEqual(loaded.errors[-1]["error"], "批次构成有误，另起新批")


class RequestBudgetToolTests(TestCase):
    def setUp(self):
        import tempfile

        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        self.budget_dir = Path(self._tmp.name)

    def test_check_request_budget_persists_count(self):
        import json as jsonlib

        from stable.services.p0_horse_completion_budget import load_race_event_request_budget_module

        check_request_budget = load_race_event_request_budget_module().check_request_budget

        artifact = self.budget_dir / "japan.json"
        check_request_budget(
            "https://www.jbis.or.jp/horse/0001/",
            artifact_path=artifact,
            max_requests=5,
            interval=0,
        )
        check_request_budget(
            "https://www.jbis.or.jp/horse/0002/",
            artifact_path=artifact,
            max_requests=5,
            interval=0,
        )
        state = jsonlib.loads(artifact.read_text(encoding="utf-8"))
        self.assertEqual(state["request_count"], 2)

    def test_check_request_budget_limit_exceeded_fail_closed(self):
        import json as jsonlib

        from stable.services.p0_horse_completion_budget import (
            load_race_event_request_budget_module,
        )

        _budget_module = load_race_event_request_budget_module()
        RequestBudgetExceeded = _budget_module.RequestBudgetExceeded
        check_request_budget = _budget_module.check_request_budget

        artifact = self.budget_dir / "japan.json"
        check_request_budget(
            "https://www.jbis.or.jp/horse/0001/",
            artifact_path=artifact,
            max_requests=1,
            interval=0,
        )
        with self.assertRaises(RequestBudgetExceeded):
            check_request_budget(
                "https://www.jbis.or.jp/horse/0002/",
                artifact_path=artifact,
                max_requests=1,
                interval=0,
            )
        state = jsonlib.loads(artifact.read_text(encoding="utf-8"))
        self.assertEqual(state["status"], "limit_exceeded")

    def test_check_request_budget_corrupted_artifact_fail_closed(self):
        from stable.services.p0_horse_completion_budget import (
            load_race_event_request_budget_module,
        )

        _budget_module = load_race_event_request_budget_module()
        RequestBudgetExceeded = _budget_module.RequestBudgetExceeded
        check_request_budget = _budget_module.check_request_budget

        artifact = self.budget_dir / "japan.json"
        artifact.write_text("not-json", encoding="utf-8")
        with self.assertRaises(RequestBudgetExceeded):
            check_request_budget(
                "https://www.jbis.or.jp/horse/0001/",
                artifact_path=artifact,
                max_requests=5,
                interval=0,
            )

    def test_host_interval_dir_derives_per_host_artifact(self):
        from stable.services.p0_horse_completion_budget import load_race_event_request_budget_module

        check_request_budget = load_race_event_request_budget_module().check_request_budget

        artifact = self.budget_dir / "japan.json"
        host_dir = self.budget_dir / "host-interval"
        check_request_budget(
            "https://www.jbis.or.jp/horse/0001/",
            artifact_path=artifact,
            max_requests=5,
            interval=0,
            host_interval_dir=host_dir,
        )
        check_request_budget(
            "https://db.netkeiba.com/horse/2010100001/",
            artifact_path=artifact,
            max_requests=5,
            interval=0,
            host_interval_dir=host_dir,
        )
        self.assertTrue((host_dir / "www.jbis.or.jp.json").exists())
        self.assertTrue((host_dir / "db.netkeiba.com.json").exists())

    def test_before_network_request_env_still_works(self):
        import json as jsonlib
        import os
        from unittest import mock

        from stable.services.p0_horse_completion_budget import (
            load_race_event_request_budget_module,
        )

        before_network_request = load_race_event_request_budget_module().before_network_request

        artifact = self.budget_dir / "race.json"
        with mock.patch.dict(
            os.environ,
            {
                "RACE_EVENT_CRAWL_MAX_REQUESTS": "3",
                "RACE_EVENT_CRAWL_REQUEST_BUDGET_ARTIFACT": str(artifact),
            },
        ):
            before_network_request("https://example.com/race/1")
        state = jsonlib.loads(artifact.read_text(encoding="utf-8"))
        self.assertEqual(state["request_count"], 1)


class P0HorseSourceRetryTests(TestCase):
    def _request(self, **overrides):
        from stable.services.p0_horse_completion_adapters import (
            P0HorseCompletionRequest,
        )

        defaults = {
            "candidate_key": "profile:1",
            "region": "japan",
            "horse_name": "テスト馬",
            "source_url": "https://www.jbis.or.jp/horse/0001234567/",
            "request_budget": 3,
            "batch_limit": 100,
        }
        defaults.update(overrides)
        return P0HorseCompletionRequest(**defaults)

    def _client(self, transport, **kwargs):
        from stable.services.p0_horse_completion_source_clients import (
            build_p0_horse_completion_source_client,
        )

        return build_p0_horse_completion_source_client(
            "japan",
            transport,
            **kwargs,
        )

    @override_settings(
        HORSE_PROFILE_COMPLETION_RETRY_MAX_ATTEMPTS=3,
        HORSE_PROFILE_COMPLETION_RETRY_BACKOFF_BASE_SECONDS=0,
    )
    def test_429_retried_then_succeeds(self):
        from unittest.mock import Mock

        transport = Mock()
        transport.get.side_effect = [
            Mock(status_code=429, text="", url="", headers={"Retry-After": "0"}),
            Mock(status_code=200, text="<html>ok</html>", url="", headers={}),
        ]
        ledger_calls = []
        client = self._client(
            transport,
            budget_hook=lambda url: ledger_calls.append(url),
        )
        response = client._get(
            "https://www.jbis.or.jp/horse/0001234567/",
            self._request(),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(transport.get.call_count, 2)
        self.assertEqual(len(ledger_calls), 2)

    @override_settings(
        HORSE_PROFILE_COMPLETION_RETRY_MAX_ATTEMPTS=3,
        HORSE_PROFILE_COMPLETION_RETRY_BACKOFF_BASE_SECONDS=0,
    )
    def test_403_not_retried(self):
        from unittest.mock import Mock

        from stable.services.p0_horse_completion_source_clients import (
            P0HorseSourceBlocked,
        )

        transport = Mock()
        transport.get.return_value = Mock(
            status_code=403, text="", url="", headers={}
        )
        client = self._client(transport)
        with self.assertRaises(P0HorseSourceBlocked) as ctx:
            client._get("https://www.jbis.or.jp/horse/0001234567/", self._request())
        self.assertEqual(transport.get.call_count, 1)
        self.assertFalse(ctx.exception.transient)
        self.assertEqual(ctx.exception.status_code, 403)

    @override_settings(
        HORSE_PROFILE_COMPLETION_RETRY_MAX_ATTEMPTS=2,
        HORSE_PROFILE_COMPLETION_RETRY_BACKOFF_BASE_SECONDS=0,
    )
    def test_retry_exhaustion_raises_transient_error(self):
        from unittest.mock import Mock

        from stable.services.p0_horse_completion_source_clients import (
            P0HorseSourceBlocked,
        )

        transport = Mock()
        transport.get.side_effect = RuntimeError("connection reset")
        client = self._client(transport)
        with self.assertRaises(P0HorseSourceBlocked) as ctx:
            client._get("https://www.jbis.or.jp/horse/0001234567/", self._request())
        self.assertEqual(transport.get.call_count, 2)
        self.assertTrue(ctx.exception.transient)

    @override_settings(
        HORSE_PROFILE_COMPLETION_RETRY_MAX_ATTEMPTS=3,
        HORSE_PROFILE_COMPLETION_RETRY_BACKOFF_BASE_SECONDS=0,
    )
    def test_retry_does_not_consume_per_candidate_budget(self):
        from unittest.mock import Mock

        transport = Mock()
        transport.get.side_effect = [
            Mock(status_code=500, text="", url="", headers={}),
            Mock(status_code=500, text="", url="", headers={}),
            Mock(status_code=200, text="<html>ok</html>", url="", headers={}),
        ]
        client = self._client(transport)
        request = self._request(request_budget=1)
        response = client._get(
            "https://www.jbis.or.jp/horse/0001234567/",
            request,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(client._request_count, 1)


class P0HorseBudgetLedgerTests(TestCase):
    def setUp(self):
        import tempfile

        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        self.budget_dir = Path(self._tmp.name)

    @override_settings(HORSE_PROFILE_COMPLETION_MAX_REQUESTS=10)
    def test_region_ledger_and_host_interval(self):
        import json as jsonlib

        from stable.services.p0_horse_completion_budget import (
            before_p0_horse_source_request,
        )

        before_p0_horse_source_request(
            "https://www.jbis.or.jp/horse/0001/",
            region="japan",
            budget_dir=self.budget_dir,
            interval=0,
        )
        before_p0_horse_source_request(
            "https://www.jbis.or.jp/horse/0002/",
            region="japan",
            budget_dir=self.budget_dir,
            interval=0,
        )
        ledger = jsonlib.loads(
            (self.budget_dir / "japan.json").read_text(encoding="utf-8")
        )
        self.assertEqual(ledger["request_count"], 2)
        self.assertTrue(
            (self.budget_dir / "host-interval" / "www.jbis.or.jp.json").exists()
        )

    @override_settings(HORSE_PROFILE_COMPLETION_MAX_REQUESTS=1)
    def test_region_ledger_limit_fail_closed(self):
        from stable.services.p0_horse_completion_budget import (
            before_p0_horse_source_request,
            load_race_event_request_budget_module,
        )

        RequestBudgetExceeded = load_race_event_request_budget_module().RequestBudgetExceeded

        before_p0_horse_source_request(
            "https://www.jbis.or.jp/horse/0001/",
            region="japan",
            budget_dir=self.budget_dir,
            interval=0,
        )
        with self.assertRaises(RequestBudgetExceeded):
            before_p0_horse_source_request(
                "https://www.jbis.or.jp/horse/0002/",
                region="japan",
                budget_dir=self.budget_dir,
                interval=0,
            )


class P0HorseBatchPrepareTests(P0HorseBatchTestBase):
    def setUp(self):
        import shutil
        import tempfile

        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        root = Path(self._tmp.name)
        self.state_dir = root / "batches"
        self.cache_dir = root / "cache"
        import datetime

        self.profile = self._profile(
            "FOREVER TEST",
            sire_text="Japan Sire",
            dam_text="Japan Dam",
            birth_date=datetime.date(2021, 2, 24),
            source_refs={
                "horse_identity_keys": ["jbis:jp-001"],
                "horse_source_urls": ["https://example.test/jbis/horse/jp-001"],
            },
        )
        self._p0_source(self.profile)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        self.manifest_path = write_batch_manifest(manifest, state_dir=self.state_dir)
        self.approved = approve_batch_manifest(self.manifest_path, reviewer="reviewer-a")
        from stable.services.p0_horse_completion_adapters import (
            p0_horse_completion_cache_path,
        )

        fixture = (
            Path(__file__).resolve().parent
            / "fixtures"
            / "p0_horse_completion"
            / "japan.json"
        )
        fixture_payload = json.loads(fixture.read_text(encoding="utf-8"))
        for record in fixture_payload["career"]["records"]:
            record.setdefault("source_name", "jbis")
        fixture_payload["career"]["records"][0]["race_date"] = "2023-11-05"
        cache_path = p0_horse_completion_cache_path(
            self.cache_dir, f"profile:{self.profile.pk}"
        )
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(
            json.dumps(fixture_payload, ensure_ascii=False, sort_keys=True),
            encoding="utf-8",
        )

    def _prepare(self, **overrides):
        from stable.services.p0_horse_completion_prepare import prepare_p0_horse_batch

        defaults = {
            "expected_sha256": self.approved["batch_sha256"],
            "allow_network": False,
            "cache_dir": self.cache_dir,
        }
        defaults.update(overrides)
        return prepare_p0_horse_batch(self.manifest_path, **defaults)

    def test_identity_source_error_is_explainable_prepare_blocker(self):
        import shutil

        if type(self) is not P0HorseBatchPrepareTests:
            return

        from stable.services.p0_horse_completion_adapters import (
            P0HorseCompletionSourceError,
        )

        class PartialExpectedIdentityClient:
            last_request_count = 0

            def fetch_source_payload(self, request):
                raise P0HorseCompletionSourceError(
                    "identity_incomplete: expected horse_name, sire_name, dam_name, and "
                    "birth_year; candidate expected fields missing: "
                    "expected_dam_name, expected_birth_year"
                )

        shutil.rmtree(self.cache_dir)
        summary = self._prepare(
            allow_network=True,
            source_client_factory=lambda region: PartialExpectedIdentityClient(),
        )
        self.assertEqual(
            summary["failure_reason_counts"]["source_cache_or_adapter_error"],
            1,
        )
        self.assertNotIn(
            "unexpected_adapter_error", summary["failure_reason_counts"]
        )
        staging_path = next((self.manifest_path.parent / "staging").iterdir())
        payload = json.loads(staging_path.read_text(encoding="utf-8"))
        self.assertNotIn("unexpected_adapter_error", payload["failure_reason"])
        self.assertEqual(
            payload["retrieval"]["error_message"],
            "identity_incomplete: expected horse_name, sire_name, dam_name, and "
            "birth_year; candidate expected fields missing: "
            "expected_dam_name, expected_birth_year",
        )

    def test_partial_career_source_error_is_explainable_prepare_blocker(self):
        import shutil

        if type(self) is not P0HorseBatchPrepareTests:
            return

        fixture = (
            Path(__file__).resolve().parent
            / "fixtures"
            / "p0_horse_completion"
            / "japan.json"
        )
        fixture_payload = json.loads(fixture.read_text(encoding="utf-8"))
        for record in fixture_payload["career"]["records"]:
            record.setdefault("source_name", "jbis")
        fixture_payload["career"]["records"][0]["finish"] = ""
        fixture_payload["career"]["records"][0]["result_status"] = ""

        class PartialCareerClient:
            last_request_count = 3

            def fetch_source_payload(self, request):
                return fixture_payload

        shutil.rmtree(self.cache_dir)
        summary = self._prepare(
            allow_network=True,
            source_client_factory=lambda region: PartialCareerClient(),
        )
        self.assertEqual(
            summary["failure_reason_counts"]["source_cache_or_adapter_error"],
            1,
        )
        self.assertNotIn(
            "unexpected_adapter_error", summary["failure_reason_counts"]
        )
        staging_path = next((self.manifest_path.parent / "staging").iterdir())
        payload = json.loads(staging_path.read_text(encoding="utf-8"))
        self.assertEqual(
            payload["retrieval"]["error_message"],
            "network working copy source payload failed validation: "
            "partial_career: record 1 lacks core evidence",
        )

    def test_prepare_cache_only_success(self):
        summary = self._prepare()
        self.assertEqual(summary["totals"]["horses"], 1)
        self.assertEqual(summary["totals"]["succeeded"], 1)
        self.assertEqual(summary["totals"]["blocked"], 0)
        artifact_dir = self.manifest_path.parent / "artifact"
        combined = artifact_dir / "combined_candidates.jsonl"
        self.assertTrue(combined.exists())
        lines = combined.read_text(encoding="utf-8").splitlines()
        self.assertEqual(len(lines), 1)
        payload = json.loads(lines[0])
        self.assertEqual(payload["candidate_key"], f"profile:{self.profile.pk}")
        self.assertTrue((artifact_dir / "batch_review.csv").exists())
        self.assertTrue((artifact_dir / "summary.json").exists())
        self.assertTrue((artifact_dir / "source_evidence_manifest.jsonl").exists())
        from stable.services.p0_horse_completion_batch import BatchRunState

        state = BatchRunState.read(self.manifest_path.parent)
        self.assertIn("prepare", state.completed_stages)
        self.assertIn("artifact", state.completed_stages)
        candidate_state = state.candidate_states[f"profile:{self.profile.pk}"]
        self.assertEqual(candidate_state["status"], "succeeded")
        self.assertEqual(len(candidate_state["outputs"]["payload"]["sha256"]), 64)

    def test_prepare_requires_approved_manifest(self):
        pending_manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        pending_path = write_batch_manifest(pending_manifest, state_dir=self.state_dir / "other")
        from stable.services.p0_horse_completion_batch import P0HorseBatchError
        from stable.services.p0_horse_completion_prepare import prepare_p0_horse_batch

        with self.assertRaises(P0HorseBatchError):
            prepare_p0_horse_batch(
                pending_path,
                allow_network=False,
                cache_dir=self.cache_dir,
            )

    def test_prepare_expected_sha_mismatch(self):
        from stable.services.p0_horse_completion_batch import P0HorseBatchError

        with self.assertRaises(P0HorseBatchError):
            self._prepare(expected_sha256="0" * 64)

    def test_prepare_resume_skips_unchanged(self):
        from stable.services.p0_horse_completion_batch import BatchRunState

        self._prepare()
        staging_file = next((self.manifest_path.parent / "staging").glob("*.json"))
        first_bytes = staging_file.read_bytes()
        summary = self._prepare()
        self.assertEqual(summary["resume"]["skipped_unchanged"], 1)
        self.assertEqual(staging_file.read_bytes(), first_bytes)
        state = BatchRunState.read(self.manifest_path.parent)
        self.assertGreaterEqual(len(state.resume_history), 1)

    def test_prepare_blocked_when_cache_missing(self):
        other = self._profile("无缓存马")
        self._p0_source(other)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        manifest_path = write_batch_manifest(manifest, state_dir=self.state_dir / "batch2")
        approved = approve_batch_manifest(manifest_path, reviewer="reviewer-a")
        from stable.services.p0_horse_completion_prepare import prepare_p0_horse_batch

        summary = prepare_p0_horse_batch(
            manifest_path,
            expected_sha256=approved["batch_sha256"],
            allow_network=False,
            cache_dir=self.cache_dir,
        )
        self.assertEqual(summary["totals"]["horses"], 2)
        self.assertEqual(summary["totals"]["succeeded"], 1)
        self.assertEqual(summary["totals"]["blocked"], 1)
        combined = (manifest_path.parent / "artifact" / "combined_candidates.jsonl")
        payloads = [
            json.loads(line)
            for line in combined.read_text(encoding="utf-8").splitlines()
        ]
        blocked = next(p for p in payloads if p["candidate_key"] == f"profile:{other.pk}")
        self.assertIn("network_disabled_cache_missing", blocked["failure_reason"])

    def test_prepare_aborts_on_budget_exhaustion(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )
        from stable.services.p0_horse_completion_budget import (
            load_race_event_request_budget_module,
        )

        no_cache = self._profile("触网马")
        self._p0_source(no_cache)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        manifest_path = write_batch_manifest(manifest, state_dir=self.state_dir / "batch3")
        approved = approve_batch_manifest(manifest_path, reviewer="reviewer-a")

        budget_error = load_race_event_request_budget_module().RequestBudgetExceeded

        class _ExhaustedClient:
            def fetch_source_payload(self, request):
                raise budget_error("budget exhausted: 1/1")

            def has_manual_supplements(self, request):
                return False

        from stable.services.p0_horse_completion_prepare import prepare_p0_horse_batch

        with self.assertRaises(P0HorseBatchError):
            prepare_p0_horse_batch(
                manifest_path,
                expected_sha256=approved["batch_sha256"],
                allow_network=True,
                cache_dir=self.cache_dir,
                source_client_factory=lambda region: _ExhaustedClient(),
            )
        state = BatchRunState.read(manifest_path.parent)
        self.assertEqual(state.stage, "prepare_failed")
        failed = state.candidate_states[f"profile:{no_cache.pk}"]
        self.assertEqual(failed["status"], "failed")


class P0HorseBatchResearchTests(P0HorseBatchPrepareTests):
    def _research(self, region=RacingRegion.JAPAN):
        from stable.services.p0_horse_completion_research import (
            build_region_research_v3,
        )

        self._prepare()
        return build_region_research_v3(
            self.manifest_path.parent / "artifact",
            region=region,
        )

    def test_converter_deterministic(self):
        first = self._research()
        second = self._research()
        self.assertEqual(
            json.dumps(first, ensure_ascii=False, sort_keys=True),
            json.dumps(second, ensure_ascii=False, sort_keys=True),
        )

    def test_research_v3_shape(self):
        research = self._research()
        self.assertEqual(research["schema_version"], "p0-horse-research.v3")
        self.assertEqual(len(research["horses"]), 1)
        horse = research["horses"][0]
        self.assertEqual(horse["region"], "japan")
        self.assertEqual(horse["identity"]["horse_name"], "FOREVER TEST")
        self.assertEqual(horse["identity"]["sire_name"], "Japan Sire")
        self.assertEqual(horse["identity"]["birth_year"], 2021)
        self.assertTrue(horse["source"]["url"].startswith("https://"))
        self.assertEqual(horse["source"]["external_horse_id"], "jp-001")
        self.assertEqual(horse["candidate"]["sample_region"], "japan")
        self.assertTrue(horse["source_evidence"])
        self.assertEqual(horse["basic_profile"]["owner_name"], "Japan Owner")
        self.assertEqual(horse["pedigree"]["sire"], "Japan Sire")
        self.assertEqual(len(horse["career"]["records"]), 4)
        self.assertEqual(horse["career"]["official_or_source_start_count"], 3)
        self.assertEqual(
            horse["career"]["record_authority_status"], "source_records_verified"
        )
        self.assertTrue(horse["aliases"])

    def test_converter_excludes_blocked_payloads(self):
        other = self._profile("无缓存马")
        self._p0_source(other)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        manifest_path = write_batch_manifest(manifest, state_dir=self.state_dir / "batch-blocked")
        approved = approve_batch_manifest(manifest_path, reviewer="reviewer-a")
        from stable.services.p0_horse_completion_prepare import prepare_p0_horse_batch
        from stable.services.p0_horse_completion_research import (
            build_region_research_v3,
        )

        prepare_p0_horse_batch(
            manifest_path,
            expected_sha256=approved["batch_sha256"],
            allow_network=False,
            cache_dir=self.cache_dir,
        )
        research = build_region_research_v3(
            manifest_path.parent / "artifact",
            region=RacingRegion.JAPAN,
        )
        names = {horse["identity"]["horse_name"] for horse in research["horses"]}
        self.assertIn("FOREVER TEST", names)
        self.assertNotIn("无缓存马", names)


class P0HorseBatchApprovalBundleTests(P0HorseBatchPrepareTests):
    def setUp(self):
        super().setUp()
        from django.contrib.auth import get_user_model

        self.reviewer = get_user_model().objects.create_user(
            username="p0-batch-reviewer",
            password="unused",
            is_superuser=True,
            is_staff=True,
        )
        self._prepare()

    def _bundle(self, **overrides):
        from stable.services.p0_horse_completion_research import (
            build_region_approval_bundle,
            build_region_research_v3,
            write_region_research,
        )

        research = build_region_research_v3(
            self.manifest_path.parent / "artifact",
            region=RacingRegion.JAPAN,
        )
        research_path, research_sha = write_region_research(
            research,
            output_dir=self.manifest_path.parent / "approval",
            region=RacingRegion.JAPAN,
        )
        defaults = {
            "research_path": research_path,
            "region": RacingRegion.JAPAN,
            "reviewer": self.reviewer,
            "output_dir": self.manifest_path.parent / "approval",
            "batch_dir": self.manifest_path.parent,
        }
        defaults.update(overrides)
        return build_region_approval_bundle(**defaults)

    def _freeze_candidate_for_builder(
        self, *, artifact_path, artifact_sha, artifact, bundle
    ):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            _append_approvals_ledger,
            load_batch_manifest,
        )
        from stable.services.p0_horse_completion_commit import (
            _build_auto_first_publish_scope,
        )
        from stable.services.p0_horse_completion_research import (
            _write_canonical,
        )

        batch_dir = self.manifest_path.parent
        manifest = load_batch_manifest(self.manifest_path)
        combined = batch_dir / "artifact" / "combined_candidates.jsonl"
        combined_sha = hashlib.sha256(combined.read_bytes()).hexdigest()
        scope = _build_auto_first_publish_scope(artifact)
        candidate = {
            "schema_version": "p0_horse_production_release_candidate.v1",
            "status": "pending_independent_release_approval",
            "batch_id": manifest["batch_id"],
            "region": RacingRegion.JAPAN,
            "executor_reviewer_id": self.reviewer.id,
            "artifact_prepared_at": artifact["prepared_at"],
            "bindings": {
                "batch_manifest_sha256": manifest["batch_sha256"],
                "combined_candidates_sha256": combined_sha,
                "research_v3_sha256": bundle["research_sha256"],
                "authority_manifest_sha256": bundle["authority_sha256"],
                "profile_mapping_decisions_sha256": bundle["mapping_sha256"],
                "production_snapshot_sha256": artifact[
                    "production_snapshot_sha256"
                ],
                "final_artifact_sha256": artifact_sha,
            },
            "expected_actions": artifact["expected_actions"],
            "auto_first_publish_scope": scope,
        }
        pending = batch_dir / "approval" / ".release_candidate_japan.pending"
        candidate_sha = _write_canonical(pending, candidate)
        candidate_path = batch_dir / "approval" / (
            f"release_candidate_japan_{candidate_sha}.json"
        )
        pending.replace(candidate_path)
        _append_approvals_ledger(
            batch_dir,
            {
                "event": "release_candidate_prepared",
                "batch_id": manifest["batch_id"],
                "region": RacingRegion.JAPAN,
                "release_candidate_sha256": candidate_sha,
                "artifact_sha256": artifact_sha,
            },
        )
        state = BatchRunState.read(batch_dir)
        history = {
            "path": str(candidate_path),
            "sha256": candidate_sha,
            "artifact_path": str(artifact_path),
            "artifact_sha256": artifact_sha,
            "publish_scope": scope,
        }
        state.artifacts[f"release_candidate:japan:{candidate_sha}"] = history
        state.artifacts["release_candidate:japan"] = history
        state.write()
        return {
            "path": candidate_path,
            "sha256": candidate_sha,
            "publish_scope": scope,
        }

    def test_bundle_binds_existing_profile(self):
        bundle = self._bundle()
        self.assertEqual(bundle["mapping"]["schema_version"], "p0-horse-profile-mapping-decisions.v1")
        self.assertEqual(bundle["mapping"]["review_status"], "approved")
        self.assertEqual(bundle["mapping"]["research_v3_sha256"], bundle["research_sha256"])
        rows = bundle["mapping"]["rows"]
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["decision"], "bind_existing")
        self.assertEqual(row["profile_id"], self.profile.pk)
        self.assertEqual(row["rejected_profile_ids"], [])
        for module in ("profile", "pedigree", "race_record", "major_wins"):
            review = row["module_reviews"][module]
            self.assertEqual(review["status"], "approved")
            self.assertGreaterEqual(review["confidence"], 90)
        self.assertTrue(row["database_mapping_snapshot"]["sha256"])
        authority = bundle["authority"]
        self.assertEqual(authority["review_status"], "approved")
        self.assertEqual(authority["horses"], [])

    def test_bundle_us_horses_fail_closed_without_authorization(self):
        us_profile = self._profile(
            "US TEST",
            region=RacingRegion.UNITED_STATES,
            sire_text="US Sire",
            dam_text="US Dam",
        )
        us_profile.birth_date = __import__("datetime").date(2020, 1, 1)
        us_profile.save()
        self._p0_source(us_profile, region=RacingRegion.UNITED_STATES)
        from stable.services.p0_horse_completion_research import (
            P0HorseBatchError,
            build_region_approval_bundle,
            build_region_research_v3,
            write_region_research,
        )

        research = build_region_research_v3(
            self.manifest_path.parent / "artifact",
            region=RacingRegion.JAPAN,
        )
        research["horses"].append(
            {
                "identity": {
                    "horse_name": "US TEST",
                    "sire_name": "US Sire",
                    "dam_name": "US Dam",
                    "birth_year": 2020,
                },
                "region": "united_states",
                "source": {"name": "hrn", "url": "https://example.test/hrn/us-test", "external_horse_id": "us-1"},
                "candidate": {"sample_region": "united_states"},
                "source_evidence": [{"source_url": "https://example.test/hrn/us-test"}],
                "basic_profile": {},
                "pedigree": {},
                "aliases": [],
                "career": {"records": [], "record_authority_status": "source_records_verified"},
                "confidence": 100,
            }
        )
        research_path, _ = write_region_research(
            research,
            output_dir=self.manifest_path.parent / "approval",
            region=RacingRegion.JAPAN,
        )
        with self.assertRaises(P0HorseBatchError):
            build_region_approval_bundle(
                research_path=research_path,
                region=RacingRegion.JAPAN,
                reviewer=self.reviewer,
                output_dir=self.manifest_path.parent / "approval",
                batch_dir=self.manifest_path.parent,
            )

    def test_bundle_writes_ledger_entry(self):
        bundle = self._bundle()
        ledger_path = self.manifest_path.parent / "approvals_ledger.jsonl"
        events = [
            json.loads(line)
            for line in ledger_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        module_events = [e for e in events if e["event"] == "region_modules_approved"]
        self.assertEqual(len(module_events), 1)
        entry = module_events[0]
        self.assertEqual(entry["region"], "japan")
        self.assertEqual(entry["mapping_sha256"], bundle["mapping_sha256"])
        self.assertEqual(entry["reviewer"], self.reviewer.get_username())

    def test_bundle_feeds_existing_commit_chain_end_to_end(self):
        from stable.services.p0_horse_completion_research import (
            _write_canonical,
            build_region_release_manifest,
        )
        from stable.services.p0_horse_production_apply import (
            commit_reviewed_p0_completion_artifact,
            dry_run_reviewed_p0_completion_artifact,
            prepare_reviewed_p0_completion_artifact,
        )

        bundle = self._bundle()
        artifact = prepare_reviewed_p0_completion_artifact(
            research_v3_path=bundle["research_path"],
            authority_manifest_path=bundle["authority_path"],
            authority_manifest_sha256=bundle["authority_sha256"],
            profile_mapping_decisions_path=bundle["mapping_path"],
            reviewer_id=self.reviewer.id,
        )
        artifact_path = self.manifest_path.parent / "approval" / "commit_artifact_japan.json"
        artifact_sha = _write_canonical(artifact_path, artifact)
        candidate = self._freeze_candidate_for_builder(
            artifact_path=artifact_path,
            artifact_sha=artifact_sha,
            artifact=artifact,
            bundle=bundle,
        )
        release = build_region_release_manifest(
            artifact_path=artifact_path,
            artifact_sha256=artifact_sha,
            bundle=bundle,
            reviewer=self.reviewer,
            approved_by="human-approver",
            batch_dir=self.manifest_path.parent,
            region=RacingRegion.JAPAN,
            release_candidate_path=candidate["path"],
            release_candidate_sha256=candidate["sha256"],
            expected_publish_scope=candidate["publish_scope"],
        )
        report = dry_run_reviewed_p0_completion_artifact(
            artifact_path=artifact_path,
            artifact_sha256=artifact_sha,
            release_manifest_path=release["release_path"],
            release_manifest_sha256=release["release_sha256"],
        )
        self.assertGreater(
            report["planned_profile_updates"] + report["planned_race_record_creates"],
            0,
        )
        commit_reviewed_p0_completion_artifact(
            artifact_path=artifact_path,
            artifact_sha256=artifact_sha,
            release_manifest_path=release["release_path"],
            release_manifest_sha256=release["release_sha256"],
            confirm_reviewed_artifact=True,
        )
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.owner_name, "Japan Owner")
        self.assertEqual(self.profile.trainer_name, "Japan Trainer")
        verification = dry_run_reviewed_p0_completion_artifact(
            artifact_path=artifact_path,
            artifact_sha256=artifact_sha,
            release_manifest_path=release["release_path"],
            release_manifest_sha256=release["release_sha256"],
        )
        self.assertEqual(verification["planned_profile_creates"], 0)
        self.assertEqual(verification["planned_profile_updates"], 0)
        self.assertEqual(verification["planned_race_record_creates"], 0)

    def test_release_manifest_without_ledger_entry_rejected(self):
        from stable.services.p0_horse_completion_research import (
            _write_canonical,
            build_region_release_manifest,
        )
        from stable.services.p0_horse_production_apply import (
            dry_run_reviewed_p0_completion_artifact,
            prepare_reviewed_p0_completion_artifact,
        )

        bundle = self._bundle()
        artifact = prepare_reviewed_p0_completion_artifact(
            research_v3_path=bundle["research_path"],
            authority_manifest_path=bundle["authority_path"],
            authority_manifest_sha256=bundle["authority_sha256"],
            profile_mapping_decisions_path=bundle["mapping_path"],
            reviewer_id=self.reviewer.id,
        )
        artifact_path = self.manifest_path.parent / "approval" / "commit_artifact_japan.json"
        artifact_sha = _write_canonical(artifact_path, artifact)
        candidate = self._freeze_candidate_for_builder(
            artifact_path=artifact_path,
            artifact_sha=artifact_sha,
            artifact=artifact,
            bundle=bundle,
        )
        release = build_region_release_manifest(
            artifact_path=artifact_path,
            artifact_sha256=artifact_sha,
            bundle=bundle,
            reviewer=self.reviewer,
            approved_by="human-approver",
            batch_dir=self.manifest_path.parent,
            region=RacingRegion.JAPAN,
            release_candidate_path=candidate["path"],
            release_candidate_sha256=candidate["sha256"],
            expected_publish_scope=candidate["publish_scope"],
        )
        # remove the release_approved ledger line to simulate an unapproved channel
        ledger_path = self.manifest_path.parent / "approvals_ledger.jsonl"
        kept = [
            line
            for line in ledger_path.read_text(encoding="utf-8").splitlines()
            if '"release_approved"' not in line
        ]
        ledger_path.write_text("\n".join(kept) + "\n", encoding="utf-8")
        with self.assertRaises(Exception) as ctx:
            dry_run_reviewed_p0_completion_artifact(
                artifact_path=artifact_path,
                artifact_sha256=artifact_sha,
                release_manifest_path=release["release_path"],
                release_manifest_sha256=release["release_sha256"],
            )
        self.assertIn("ledger", str(ctx.exception))

    def test_release_builder_requires_candidate_sha_before_any_evidence_write(self):
        from stable.services.p0_horse_completion_batch import P0HorseBatchError
        from stable.services.p0_horse_completion_research import (
            build_region_release_manifest,
        )

        bundle = self._bundle()
        batch_dir = self.manifest_path.parent
        ledger_path = batch_dir / "approvals_ledger.jsonl"
        ledger_before = ledger_path.read_bytes()
        release_before = {
            path.name
            for path in (batch_dir / "approval").glob("release_manifest_*.json")
        }
        with self.assertRaisesRegex(P0HorseBatchError, "candidate SHA"):
            build_region_release_manifest(
                artifact_path=batch_dir / "approval" / "missing.json",
                artifact_sha256="0" * 64,
                bundle=bundle,
                reviewer=self.reviewer,
                approved_by="human-approver",
                batch_dir=batch_dir,
                region=RacingRegion.JAPAN,
            )
        self.assertEqual(ledger_path.read_bytes(), ledger_before)
        self.assertEqual(
            {
                path.name
                for path in (batch_dir / "approval").glob(
                    "release_manifest_*.json"
                )
            },
            release_before,
        )

    def test_release_builder_rejects_forged_candidate_context_without_writes(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
            _append_approvals_ledger,
            load_batch_manifest,
        )
        from stable.services.p0_horse_completion_research import (
            _write_canonical,
            build_region_release_manifest,
        )
        from stable.services.p0_horse_production_apply import (
            prepare_reviewed_p0_completion_artifact,
        )

        bundle = self._bundle()
        artifact = prepare_reviewed_p0_completion_artifact(
            research_v3_path=bundle["research_path"],
            authority_manifest_path=bundle["authority_path"],
            authority_manifest_sha256=bundle["authority_sha256"],
            profile_mapping_decisions_path=bundle["mapping_path"],
            reviewer_id=self.reviewer.id,
        )
        batch_dir = self.manifest_path.parent
        artifact_path = batch_dir / "approval" / "commit_artifact_japan.json"
        artifact_sha = _write_canonical(artifact_path, artifact)
        genuine = self._freeze_candidate_for_builder(
            artifact_path=artifact_path,
            artifact_sha=artifact_sha,
            artifact=artifact,
            bundle=bundle,
        )
        common = {
            "artifact_path": artifact_path,
            "artifact_sha256": artifact_sha,
            "bundle": bundle,
            "reviewer": self.reviewer,
            "approved_by": "human-approver",
            "batch_dir": batch_dir,
            "region": RacingRegion.JAPAN,
            "expected_publish_scope": genuine["publish_scope"],
        }
        ledger_path = batch_dir / "approvals_ledger.jsonl"
        release_before = set(
            (batch_dir / "approval").glob("release_manifest_*.json")
        )

        with self.assertRaises(P0HorseBatchError):
            build_region_release_manifest(
                **common,
                release_candidate_path=genuine["path"],
                release_candidate_sha256="f" * 64,
            )

        genuine_payload = json.loads(
            genuine["path"].read_text(encoding="utf-8")
        )
        forged_payload = {
            **genuine_payload,
            "expected_actions": {
                **genuine_payload["expected_actions"],
                "profile_updates": 999,
            },
        }
        forged_pending = batch_dir / "approval" / ".forged_candidate.pending"
        forged_sha = _write_canonical(forged_pending, forged_payload)
        forged_path = batch_dir / "approval" / (
            f"release_candidate_japan_{forged_sha}.json"
        )
        forged_pending.replace(forged_path)
        state = BatchRunState.read(batch_dir)
        forged_history = {
            "path": str(forged_path),
            "sha256": forged_sha,
            "artifact_path": str(artifact_path),
            "artifact_sha256": artifact_sha,
            "publish_scope": genuine["publish_scope"],
        }
        state.artifacts[
            f"release_candidate:japan:{forged_sha}"
        ] = forged_history
        state.write()
        manifest = load_batch_manifest(self.manifest_path)
        _append_approvals_ledger(
            batch_dir,
            {
                "event": "release_candidate_prepared",
                "batch_id": manifest["batch_id"],
                "region": RacingRegion.JAPAN,
                "release_candidate_sha256": forged_sha,
                "artifact_sha256": artifact_sha,
            },
        )
        ledger_before = ledger_path.read_bytes()
        with self.assertRaisesRegex(P0HorseBatchError, "release context"):
            build_region_release_manifest(
                **common,
                release_candidate_path=forged_path,
                release_candidate_sha256=forged_sha,
            )
        self.assertEqual(ledger_path.read_bytes(), ledger_before)
        self.assertEqual(
            set((batch_dir / "approval").glob("release_manifest_*.json")),
            release_before,
        )

        genuine_backup = genuine["path"].with_suffix(".backup")
        genuine["path"].replace(genuine_backup)
        genuine["path"].symlink_to(genuine_backup)
        ledger_before = ledger_path.read_bytes()
        with self.assertRaisesRegex(P0HorseBatchError, "regular file"):
            build_region_release_manifest(
                **common,
                release_candidate_path=genuine["path"],
                release_candidate_sha256=genuine["sha256"],
            )
        self.assertEqual(ledger_path.read_bytes(), ledger_before)
        self.assertEqual(
            set((batch_dir / "approval").glob("release_manifest_*.json")),
            release_before,
        )


class P0HorseBatchReviewWorkbookTests(P0HorseBatchPrepareTests):
    def test_workbook_sheets_and_exception_sampling(self):
        from openpyxl import load_workbook

        from stable.services.p0_horse_completion_batch import load_batch_manifest
        from stable.services.p0_horse_completion_review import (
            EXCEPTION_SHEET,
            SUMMARY_SHEET,
            build_batch_review_workbook,
        )

        self._prepare()
        manifest = load_batch_manifest(self.manifest_path)
        output = Path(self._tmp.name) / "review" / f"{manifest['batch_id']}.xlsx"
        build_batch_review_workbook(
            manifest=manifest,
            artifact_dir=self.manifest_path.parent / "artifact",
            output_path=output,
        )
        self.assertTrue(output.exists())
        workbook = load_workbook(output)
        self.assertIn(SUMMARY_SHEET, workbook.sheetnames)
        self.assertIn("日本", workbook.sheetnames)
        self.assertIn(EXCEPTION_SHEET, workbook.sheetnames)
        japan_sheet = workbook["日本"]
        self.assertEqual(japan_sheet.max_row, 2)
        header = [cell.value for cell in japan_sheet[1]]
        self.assertIn("马名", header)
        self.assertIn("生涯缺口", header)
        name_col = header.index("马名") + 1
        self.assertEqual(japan_sheet.cell(row=2, column=name_col).value, "FOREVER TEST")

    def test_workbook_lists_blocked_horse_in_exceptions(self):
        from openpyxl import load_workbook

        from stable.services.p0_horse_completion_batch import load_batch_manifest
        from stable.services.p0_horse_completion_prepare import prepare_p0_horse_batch
        from stable.services.p0_horse_completion_review import (
            EXCEPTION_SHEET,
            build_batch_review_workbook,
        )

        other = self._profile("无缓存马")
        self._p0_source(other)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        manifest_path = write_batch_manifest(manifest, state_dir=self.state_dir / "batch-review")
        approved = approve_batch_manifest(manifest_path, reviewer="reviewer-a")
        prepare_p0_horse_batch(
            manifest_path,
            expected_sha256=approved["batch_sha256"],
            allow_network=False,
            cache_dir=self.cache_dir,
        )
        output = Path(self._tmp.name) / "review" / "blocked.xlsx"
        build_batch_review_workbook(
            manifest=load_batch_manifest(manifest_path),
            artifact_dir=manifest_path.parent / "artifact",
            output_path=output,
        )
        workbook = load_workbook(output)
        sheet = workbook[EXCEPTION_SHEET]
        names = [sheet.cell(row=row, column=5).value for row in range(2, sheet.max_row + 1)]
        self.assertIn("无缓存马", names)
        self.assertNotIn("FOREVER TEST", names)


class P0HorseBatchCommandPipelineTests(P0HorseBatchPrepareTests):
    def setUp(self):
        super().setUp()
        from django.contrib.auth import get_user_model

        self.reviewer = get_user_model().objects.create_user(
            username="p0-pipeline-reviewer",
            password="unused",
            is_superuser=True,
            is_staff=True,
        )

    def _call(self, *command_args) -> dict:
        from io import StringIO

        from django.core.management import call_command

        out = StringIO()
        with override_settings(
            HORSE_PROFILE_COMPLETION_BATCH_STATE_DIR=str(self.state_dir),
            HORSE_PROFILE_COMPLETION_CACHE_DIR=str(self.cache_dir),
            HORSE_PROFILE_COMPLETION_REVIEW_OUTPUT_DIR=str(
                Path(self._tmp.name) / "review"
            ),
        ):
            call_command("p0_horse_completion_batch", *command_args, "--json", stdout=out)
        text = out.getvalue()
        marker = text.rfind("\n{")
        payload = text[marker + 1 :] if marker != -1 else text
        return json.loads(payload)

    def _prepare_release_candidate(self) -> dict:
        return self._call(
            "--prepare-release",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )

    def test_prepare_release_is_deterministic_read_only_and_unapproved(self):
        from stable.models import (
            HorseP0Source,
            HorseProfile,
            HorseProfileDataCandidate,
            HorseRaceRecord,
            OperationLog,
            TaskExecutionLog,
        )
        from stable.services.p0_horse_completion_batch import BatchRunState

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        before = {
            "profiles": HorseProfile.objects.count(),
            "records": HorseRaceRecord.objects.count(),
            "sources": HorseP0Source.objects.count(),
            "candidates": HorseProfileDataCandidate.objects.count(),
            "operation_logs": OperationLog.objects.count(),
            "task_logs": TaskExecutionLog.objects.count(),
            "review_status": self.profile.review_status,
        }
        first = self._prepare_release_candidate()
        first_bytes = Path(first["release_candidate_path"]).read_bytes()
        second = self._prepare_release_candidate()
        second_bytes = Path(second["release_candidate_path"]).read_bytes()
        candidate = json.loads(first_bytes)
        artifact = json.loads(
            Path(first["artifact_path"]).read_text(encoding="utf-8")
        )

        self.assertEqual(first["release_candidate_sha256"], second["release_candidate_sha256"])
        self.assertEqual(first_bytes, second_bytes)
        self.assertEqual(
            candidate["schema_version"],
            "p0_horse_production_release_candidate.v1",
        )
        self.assertEqual(candidate["status"], "pending_independent_release_approval")
        self.assertNotIn("approved_by", candidate)
        self.assertEqual(
            set(candidate["bindings"]),
            {
                "batch_manifest_sha256",
                "combined_candidates_sha256",
                "research_v3_sha256",
                "authority_manifest_sha256",
                "profile_mapping_decisions_sha256",
                "production_snapshot_sha256",
                "final_artifact_sha256",
            },
        )
        self.assertEqual(candidate["expected_actions"]["profile_updates"], 1)
        self.assertEqual(
            candidate["bindings"]["research_v3_sha256"],
            artifact["inputs"]["research_v3"]["sha256"],
        )
        self.assertEqual(
            candidate["bindings"]["authority_manifest_sha256"],
            artifact["inputs"]["authority_manifest"]["sha256"],
        )
        self.assertEqual(
            candidate["bindings"]["profile_mapping_decisions_sha256"],
            artifact["inputs"]["profile_mapping_decisions"]["sha256"],
        )
        self.assertEqual(
            candidate["auto_first_publish_scope"]["existing_profiles"],
            [
                {
                    "profile_id": self.profile.pk,
                    "review_status": self.profile.review_status,
                    "hidden": False,
                    "manual_lock": False,
                    "disposition": "attempt_publish_after_commit",
                }
            ],
        )
        state = BatchRunState.read(self.manifest_path.parent)
        self.assertEqual(
            state.artifacts["release_candidate:japan"]["sha256"],
            first["release_candidate_sha256"],
        )
        ledger = [
            json.loads(line)
            for line in (
                self.manifest_path.parent / "approvals_ledger.jsonl"
            ).read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        self.assertEqual(
            sum(entry.get("event") == "release_candidate_prepared" for entry in ledger),
            1,
        )
        self.assertFalse(any(entry.get("event") == "release_approved" for entry in ledger))
        self.profile.refresh_from_db()
        self.assertEqual(
            before,
            {
                "profiles": HorseProfile.objects.count(),
                "records": HorseRaceRecord.objects.count(),
                "sources": HorseP0Source.objects.count(),
                "candidates": HorseProfileDataCandidate.objects.count(),
                "operation_logs": OperationLog.objects.count(),
                "task_logs": TaskExecutionLog.objects.count(),
                "review_status": self.profile.review_status,
            },
        )

    def test_publish_scope_is_derived_only_from_artifact_rows(self):
        from stable.services.p0_horse_completion_commit import (
            _build_auto_first_publish_scope,
        )

        unrelated_blocker = self._profile("同批阻断马")
        artifact = {
            "rows": [
                {
                    "deterministic_identity_key": "a" * 64,
                    "identity": {"horse_name": self.profile.original_name},
                    "resolution": {
                        "decision": "bind_existing",
                        "profile_id": self.profile.pk,
                    },
                },
                {
                    "deterministic_identity_key": "b" * 64,
                    "identity": {"horse_name": "本次新建马"},
                    "resolution": {"decision": "create_new"},
                },
            ]
        }
        scope = _build_auto_first_publish_scope(artifact)
        self.assertEqual(
            [row["profile_id"] for row in scope["existing_profiles"]],
            [self.profile.pk],
        )
        self.assertEqual(
            scope["create_new_identities"],
            [
                {
                    "deterministic_identity_key": "b" * 64,
                    "horse_name": "本次新建马",
                    "disposition": "attempt_publish_after_commit",
                }
            ],
        )
        self.assertNotIn(
            unrelated_blocker.pk,
            [row["profile_id"] for row in scope["existing_profiles"]],
        )

    def test_full_pipeline_via_command(self):
        prepared = self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self.assertEqual(prepared["totals"]["succeeded"], 1)
        self.assertTrue(Path(prepared["review_workbook"]).exists())

        bundled = self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        self.assertEqual(bundled["horse_count"], 1)
        candidate = self._prepare_release_candidate()

        committed = self._call(
            "--commit",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
            "--approved-by",
            "human-approver",
            "--release-candidate-sha256",
            candidate["release_candidate_sha256"],
            "--confirm-reviewed-artifact",
        )
        self.assertTrue(committed["idempotent_verification"]["passed"])
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.owner_name, "Japan Owner")
        self.assertEqual(self.profile.completeness_status, "complete_profile_full")

        run = HorseProfileCompletionRun.objects.get(id=committed["completion_run_id"])
        self.assertEqual(
            [
                item["profile_id"]
                for item in run.parameters["p0_batch"]["publish_scope"][
                    "existing_profiles"
                ]
            ],
            [self.profile.pk],
        )
        self.assertEqual(run.parameters["p0_batch"]["region"], "japan")
        self.assertTrue(run.summary["idempotent_verification"]["passed"])

    def _build_direct_apply_v2_context(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
        )
        from stable.services.p0_horse_completion_research import (
            build_region_release_manifest,
        )

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate_result = self._prepare_release_candidate()
        candidate_sha = candidate_result["release_candidate_sha256"]
        state = BatchRunState.read(self.manifest_path.parent)
        history = state.artifacts[
            f"release_candidate:japan:{candidate_sha}"
        ]
        artifact_path = Path(history["artifact_path"])
        artifact_sha = history["artifact_sha256"]
        release = build_region_release_manifest(
            artifact_path=artifact_path,
            artifact_sha256=artifact_sha,
            bundle=state.artifacts["bundle:japan"],
            reviewer=self.reviewer,
            approved_by="human-approver",
            batch_dir=self.manifest_path.parent,
            region="japan",
            release_candidate_path=history["path"],
            release_candidate_sha256=candidate_sha,
            expected_publish_scope=history["publish_scope"],
        )
        return {
            "candidate_sha": candidate_sha,
            "artifact_path": artifact_path,
            "artifact_sha": artifact_sha,
            "release": release,
        }

    def test_direct_apply_rejects_v2_release_after_supersede(self):
        from stable.services.p0_horse_completion_batch import (
            _append_approvals_ledger,
        )
        from stable.services.p0_horse_production_apply import (
            P0ReviewedArtifactError,
            commit_reviewed_p0_completion_artifact,
            dry_run_reviewed_p0_completion_artifact,
        )

        context = self._build_direct_apply_v2_context()
        candidate_sha = context["candidate_sha"]
        artifact_path = context["artifact_path"]
        artifact_sha = context["artifact_sha"]
        release = context["release"]
        active = dry_run_reviewed_p0_completion_artifact(
            artifact_path=artifact_path,
            artifact_sha256=artifact_sha,
            release_manifest_path=release["release_path"],
            release_manifest_sha256=release["release_sha256"],
        )
        self.assertEqual(active["planned_profile_updates"], 1)

        _append_approvals_ledger(
            self.manifest_path.parent,
            {
                "event": "release_superseded",
                "region": "japan",
                "old_release_candidate_sha256": candidate_sha,
                "old_release_manifest_sha256": release["release_sha256"],
                "new_release_candidate_sha256": "b" * 64,
                "new_release_manifest_sha256": "c" * 64,
            },
        )
        before_status = self.profile.review_status
        before_records = self.profile.race_records.count()
        for apply in (
            dry_run_reviewed_p0_completion_artifact,
            commit_reviewed_p0_completion_artifact,
        ):
            kwargs = {
                "artifact_path": artifact_path,
                "artifact_sha256": artifact_sha,
                "release_manifest_path": release["release_path"],
                "release_manifest_sha256": release["release_sha256"],
            }
            if apply is commit_reviewed_p0_completion_artifact:
                kwargs["confirm_reviewed_artifact"] = True
            with self.assertRaisesRegex(
                P0ReviewedArtifactError, "superseded"
            ):
                apply(**kwargs)
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.review_status, before_status)
        self.assertEqual(self.profile.race_records.count(), before_records)

    def test_direct_apply_rejects_state_or_manifest_abandoned_without_db_writes(self):
        from stable.models import (
            HorseProfileDataCandidate,
            HorseRaceRecord,
            OperationLog,
        )
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            load_batch_manifest,
            mark_batch_manifest_status,
        )
        from stable.services.p0_horse_production_apply import (
            P0ReviewedArtifactError,
            commit_reviewed_p0_completion_artifact,
            dry_run_reviewed_p0_completion_artifact,
        )

        context = self._build_direct_apply_v2_context()
        apply_kwargs = {
            "artifact_path": context["artifact_path"],
            "artifact_sha256": context["artifact_sha"],
            "release_manifest_path": context["release"]["release_path"],
            "release_manifest_sha256": context["release"]["release_sha256"],
        }
        counted_models = (
            HorseProfile,
            HorseProfileDataCandidate,
            HorseRaceRecord,
            HorseProfileCompletionRun,
            OperationLog,
        )
        baseline = {
            model: model.objects.count() for model in counted_models
        }

        state = BatchRunState.read(self.manifest_path.parent)
        original_stage = state.stage
        state.stage = "abandoned"
        state.write()
        for apply in (
            dry_run_reviewed_p0_completion_artifact,
            commit_reviewed_p0_completion_artifact,
        ):
            kwargs = dict(apply_kwargs)
            if apply is commit_reviewed_p0_completion_artifact:
                kwargs["confirm_reviewed_artifact"] = True
            with self.assertRaisesRegex(
                P0ReviewedArtifactError, "abandoned"
            ):
                apply(**kwargs)
        self.assertEqual(
            {model: model.objects.count() for model in counted_models},
            baseline,
        )

        state.stage = original_stage
        state.write()
        mark_batch_manifest_status(self.manifest_path, status="abandoned")
        self.assertEqual(
            load_batch_manifest(self.manifest_path)["status"],
            "abandoned",
        )
        for apply in (
            dry_run_reviewed_p0_completion_artifact,
            commit_reviewed_p0_completion_artifact,
        ):
            kwargs = dict(apply_kwargs)
            if apply is commit_reviewed_p0_completion_artifact:
                kwargs["confirm_reviewed_artifact"] = True
            with self.assertRaisesRegex(
                P0ReviewedArtifactError, "abandoned"
            ):
                apply(**kwargs)
        self.assertEqual(
            {model: model.objects.count() for model in counted_models},
            baseline,
        )

    def test_direct_apply_validates_batch_manifest_schema_and_internal_sha(self):
        from stable.services.p0_horse_production_apply import (
            P0ReviewedArtifactError,
            dry_run_reviewed_p0_completion_artifact,
        )

        context = self._build_direct_apply_v2_context()
        apply_kwargs = {
            "artifact_path": context["artifact_path"],
            "artifact_sha256": context["artifact_sha"],
            "release_manifest_path": context["release"]["release_path"],
            "release_manifest_sha256": context["release"]["release_sha256"],
        }
        original = self.manifest_path.read_bytes()
        original_payload = json.loads(original)
        mutations = (
            {"schema_version": "invalid-batch-schema"},
            {"batch_sha256": "0" * 64},
        )
        for mutation in mutations:
            with self.subTest(mutation=mutation):
                payload = {**original_payload, **mutation}
                self.manifest_path.write_text(
                    json.dumps(payload, ensure_ascii=False, sort_keys=True)
                    + "\n",
                    encoding="utf-8",
                )
                with self.assertRaisesRegex(
                    P0ReviewedArtifactError, "schema or internal SHA"
                ):
                    dry_run_reviewed_p0_completion_artifact(**apply_kwargs)
        self.manifest_path.write_bytes(original)

    def test_direct_v2_uncommitted_rejects_current_manifest_and_combined_drift(self):
        from stable.services.p0_horse_completion_batch import _manifest_sha256
        from stable.services.p0_horse_production_apply import (
            P0ReviewedArtifactError,
            commit_reviewed_p0_completion_artifact,
            dry_run_reviewed_p0_completion_artifact,
        )

        context = self._build_direct_apply_v2_context()
        apply_kwargs = {
            "artifact_path": context["artifact_path"],
            "artifact_sha256": context["artifact_sha"],
            "release_manifest_path": context["release"]["release_path"],
            "release_manifest_sha256": context["release"]["release_sha256"],
        }
        original_manifest = self.manifest_path.read_bytes()
        manifest = json.loads(original_manifest)
        manifest["reviewer"] = "different-reviewer"
        manifest["batch_sha256"] = _manifest_sha256(manifest)
        self.manifest_path.write_text(
            json.dumps(manifest, ensure_ascii=False, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        with self.assertRaisesRegex(
            P0ReviewedArtifactError, "batch manifest binding drift"
        ):
            dry_run_reviewed_p0_completion_artifact(**apply_kwargs)

        self.manifest_path.write_bytes(original_manifest)
        combined_path = (
            self.manifest_path.parent / "artifact" / "combined_candidates.jsonl"
        )
        combined_path.write_bytes(combined_path.read_bytes() + b"\n")
        with self.assertRaisesRegex(
            P0ReviewedArtifactError, "combined candidates binding drift"
        ):
            commit_reviewed_p0_completion_artifact(
                **apply_kwargs,
                confirm_reviewed_artifact=True,
            )
        self.assertFalse(
            HorseProfileCompletionRun.objects.filter(status="committed").exists()
        )

    def test_direct_v2_committed_recovery_ignores_current_input_drift(self):
        from stable.services.p0_horse_completion_batch import _manifest_sha256
        from stable.services.p0_horse_production_apply import (
            P0ReviewedArtifactError,
            commit_reviewed_p0_completion_artifact,
            dry_run_reviewed_p0_completion_artifact,
        )

        context = self._build_direct_apply_v2_context()
        apply_kwargs = {
            "artifact_path": context["artifact_path"],
            "artifact_sha256": context["artifact_sha"],
            "release_manifest_path": context["release"]["release_path"],
            "release_manifest_sha256": context["release"]["release_sha256"],
        }
        first = commit_reviewed_p0_completion_artifact(
            **apply_kwargs,
            confirm_reviewed_artifact=True,
        )
        self.assertGreater(first["database_write_count"], 0)

        manifest = json.loads(self.manifest_path.read_text(encoding="utf-8"))
        manifest["reviewer"] = "post-commit-drift"
        manifest["batch_sha256"] = _manifest_sha256(manifest)
        self.manifest_path.write_text(
            json.dumps(manifest, ensure_ascii=False, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        combined_path = (
            self.manifest_path.parent / "artifact" / "combined_candidates.jsonl"
        )
        combined_path.write_bytes(combined_path.read_bytes() + b"\n")

        dry_run = dry_run_reviewed_p0_completion_artifact(**apply_kwargs)
        repeated = commit_reviewed_p0_completion_artifact(
            **apply_kwargs,
            confirm_reviewed_artifact=True,
        )
        self.assertEqual(dry_run["planned_profile_creates"], 0)
        self.assertEqual(dry_run["planned_profile_updates"], 0)
        self.assertEqual(repeated["database_write_count"], 0)

        committed_run = HorseProfileCompletionRun.objects.get(
            status="committed"
        )
        committed_run.artifact_path = f"{committed_run.artifact_path}.other"
        committed_run.save(update_fields=["artifact_path", "updated_at"])
        with self.assertRaisesRegex(
            P0ReviewedArtifactError, "batch manifest binding drift"
        ):
            dry_run_reviewed_p0_completion_artifact(**apply_kwargs)

    def test_direct_v2_commit_holds_reentrant_execution_lock_after_validation(self):
        from concurrent.futures import ThreadPoolExecutor
        from contextlib import contextmanager
        from threading import Event
        from unittest import mock

        from stable.services import (
            p0_horse_completion_commit as commit_module,
            p0_horse_production_apply as apply_module,
        )
        from stable.services.p0_horse_completion_batch import (
            _append_approvals_ledger,
        )

        context = self._build_direct_apply_v2_context()
        loaded_artifact = apply_module._load_artifact(
            context["artifact_path"],
            context["artifact_sha"],
        )
        validation_finished = Event()
        release_direct = Event()
        supersede_started = Event()
        supersede_finished = Event()

        @contextmanager
        def pause_before_transaction(_rows):
            validation_finished.set()
            self.assertTrue(release_direct.wait(timeout=5))
            raise apply_module.P0ReviewedArtifactError("stop before database")
            yield

        def activate_candidate_b(*_args, **_kwargs):
            _append_approvals_ledger(
                self.manifest_path.parent,
                {
                    "event": "release_superseded",
                    "region": "japan",
                    "old_release_candidate_sha256": context["candidate_sha"],
                    "old_release_manifest_sha256": context["release"][
                        "release_sha256"
                    ],
                    "new_release_candidate_sha256": "b" * 64,
                    "new_release_manifest_sha256": "c" * 64,
                },
            )
            return {"status": "candidate-b-active"}

        def supersede():
            supersede_started.set()
            commit_module.commit_p0_horse_batch_region(
                self.manifest_path,
                region="japan",
                reviewer=self.reviewer,
                approved_by="human-approver",
                release_candidate_sha256="b" * 64,
                state_dir=self.state_dir,
                confirm_reviewed_artifact=True,
            )
            supersede_finished.set()

        with mock.patch.object(
            apply_module,
            "_validate_reviewer",
            return_value=self.reviewer,
        ), mock.patch.object(
            apply_module,
            "_load_artifact",
            return_value=loaded_artifact,
        ), mock.patch.object(
            apply_module,
            "_artifact_has_exact_committed_run",
            return_value=False,
        ), mock.patch.object(
            apply_module,
            "_identity_session_lock_scope",
            side_effect=pause_before_transaction,
        ), mock.patch.object(
            commit_module,
            "_commit_p0_horse_batch_region_locked",
            side_effect=activate_candidate_b,
        ):
            with ThreadPoolExecutor(max_workers=2) as executor:
                direct = executor.submit(
                    apply_module.commit_reviewed_p0_completion_artifact,
                    artifact_path=context["artifact_path"],
                    artifact_sha256=context["artifact_sha"],
                    release_manifest_path=context["release"]["release_path"],
                    release_manifest_sha256=context["release"]["release_sha256"],
                    confirm_reviewed_artifact=True,
                )
                self.assertTrue(validation_finished.wait(timeout=5))
                competing = executor.submit(supersede)
                self.assertTrue(supersede_started.wait(timeout=5))
                self.assertFalse(supersede_finished.wait(timeout=0.2))
                release_direct.set()
                with self.assertRaisesRegex(
                    apply_module.P0ReviewedArtifactError, "stop before database"
                ):
                    direct.result(timeout=5)
                competing.result(timeout=5)

    def test_direct_v2_commit_revalidates_supersede_after_waiting_for_lock(self):
        from concurrent.futures import ThreadPoolExecutor
        from threading import Event
        from unittest import mock

        from stable.services import (
            p0_horse_completion_commit as commit_module,
            p0_horse_production_apply as apply_module,
        )
        from stable.models import (
            HorseProfileDataCandidate,
            HorseRaceRecord,
            OperationLog,
            TaskExecutionLog,
        )
        from stable.services.p0_horse_completion_batch import (
            _append_approvals_ledger,
            batch_execution_window,
        )

        context = self._build_direct_apply_v2_context()
        loaded_artifact = apply_module._load_artifact(
            context["artifact_path"],
            context["artifact_sha"],
        )
        direct_started = Event()
        counted_models = (
            HorseProfile,
            HorseProfileDataCandidate,
            HorseRaceRecord,
            HorseProfileCompletionRun,
            OperationLog,
            TaskExecutionLog,
        )
        baseline = {
            model: model.objects.count() for model in counted_models
        }

        def direct_commit():
            direct_started.set()
            return apply_module.commit_reviewed_p0_completion_artifact(
                artifact_path=context["artifact_path"],
                artifact_sha256=context["artifact_sha"],
                release_manifest_path=context["release"]["release_path"],
                release_manifest_sha256=context["release"]["release_sha256"],
                confirm_reviewed_artifact=True,
            )

        def activate_candidate_b(*_args, **_kwargs):
            _append_approvals_ledger(
                self.manifest_path.parent,
                {
                    "event": "release_superseded",
                    "region": "japan",
                    "old_release_candidate_sha256": context["candidate_sha"],
                    "old_release_manifest_sha256": context["release"][
                        "release_sha256"
                    ],
                    "new_release_candidate_sha256": "b" * 64,
                    "new_release_manifest_sha256": "c" * 64,
                },
            )
            return {"status": "candidate-b-active"}

        executor = ThreadPoolExecutor(max_workers=1)
        try:
            with mock.patch.object(
                apply_module,
                "_validate_reviewer",
                return_value=self.reviewer,
            ), mock.patch.object(
                apply_module,
                "_load_artifact",
                return_value=loaded_artifact,
            ), mock.patch.object(
                apply_module,
                "_artifact_has_exact_committed_run",
                return_value=False,
            ), mock.patch.object(
                commit_module,
                "_commit_p0_horse_batch_region_locked",
                side_effect=activate_candidate_b,
            ):
                with batch_execution_window(self.manifest_path.parent):
                    commit_module.commit_p0_horse_batch_region(
                        self.manifest_path,
                        region="japan",
                        reviewer=self.reviewer,
                        approved_by="human-approver",
                        release_candidate_sha256="b" * 64,
                        state_dir=self.state_dir,
                        confirm_reviewed_artifact=True,
                    )
                    direct = executor.submit(direct_commit)
                    self.assertTrue(direct_started.wait(timeout=5))
                    self.assertFalse(direct.done())
                with self.assertRaisesRegex(
                    apply_module.P0ReviewedArtifactError, "superseded"
                ):
                    direct.result(timeout=5)
        finally:
            executor.shutdown(wait=True)
        self.assertEqual(
            {model: model.objects.count() for model in counted_models},
            baseline,
        )

    def test_legacy_auto_publish_ledger_is_normalized_without_rewrite(self):
        import shutil

        from stable.services.p0_horse_completion_batch import (
            read_approvals_ledger,
        )

        fixture = (
            Path(__file__).resolve().parent
            / "fixtures"
            / "p0_horse_completion"
            / "approvals_ledger_legacy_auto_first_publish.jsonl"
        )
        ledger_path = self.manifest_path.parent / "approvals_ledger.jsonl"
        shutil.copyfile(fixture, ledger_path)
        before = ledger_path.read_bytes()
        entries = read_approvals_ledger(self.manifest_path.parent)
        publish = next(
            entry
            for entry in entries
            if entry["event"] == "auto_first_publish"
        )
        self.assertEqual(publish["frozen_exclusions"], [])
        self.assertEqual(publish["frozen_exclusion_counts"], {})
        self.assertEqual(ledger_path.read_bytes(), before)

    def test_v2_auto_publish_ledger_requires_frozen_exclusions(self):
        from stable.services.p0_horse_completion_batch import (
            read_approvals_ledger,
        )

        ledger_path = self.manifest_path.parent / "approvals_ledger.jsonl"
        ledger_path.write_text(
            json.dumps(
                {
                    "event": "auto_first_publish",
                    "event_schema": "p0_horse_auto_first_publish.v2",
                    "batch_id": self.approved["batch_id"],
                    "region": "japan",
                    "artifact_sha256": "a" * 64,
                    "published": 1,
                    "skipped_already_published": 0,
                    "blocked": 0,
                    "published_profile_ids": [self.profile.pk],
                    "at": "2026-07-24T00:00:00Z",
                },
                sort_keys=True,
            )
            + "\n",
            encoding="utf-8",
        )
        with self.assertRaisesRegex(P0HorseBatchError, "partial"):
            read_approvals_ledger(self.manifest_path.parent)

    def test_abandon_rejects_committed_manifest_without_mutation(self):
        from django.core.management.base import CommandError

        from stable.services.p0_horse_completion_batch import (
            mark_batch_manifest_status,
        )

        mark_batch_manifest_status(self.manifest_path, status="committed")
        state_path = self.manifest_path.parent / "state.json"
        self.assertFalse(state_path.exists())
        manifest_before = self.manifest_path.read_bytes()
        with self.assertRaisesRegex(CommandError, "committed"):
            self._call(
                "--abandon",
                str(self.manifest_path),
                "--note",
                "不得伪装撤回",
            )
        self.assertEqual(self.manifest_path.read_bytes(), manifest_before)
        self.assertFalse(state_path.exists())

    def test_abandon_rejects_commit_or_publish_checkpoint_without_mutation(self):
        from django.core.management.base import CommandError

        from stable.services.p0_horse_completion_batch import BatchRunState

        state = BatchRunState.create(
            batch_id=self.approved["batch_id"],
            run_dir=self.manifest_path.parent,
        )
        for checkpoint in ("commit:japan", "publish:japan"):
            with self.subTest(checkpoint=checkpoint):
                state.artifacts = {checkpoint: {"status": "recorded"}}
                state.completed_stages = []
                state.write()
                manifest_before = self.manifest_path.read_bytes()
                state_before = (
                    self.manifest_path.parent / "state.json"
                ).read_bytes()
                with self.assertRaisesRegex(CommandError, "checkpoint"):
                    self._call(
                        "--abandon",
                        str(self.manifest_path),
                        "--note",
                        "已有生产 checkpoint",
                    )
                self.assertEqual(
                    self.manifest_path.read_bytes(), manifest_before
                )
                self.assertEqual(
                    (self.manifest_path.parent / "state.json").read_bytes(),
                    state_before,
                )

    def test_abandon_rejects_committed_candidate_run_without_checkpoint(self):
        from django.core.management.base import CommandError

        from stable.models import HorseCompletionRunStatus
        from stable.services.p0_horse_completion_batch import BatchRunState

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate = self._prepare_release_candidate()
        state = BatchRunState.read(self.manifest_path.parent)
        history = state.artifacts[
            "release_candidate:japan:"
            + candidate["release_candidate_sha256"]
        ]
        HorseProfileCompletionRun.objects.create(
            status=HorseCompletionRunStatus.COMMITTED,
            artifact_path=history["artifact_path"],
            summary={"artifact_sha256": history["artifact_sha256"]},
        )
        manifest_before = self.manifest_path.read_bytes()
        state_before = (self.manifest_path.parent / "state.json").read_bytes()
        with self.assertRaisesRegex(CommandError, "already committed"):
            self._call(
                "--abandon",
                str(self.manifest_path),
                "--note",
                "state checkpoint 已被删除",
            )
        self.assertEqual(self.manifest_path.read_bytes(), manifest_before)
        self.assertEqual(
            (self.manifest_path.parent / "state.json").read_bytes(),
            state_before,
        )

    def test_ledger_append_rejects_malformed_tail_without_writing(self):
        from stable.services.p0_horse_completion_batch import (
            _append_approvals_ledger,
        )

        ledger_path = self.manifest_path.parent / "approvals_ledger.jsonl"
        with ledger_path.open("a", encoding="utf-8") as handle:
            handle.write('{"event":"release_approved"')
        before = ledger_path.read_bytes()
        with self.assertRaisesRegex(P0HorseBatchError, "malformed"):
            _append_approvals_ledger(
                self.manifest_path.parent,
                {
                    "event": "release_candidate_prepared",
                    "batch_id": self.approved["batch_id"],
                    "region": "japan",
                    "release_candidate_sha256": "a" * 64,
                    "artifact_sha256": "b" * 64,
                },
            )
        self.assertEqual(ledger_path.read_bytes(), before)

    def test_ledger_append_flushes_and_fsyncs_complete_line(self):
        from unittest import mock

        from stable.services import p0_horse_completion_batch as batch_module

        ledger_path = self.manifest_path.parent / "approvals_ledger.jsonl"
        with mock.patch.object(batch_module.os, "fsync") as fsync:
            batch_module._append_approvals_ledger(
                self.manifest_path.parent,
                {
                    "event": "release_candidate_prepared",
                    "batch_id": self.approved["batch_id"],
                    "region": "japan",
                    "release_candidate_sha256": "a" * 64,
                    "artifact_sha256": "b" * 64,
                },
            )
        fsync.assert_called_once()
        final_line = ledger_path.read_text(encoding="utf-8").splitlines()[-1]
        self.assertEqual(
            json.loads(final_line)["release_candidate_sha256"],
            "a" * 64,
        )

    def test_commit_requires_confirm_flag(self):
        from django.core.management.base import CommandError

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate = self._prepare_release_candidate()
        with self.assertRaises(CommandError):
            self._call(
                "--commit",
                str(self.manifest_path),
                "--region",
                "japan",
                "--reviewer-id",
                str(self.reviewer.id),
                "--approved-by",
                "human-approver",
                "--release-candidate-sha256",
                candidate["release_candidate_sha256"],
            )

    def test_commit_requires_exact_release_candidate_sha(self):
        from django.core.management.base import CommandError

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        self._prepare_release_candidate()
        with self.assertRaisesRegex(CommandError, "candidate SHA-256 mismatch"):
            self._call(
                "--commit",
                str(self.manifest_path),
                "--region",
                "japan",
                "--reviewer-id",
                str(self.reviewer.id),
                "--approved-by",
                "human-approver",
                "--release-candidate-sha256",
                "0" * 64,
                "--confirm-reviewed-artifact",
            )

    def test_commit_rejects_each_release_candidate_binding_drift(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )
        from stable.services.p0_horse_completion_commit import (
            commit_p0_horse_batch_region,
        )
        from stable.services.p0_horse_completion_research import _write_canonical

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        prepared = self._prepare_release_candidate()
        candidate_path = Path(prepared["release_candidate_path"])
        original = json.loads(candidate_path.read_text(encoding="utf-8"))
        for field in original["bindings"]:
            with self.subTest(binding=field):
                mutated = json.loads(json.dumps(original))
                mutated["bindings"][field] = "f" * 64
                mutated_sha = _write_canonical(candidate_path, mutated)
                state = BatchRunState.read(self.manifest_path.parent)
                state.artifacts["release_candidate:japan"]["sha256"] = mutated_sha
                state.write()
                with self.assertRaisesRegex(
                    P0HorseBatchError, "SHA-256 mismatch|bindings drifted"
                ):
                    commit_p0_horse_batch_region(
                        self.manifest_path,
                        region="japan",
                        reviewer=self.reviewer,
                        approved_by="human-approver",
                        release_candidate_sha256=mutated_sha,
                        state_dir=self.state_dir,
                        confirm_reviewed_artifact=True,
                    )

    def test_publish_scope_records_all_existing_dispositions(self):
        from django.utils import timezone

        from stable.services.p0_horse_completion_commit import (
            _build_auto_first_publish_scope,
        )

        published = self._profile("已发布马")
        published.review_status = "published"
        published.save(update_fields=["review_status"])
        hidden = self._profile("隐藏马")
        hidden.review_status = "hidden"
        hidden.hidden_at = timezone.now()
        hidden.save(update_fields=["review_status", "hidden_at"])
        locked = self._profile("人工锁马")
        locked.manual_lock_flags = {"auto_publish_blocked": True}
        locked.save(update_fields=["manual_lock_flags"])
        rows = []
        for index, profile in enumerate(
            [self.profile, published, hidden, locked], start=1
        ):
            rows.append(
                {
                    "deterministic_identity_key": str(index) * 64,
                    "identity": {"horse_name": profile.original_name},
                    "resolution": {
                        "decision": "bind_existing",
                        "profile_id": profile.pk,
                    },
                }
            )
        scope = _build_auto_first_publish_scope({"rows": rows})
        dispositions = {
            item["profile_id"]: item["disposition"]
            for item in scope["existing_profiles"]
        }
        self.assertEqual(dispositions[self.profile.pk], "attempt_publish_after_commit")
        self.assertEqual(dispositions[published.pk], "skip_already_published")
        self.assertEqual(dispositions[hidden.pk], "block_hidden")
        self.assertEqual(dispositions[locked.pk], "block_manual_lock")

    def test_prepare_and_bundle_hold_shared_lock_during_generation(self):
        from concurrent.futures import ThreadPoolExecutor
        from unittest import mock

        from stable.services import (
            p0_horse_completion_prepare as prepare_module,
            p0_horse_completion_research as research_module,
        )
        from stable.services.p0_horse_completion_batch import (
            P0HorseBatchError,
            batch_serial_window,
        )

        def competing_lock_error():
            def compete():
                with batch_serial_window(self.state_dir):
                    return None

            with ThreadPoolExecutor(max_workers=1) as executor:
                return executor.submit(compete).exception(timeout=5)

        real_prepare = prepare_module.prepare_p0_horse_batch

        def controlled_prepare(*args, **kwargs):
            self.assertIsInstance(competing_lock_error(), P0HorseBatchError)
            return real_prepare(*args, **kwargs)

        with mock.patch.object(
            prepare_module,
            "prepare_p0_horse_batch",
            side_effect=controlled_prepare,
        ):
            self._call(
                "--prepare",
                str(self.manifest_path),
                "--expected-sha256",
                self.approved["batch_sha256"],
            )

        real_research = research_module.build_region_research_v3

        def controlled_research(*args, **kwargs):
            self.assertIsInstance(competing_lock_error(), P0HorseBatchError)
            return real_research(*args, **kwargs)

        with mock.patch.object(
            research_module,
            "build_region_research_v3",
            side_effect=controlled_research,
        ):
            self._call(
                "--bundle",
                str(self.manifest_path),
                "--region",
                "japan",
                "--reviewer-id",
                str(self.reviewer.id),
            )

    def test_prepare_execution_window_blocks_same_batch_commit(self):
        from concurrent.futures import ThreadPoolExecutor
        from threading import Event
        from unittest import mock

        from django.core.management.base import CommandError

        from stable.services import (
            p0_horse_completion_commit as commit_module,
            p0_horse_completion_prepare as prepare_module,
        )
        from stable.services.p0_horse_completion_batch import P0HorseBatchError

        prepare_entered = Event()
        release_prepare = Event()
        commit_started = Event()
        commit_entered = Event()

        def controlled_prepare(*_args, **_kwargs):
            prepare_entered.set()
            self.assertTrue(release_prepare.wait(timeout=5))
            raise P0HorseBatchError("stop controlled prepare")

        def controlled_commit(*_args, **_kwargs):
            commit_entered.set()
            return {"status": "commit-entered-after-prepare"}

        def run_commit():
            commit_started.set()
            return commit_module.commit_p0_horse_batch_region(
                self.manifest_path,
                region="japan",
                reviewer=self.reviewer,
                approved_by="human-approver",
                release_candidate_sha256="a" * 64,
                state_dir=self.state_dir,
                confirm_reviewed_artifact=True,
            )

        with mock.patch.object(
            prepare_module,
            "prepare_p0_horse_batch",
            side_effect=controlled_prepare,
        ), mock.patch.object(
            commit_module,
            "_commit_p0_horse_batch_region_locked",
            side_effect=controlled_commit,
        ):
            with ThreadPoolExecutor(max_workers=2) as executor:
                preparing = executor.submit(
                    self._call,
                    "--prepare",
                    str(self.manifest_path),
                    "--expected-sha256",
                    self.approved["batch_sha256"],
                )
                self.assertTrue(prepare_entered.wait(timeout=5))
                committing = executor.submit(run_commit)
                self.assertTrue(commit_started.wait(timeout=5))
                self.assertFalse(commit_entered.wait(timeout=0.2))
                release_prepare.set()
                with self.assertRaisesRegex(CommandError, "controlled prepare"):
                    preparing.result(timeout=5)
                self.assertEqual(
                    committing.result(timeout=5)["status"],
                    "commit-entered-after-prepare",
                )

    def test_prepare_release_rejects_symlink_and_non_regular_snapshot_targets(self):
        from django.core.management.base import CommandError

        from stable.services.p0_horse_completion_batch import BatchRunState

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        batch_dir = self.manifest_path.parent
        bundle = BatchRunState.read(batch_dir).artifacts["bundle:japan"]
        target = (
            batch_dir
            / "approval"
            / "input_snapshots"
            / f"research_v3_{bundle['research_sha256']}.json"
        )
        target.parent.mkdir(parents=True, exist_ok=True)
        ledger_path = batch_dir / "approvals_ledger.jsonl"
        ledger_before = ledger_path.read_bytes()
        state_before = (batch_dir / "state.json").read_bytes()

        target.symlink_to(Path(bundle["research_path"]))
        with self.assertRaisesRegex(CommandError, "regular file"):
            self._prepare_release_candidate()
        target.unlink()
        target.mkdir()
        with self.assertRaisesRegex(CommandError, "regular file"):
            self._prepare_release_candidate()

        self.assertEqual(ledger_path.read_bytes(), ledger_before)
        self.assertEqual((batch_dir / "state.json").read_bytes(), state_before)
        self.assertFalse(
            list((batch_dir / "approval").glob("release_candidate_japan_*.json"))
        )

    def _assert_prepare_release_rejects_stale_bundle_declaration(
        self, replaced_input: str
    ):
        from django.core.management.base import CommandError

        from stable.services.p0_horse_completion_batch import BatchRunState
        from stable.services.p0_horse_completion_research import _write_canonical

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        batch_dir = self.manifest_path.parent
        state = BatchRunState.read(batch_dir)
        bundle = state.artifacts["bundle:japan"]
        research_path = Path(bundle["research_path"])
        mapping_path = Path(bundle["mapping_path"])
        authority_path = Path(bundle["authority_path"])
        research = json.loads(research_path.read_text(encoding="utf-8"))
        mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
        authority = json.loads(authority_path.read_text(encoding="utf-8"))
        if replaced_input == "research":
            research["_replacement_marker"] = "research"
            research_sha = _write_canonical(research_path, research)
            mapping["research_v3_sha256"] = research_sha
            _write_canonical(mapping_path, mapping)
        elif replaced_input == "mapping":
            mapping["_replacement_marker"] = "mapping"
            _write_canonical(mapping_path, mapping)
        else:
            authority["_replacement_marker"] = "authority"
            authority_sha = _write_canonical(authority_path, authority)
            research["career_authority_review_application"][
                "review_artifact_sha256"
            ] = authority_sha
            research_sha = _write_canonical(research_path, research)
            mapping["research_v3_sha256"] = research_sha
            _write_canonical(mapping_path, mapping)

        ledger_path = batch_dir / "approvals_ledger.jsonl"
        ledger_before = ledger_path.read_bytes()
        state_before = (batch_dir / "state.json").read_bytes()
        approval_dir = batch_dir / "approval"
        evidence_before = {
            path.name
            for pattern in ("commit_artifact_japan_*", "release_candidate_japan_*")
            for path in approval_dir.glob(pattern)
        }
        with self.assertRaisesRegex(CommandError, "bundle .* SHA"):
            self._prepare_release_candidate()
        self.assertEqual((batch_dir / "state.json").read_bytes(), state_before)
        self.assertEqual(ledger_path.read_bytes(), ledger_before)
        self.assertEqual(
            {
                path.name
                for pattern in (
                    "commit_artifact_japan_*",
                    "release_candidate_japan_*",
                )
                for path in approval_dir.glob(pattern)
            },
            evidence_before,
        )

    def test_prepare_release_rejects_replaced_research_before_evidence(self):
        self._assert_prepare_release_rejects_stale_bundle_declaration("research")

    def test_prepare_release_rejects_replaced_mapping_before_evidence(self):
        self._assert_prepare_release_rejects_stale_bundle_declaration("mapping")

    def test_prepare_release_rejects_replaced_authority_before_evidence(self):
        self._assert_prepare_release_rejects_stale_bundle_declaration("authority")

    def test_prepare_network_flag_requires_setting(self):
        from django.core.management.base import CommandError

        with self.assertRaises(CommandError):
            self._call(
                "--prepare",
                str(self.manifest_path),
                "--expected-sha256",
                self.approved["batch_sha256"],
                "--allow-network",
            )

    def test_commit_rejects_stale_bundle_after_rerun(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )
        from stable.services.p0_horse_completion_commit import (
            commit_p0_horse_batch_region,
        )

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate = self._prepare_release_candidate()
        # simulate a rerun that republished the combined artifact (new bytes)
        combined = self.manifest_path.parent / "artifact" / "combined_candidates.jsonl"
        combined.write_text(combined.read_text(encoding="utf-8") + "\n", encoding="utf-8")
        with self.assertRaises(P0HorseBatchError) as ctx:
            commit_p0_horse_batch_region(
                self.manifest_path,
                region="japan",
                reviewer=self.reviewer,
                approved_by="human-approver",
                release_candidate_sha256=candidate[
                    "release_candidate_sha256"
                ],
                state_dir=self.state_dir,
                confirm_reviewed_artifact=True,
            )
        self.assertIn("stale", str(ctx.exception))

    def test_commit_reads_combined_only_after_entering_serial_window(self):
        from concurrent.futures import ThreadPoolExecutor
        from contextlib import contextmanager
        from threading import Event
        from unittest import mock

        from stable.services import p0_horse_completion_commit as commit_module
        from stable.services.p0_horse_completion_batch import P0HorseBatchError

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate = self._prepare_release_candidate()
        combined = (
            self.manifest_path.parent
            / "artifact"
            / "combined_candidates.jsonl"
        )

        lock_entered = Event()
        release_lock = Event()

        @contextmanager
        def controlled_serial_window(_state_dir):
            lock_entered.set()
            self.assertTrue(release_lock.wait(timeout=5))
            yield

        with mock.patch.object(
            commit_module,
            "_serial_window",
            side_effect=controlled_serial_window,
        ), mock.patch.object(
            commit_module,
            "_artifact_was_committed",
            return_value=False,
        ), mock.patch.object(
            commit_module,
            "build_region_release_manifest",
            side_effect=P0HorseBatchError("formal release builder reached"),
        ) as release_builder:
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(
                    commit_module.commit_p0_horse_batch_region,
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=candidate[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
                self.assertTrue(lock_entered.wait(timeout=5))
                combined.write_text(
                    combined.read_text(encoding="utf-8") + "\n",
                    encoding="utf-8",
                )
                release_lock.set()
                error = future.exception(timeout=5)
        self.assertIsInstance(error, P0HorseBatchError)
        self.assertRegex(str(error), "stale|drifted")
        release_builder.assert_not_called()

    def test_post_commit_checkpoint_merges_concurrent_bundle_state(self):
        from concurrent.futures import ThreadPoolExecutor
        from threading import Event
        from unittest import mock

        from stable.services import p0_horse_completion_commit as commit_module
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            batch_serial_window,
        )

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate = self._prepare_release_candidate()
        real_dry_run = commit_module.dry_run_reviewed_p0_completion_artifact
        calls = {"count": 0}
        writer_entered = Event()
        writer_finished = Event()

        def concurrent_bundle_writer():
            writer_entered.set()
            with batch_serial_window(self.state_dir):
                latest = BatchRunState.read(self.manifest_path.parent)
                bundle = dict(latest.artifacts["bundle:japan"])
                bundle["concurrent_writer_marker"] = "preserve-me"
                latest.artifacts["bundle:japan"] = bundle
                latest.write()
            writer_finished.set()

        def pause_after_database_commit(**kwargs):
            calls["count"] += 1
            report = real_dry_run(**kwargs)
            if calls["count"] == 2:
                with ThreadPoolExecutor(max_workers=1) as executor:
                    future = executor.submit(concurrent_bundle_writer)
                    self.assertTrue(writer_entered.wait(timeout=5))
                    self.assertTrue(writer_finished.wait(timeout=5))
                    future.result(timeout=5)
            return report

        with mock.patch.object(
            commit_module,
            "dry_run_reviewed_p0_completion_artifact",
            side_effect=pause_after_database_commit,
        ):
            committed = commit_module.commit_p0_horse_batch_region(
                self.manifest_path,
                region="japan",
                reviewer=self.reviewer,
                approved_by="human-approver",
                release_candidate_sha256=candidate[
                    "release_candidate_sha256"
                ],
                state_dir=self.state_dir,
                confirm_reviewed_artifact=True,
            )
        state = BatchRunState.read(self.manifest_path.parent)
        self.assertEqual(
            state.artifacts["bundle:japan"]["concurrent_writer_marker"],
            "preserve-me",
        )
        self.assertIn("commit:japan", state.artifacts)
        self.assertIn("publish:japan", state.artifacts)
        self.assertTrue(committed["idempotent_verification"]["passed"])

    def test_execution_lock_serializes_two_same_batch_commits(self):
        from concurrent.futures import ThreadPoolExecutor
        from threading import Event
        from unittest import mock

        from stable.services import p0_horse_completion_commit as commit_module

        first_entered = Event()
        release_first = Event()
        second_started = Event()
        second_entered = Event()
        order = []

        def controlled_commit(*args, **kwargs):
            candidate_sha = kwargs["release_candidate_sha256"]
            order.append(candidate_sha)
            if candidate_sha == "a" * 64:
                first_entered.set()
                self.assertTrue(release_first.wait(timeout=5))
            else:
                second_entered.set()
            return {"candidate": candidate_sha}

        def run_second():
            second_started.set()
            return commit_module.commit_p0_horse_batch_region(
                self.manifest_path,
                region="japan",
                reviewer=self.reviewer,
                approved_by="human-approver",
                release_candidate_sha256="b" * 64,
                state_dir=self.state_dir,
                confirm_reviewed_artifact=True,
            )

        with mock.patch.object(
            commit_module,
            "_commit_p0_horse_batch_region_locked",
            side_effect=controlled_commit,
        ):
            with ThreadPoolExecutor(max_workers=2) as executor:
                first = executor.submit(
                    commit_module.commit_p0_horse_batch_region,
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256="a" * 64,
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
                self.assertTrue(first_entered.wait(timeout=5))
                second = executor.submit(run_second)
                self.assertTrue(second_started.wait(timeout=5))
                self.assertFalse(second_entered.wait(timeout=0.2))
                release_first.set()
                first.result(timeout=5)
                second.result(timeout=5)
        self.assertEqual(order, ["a" * 64, "b" * 64])

    def test_prepare_release_waits_for_commit_db_window_then_rejects_committed(self):
        from concurrent.futures import ThreadPoolExecutor
        from contextlib import contextmanager
        from threading import Event, get_ident
        from unittest import mock

        from stable.services import p0_horse_completion_commit as commit_module
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate = self._prepare_release_candidate()
        batch_dir = self.manifest_path.parent
        ledger_path = batch_dir / "approvals_ledger.jsonl"
        candidate_state = BatchRunState.read(batch_dir)
        candidate_history = candidate_state.artifacts[
            "release_candidate:japan:"
            + candidate["release_candidate_sha256"]
        ]
        frozen_artifact = json.loads(
            Path(candidate_history["artifact_path"]).read_text(
                encoding="utf-8"
            )
        )
        frozen_publish_scope = json.loads(
            Path(candidate["release_candidate_path"]).read_text(
                encoding="utf-8"
            )
        )["auto_first_publish_scope"]
        db_window_entered = Event()
        release_db_window = Event()
        commit_body_finished = Event()
        release_execution_window = Event()
        prepare_started = Event()
        prepare_reached_bundle = Event()
        prepare_finished = Event()
        commit_thread_id = {"value": None}
        real_execution_window = commit_module.batch_execution_window

        @contextmanager
        def controlled_execution_window(path):
            with real_execution_window(path):
                try:
                    yield
                finally:
                    if (
                        get_ident() == commit_thread_id["value"]
                        and db_window_entered.is_set()
                    ):
                        commit_body_finished.set()
                        self.assertTrue(
                            release_execution_window.wait(timeout=5)
                        )

        zero_report = {
            "planned_profile_creates": 0,
            "planned_profile_updates": 0,
            "planned_race_record_creates": 0,
            "planned_race_record_updates": 0,
            "planned_module_audits": 0,
        }

        def controlled_db_commit(**_kwargs):
            db_window_entered.set()
            self.assertTrue(release_db_window.wait(timeout=5))
            return {"status": "committed"}

        def controlled_publish(
            _manifest,
            *,
            batch_dir,
            state_dir,
            region,
            **_kwargs,
        ):
            report = {
                "published": 0,
                "skipped_already_published": 0,
                "blocked": 0,
                "blocked_reasons": {},
                "published_profile_ids": [],
                "errors": [],
                "profile_ids": [],
                "frozen_exclusions": [],
                "frozen_exclusion_counts": {},
            }
            with commit_module._serial_window(state_dir):
                state = BatchRunState.read(batch_dir)
                stage = f"publish:{region}"
                state.artifacts[stage] = report
                if stage not in state.completed_stages:
                    state.completed_stages.append(stage)
                state.write()
            return report

        def reject_if_prepare_crosses_execution_boundary(state, region):
            prepare_reached_bundle.set()
            raise P0HorseBatchError(
                "prepare-release entered before commit execution exited"
            )

        def run_commit():
            commit_thread_id["value"] = get_ident()
            return commit_module.commit_p0_horse_batch_region(
                self.manifest_path,
                region="japan",
                reviewer=self.reviewer,
                approved_by="human-approver",
                release_candidate_sha256=candidate[
                    "release_candidate_sha256"
                ],
                state_dir=self.state_dir,
                confirm_reviewed_artifact=True,
            )

        def run_prepare_release():
            prepare_started.set()
            try:
                return commit_module.prepare_p0_horse_batch_release_candidate(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    state_dir=self.state_dir,
                )
            finally:
                prepare_finished.set()

        completion_runs = mock.Mock()
        completion_runs.order_by.return_value.first.return_value = None
        with mock.patch.object(
            commit_module,
            "batch_execution_window",
            side_effect=controlled_execution_window,
        ), mock.patch.object(
            commit_module,
            "prepare_reviewed_p0_completion_artifact",
            return_value=frozen_artifact,
        ), mock.patch.object(
            commit_module,
            "_build_auto_first_publish_scope",
            return_value=frozen_publish_scope,
        ), mock.patch.object(
            commit_module,
            "_artifact_was_committed",
            return_value=False,
        ), mock.patch.object(
            commit_module,
            "dry_run_reviewed_p0_completion_artifact",
            return_value=zero_report,
        ), mock.patch.object(
            commit_module,
            "commit_reviewed_p0_completion_artifact",
            side_effect=controlled_db_commit,
        ), mock.patch.object(
            commit_module,
            "_run_region_publish",
            side_effect=controlled_publish,
        ), mock.patch.object(
            commit_module,
            "_region_bundle",
            side_effect=reject_if_prepare_crosses_execution_boundary,
        ), mock.patch(
            "stable.models.HorseProfileCompletionRun.objects.filter",
            return_value=completion_runs,
        ):
            with ThreadPoolExecutor(max_workers=2) as executor:
                committing = executor.submit(run_commit)
                if not db_window_entered.wait(timeout=5):
                    release_execution_window.set()
                    committing.result(timeout=5)
                    self.fail("commit did not reach database window")
                preparing = executor.submit(run_prepare_release)
                self.assertTrue(prepare_started.wait(timeout=5))
                crossed_during_db = prepare_reached_bundle.wait(timeout=0.2)
                release_db_window.set()
                self.assertTrue(commit_body_finished.wait(timeout=5))
                finished_before_commit_exit = prepare_finished.is_set()
                state_after_commit = (batch_dir / "state.json").read_bytes()
                ledger_after_commit = ledger_path.read_bytes()
                candidate_after_commit = {
                    path.name: path.read_bytes()
                    for path in (batch_dir / "approval").glob(
                        "release_candidate_japan_*.json"
                    )
                }
                release_execution_window.set()
                committing.result(timeout=5)
                with self.assertRaisesRegex(P0HorseBatchError, "committed"):
                    preparing.result(timeout=5)

        self.assertFalse(crossed_during_db)
        self.assertFalse(finished_before_commit_exit)
        self.assertEqual((batch_dir / "state.json").read_bytes(), state_after_commit)
        self.assertEqual(ledger_path.read_bytes(), ledger_after_commit)
        self.assertEqual(
            {
                path.name: path.read_bytes()
                for path in (batch_dir / "approval").glob(
                    "release_candidate_japan_*.json"
                )
            },
            candidate_after_commit,
        )

    def test_prepare_release_waits_for_abandon_exit_then_rejects_without_mutation(
        self,
    ):
        from concurrent.futures import ThreadPoolExecutor
        from contextlib import contextmanager
        from threading import Event, get_ident
        from unittest import mock

        from stable.services import (
            p0_horse_completion_batch as batch_module,
            p0_horse_completion_commit as commit_module,
        )
        from stable.services.p0_horse_completion_batch import P0HorseBatchError

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        self._prepare_release_candidate()
        batch_dir = self.manifest_path.parent
        ledger_path = batch_dir / "approvals_ledger.jsonl"
        abandon_body_finished = Event()
        release_abandon_execution = Event()
        prepare_started = Event()
        prepare_finished = Event()
        abandon_thread_id = {"value": None}
        real_execution_window = batch_module.batch_execution_window

        @contextmanager
        def controlled_abandon_execution(path):
            with real_execution_window(path):
                try:
                    yield
                finally:
                    if get_ident() == abandon_thread_id["value"]:
                        abandon_body_finished.set()
                        self.assertTrue(
                            release_abandon_execution.wait(timeout=5)
                        )

        def run_abandon():
            abandon_thread_id["value"] = get_ident()
            return self._call(
                "--abandon",
                str(self.manifest_path),
                "--note",
                "prepare-release 并发终止测试",
            )

        def run_prepare_release():
            prepare_started.set()
            try:
                return commit_module.prepare_p0_horse_batch_release_candidate(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    state_dir=self.state_dir,
                )
            finally:
                prepare_finished.set()

        with mock.patch.object(
            batch_module,
            "batch_execution_window",
            side_effect=controlled_abandon_execution,
        ), mock.patch.object(
            batch_module,
            "ensure_batch_can_be_abandoned",
            return_value=None,
        ):
            with ThreadPoolExecutor(max_workers=2) as executor:
                abandoning = executor.submit(run_abandon)
                self.assertTrue(abandon_body_finished.wait(timeout=5))
                state_after_abandon = (batch_dir / "state.json").read_bytes()
                manifest_after_abandon = self.manifest_path.read_bytes()
                ledger_after_abandon = ledger_path.read_bytes()
                candidate_after_abandon = {
                    path.name: path.read_bytes()
                    for path in (batch_dir / "approval").glob(
                        "release_candidate_japan_*.json"
                    )
                }
                preparing = executor.submit(run_prepare_release)
                self.assertTrue(prepare_started.wait(timeout=5))
                finished_before_abandon_exit = prepare_finished.wait(
                    timeout=0.2
                )
                release_abandon_execution.set()
                abandoning.result(timeout=5)
                with self.assertRaisesRegex(P0HorseBatchError, "abandoned"):
                    preparing.result(timeout=5)

        self.assertFalse(finished_before_abandon_exit)
        self.assertEqual((batch_dir / "state.json").read_bytes(), state_after_abandon)
        self.assertEqual(self.manifest_path.read_bytes(), manifest_after_abandon)
        self.assertEqual(ledger_path.read_bytes(), ledger_after_abandon)
        self.assertEqual(
            {
                path.name: path.read_bytes()
                for path in (batch_dir / "approval").glob(
                    "release_candidate_japan_*.json"
                )
            },
            candidate_after_abandon,
        )

    def test_abandon_waits_for_execution_window_then_stops_batch(self):
        from concurrent.futures import ThreadPoolExecutor
        from threading import Event
        from unittest import mock

        from stable.services import p0_horse_completion_commit as commit_module
        from stable.services.p0_horse_completion_batch import BatchRunState

        commit_entered = Event()
        release_commit = Event()
        abandon_started = Event()
        abandon_finished = Event()

        def controlled_commit(*args, **kwargs):
            commit_entered.set()
            self.assertTrue(release_commit.wait(timeout=5))
            return {"status": "finished-before-abandon"}

        def run_abandon():
            abandon_started.set()
            result = self._call(
                "--abandon",
                str(self.manifest_path),
                "--note",
                "并发终止测试",
            )
            abandon_finished.set()
            return result

        with mock.patch.object(
            commit_module,
            "_commit_p0_horse_batch_region_locked",
            side_effect=controlled_commit,
        ):
            with ThreadPoolExecutor(max_workers=2) as executor:
                commit = executor.submit(
                    commit_module.commit_p0_horse_batch_region,
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256="a" * 64,
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
                self.assertTrue(commit_entered.wait(timeout=5))
                abandon = executor.submit(run_abandon)
                self.assertTrue(abandon_started.wait(timeout=5))
                self.assertFalse(abandon_finished.wait(timeout=0.2))
                release_commit.set()
                commit.result(timeout=5)
                result = abandon.result(timeout=5)
        self.assertEqual(result["status"], "abandoned")
        self.assertEqual(
            BatchRunState.read(self.manifest_path.parent).stage,
            "abandoned",
        )

    def test_commit_after_successful_abandon_has_no_db_or_publish_effects(self):
        from unittest import mock

        from stable.services import (
            horse_profile_publish,
            p0_horse_completion_commit as commit_module,
        )
        from stable.services.p0_horse_completion_batch import P0HorseBatchError

        self._call(
            "--abandon",
            str(self.manifest_path),
            "--note",
            "终止后禁止提交",
        )
        with mock.patch.object(
            commit_module,
            "commit_reviewed_p0_completion_artifact",
        ) as db_commit, mock.patch.object(
            horse_profile_publish,
            "auto_publish_profiles",
        ) as publish:
            with self.assertRaisesRegex(P0HorseBatchError, "abandoned"):
                commit_module.commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256="a" * 64,
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
        db_commit.assert_not_called()
        publish.assert_not_called()

    def test_commit_marks_manifest_committed_and_leaves_inflight(self):
        self.test_full_pipeline_via_command()
        manifest = json.loads(self.manifest_path.read_text(encoding="utf-8"))
        self.assertEqual(manifest["status"], "committed")
        selected = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        names = {horse["horse_name"] for horse in selected["horses"]}
        self.assertNotIn("FOREVER TEST", names)

    def test_repeated_commit_reuses_v2_release_manifest_and_approval(self):
        self.test_full_pipeline_via_command()
        from stable.services.p0_horse_completion_batch import BatchRunState

        state = BatchRunState.read(self.manifest_path.parent)
        candidate_sha = state.artifacts["release_candidate:japan"]["sha256"]
        first_release_sha = state.artifacts["commit:japan"]["release_sha256"]
        release_path = Path(state.artifacts["commit:japan"]["release_path"])
        first_bytes = release_path.read_bytes()
        second = self._call(
            "--commit",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
            "--approved-by",
            "human-approver",
            "--release-candidate-sha256",
            candidate_sha,
            "--confirm-reviewed-artifact",
        )
        self.assertEqual(second["release_sha256"], first_release_sha)
        self.assertEqual(release_path.read_bytes(), first_bytes)
        ledger = [
            json.loads(line)
            for line in (
                self.manifest_path.parent / "approvals_ledger.jsonl"
            ).read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        self.assertEqual(
            sum(entry.get("event") == "release_approved" for entry in ledger),
            1,
        )

    def test_release_recovery_rejects_tampered_bytes_without_reapproval(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )
        from stable.services.p0_horse_completion_commit import (
            commit_p0_horse_batch_region,
        )

        self.test_full_pipeline_via_command()
        state = BatchRunState.read(self.manifest_path.parent)
        commit_state = state.artifacts["commit:japan"]
        release_path = Path(commit_state["release_path"])
        candidate_sha = commit_state["release_candidate_sha256"]
        ledger_path = self.manifest_path.parent / "approvals_ledger.jsonl"
        kept = [
            line
            for line in ledger_path.read_text(encoding="utf-8").splitlines()
            if not (
                json.loads(line).get("event") == "release_approved"
                and json.loads(line).get("release_manifest_sha256")
                == commit_state["release_sha256"]
            )
        ]
        ledger_path.write_text("\n".join(kept) + "\n", encoding="utf-8")
        original = json.loads(release_path.read_text(encoding="utf-8"))
        ledger_before = ledger_path.read_bytes()
        mutations = {
            "approved_at": lambda value: value.update(
                {"approved_at": "2026-07-24T00:00:00"}
            ),
            "decision_reference": lambda value: value.update(
                {"decision_reference": "tampered"}
            ),
            "approvals_ledger_path": lambda value: value.update(
                {"approvals_ledger_path": "/tmp/forged-ledger.jsonl"}
            ),
            "extra_key": lambda value: value.update(
                {"unexpected_key": "must fail"}
            ),
        }
        for name, mutate in mutations.items():
            with self.subTest(field=name):
                tampered = json.loads(json.dumps(original))
                mutate(tampered)
                release_path.write_text(
                    json.dumps(
                        tampered, ensure_ascii=False, sort_keys=True
                    )
                    + "\n",
                    encoding="utf-8",
                )
                with self.assertRaisesRegex(
                    P0HorseBatchError, "filename SHA|signing contract"
                ):
                    commit_p0_horse_batch_region(
                        self.manifest_path,
                        region="japan",
                        reviewer=self.reviewer,
                        approved_by="human-approver",
                        release_candidate_sha256=candidate_sha,
                        state_dir=self.state_dir,
                        confirm_reviewed_artifact=True,
                    )
                self.assertEqual(ledger_path.read_bytes(), ledger_before)

    def test_release_recovery_rejects_symlink_and_nonregular_file(self):
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )
        from stable.services.p0_horse_completion_commit import (
            commit_p0_horse_batch_region,
        )

        self.test_full_pipeline_via_command()
        state = BatchRunState.read(self.manifest_path.parent)
        commit_state = state.artifacts["commit:japan"]
        release_path = Path(commit_state["release_path"])
        backup = release_path.with_suffix(".backup")
        release_path.replace(backup)
        release_path.mkdir()
        ledger_path = self.manifest_path.parent / "approvals_ledger.jsonl"
        ledger_before = ledger_path.read_bytes()
        with self.assertRaisesRegex(P0HorseBatchError, "regular file"):
            commit_p0_horse_batch_region(
                self.manifest_path,
                region="japan",
                reviewer=self.reviewer,
                approved_by="human-approver",
                release_candidate_sha256=commit_state[
                    "release_candidate_sha256"
                ],
                state_dir=self.state_dir,
                confirm_reviewed_artifact=True,
            )
        self.assertEqual(ledger_path.read_bytes(), ledger_before)
        release_path.rmdir()
        release_path.symlink_to(backup)
        with self.assertRaisesRegex(P0HorseBatchError, "regular file"):
            commit_p0_horse_batch_region(
                self.manifest_path,
                region="japan",
                reviewer=self.reviewer,
                approved_by="human-approver",
                release_candidate_sha256=commit_state[
                    "release_candidate_sha256"
                ],
                state_dir=self.state_dir,
                confirm_reviewed_artifact=True,
            )
        self.assertEqual(ledger_path.read_bytes(), ledger_before)

    def test_new_candidate_supersedes_approved_uncommitted_candidate_immutably(self):
        from unittest import mock

        from django.utils import timezone

        from stable.services import p0_horse_completion_commit as commit_module
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate_a = self._prepare_release_candidate()
        with mock.patch.object(
            commit_module,
            "dry_run_reviewed_p0_completion_artifact",
            side_effect=P0HorseBatchError("injected before database commit"),
        ):
            with self.assertRaisesRegex(P0HorseBatchError, "injected"):
                commit_module.commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=candidate_a[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
        state = BatchRunState.read(self.manifest_path.parent)
        history_a = state.artifacts[
            "release_candidate:japan:"
            + candidate_a["release_candidate_sha256"]
        ]
        a_paths = {
            "artifact": Path(history_a["artifact_path"]),
            "candidate": Path(history_a["path"]),
            "release": Path(history_a["release_path"]),
        }
        a_bytes = {key: path.read_bytes() for key, path in a_paths.items()}
        self.assertIn(candidate_a["artifact_sha256"], a_paths["artifact"].name)
        self.assertIn(
            candidate_a["release_candidate_sha256"],
            a_paths["candidate"].name,
        )
        self.assertIn(history_a["release_sha256"], a_paths["release"].name)

        self.profile.hidden_at = timezone.now()
        self.profile.review_status = "hidden"
        self.profile.save(update_fields=["hidden_at", "review_status"])
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate_b = self._prepare_release_candidate()
        self.assertNotEqual(
            candidate_b["release_candidate_sha256"],
            candidate_a["release_candidate_sha256"],
        )
        committed_b = self._call(
            "--commit",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
            "--approved-by",
            "human-approver",
            "--release-candidate-sha256",
            candidate_b["release_candidate_sha256"],
            "--confirm-reviewed-artifact",
        )
        self.assertTrue(committed_b["idempotent_verification"]["passed"])
        for key, path in a_paths.items():
            self.assertEqual(path.read_bytes(), a_bytes[key], key)
        state = BatchRunState.read(self.manifest_path.parent)
        history_b = state.artifacts[
            "release_candidate:japan:"
            + candidate_b["release_candidate_sha256"]
        ]
        self.assertNotEqual(history_a["release_path"], history_b["release_path"])
        ledger = [
            json.loads(line)
            for line in (
                self.manifest_path.parent / "approvals_ledger.jsonl"
            ).read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        superseded = [
            entry for entry in ledger if entry.get("event") == "release_superseded"
        ]
        self.assertEqual(len(superseded), 1)
        self.assertEqual(
            superseded[0]["old_release_candidate_sha256"],
            candidate_a["release_candidate_sha256"],
        )
        self.assertEqual(
            superseded[0]["new_release_candidate_sha256"],
            candidate_b["release_candidate_sha256"],
        )

    def test_supersede_recovery_orders_old_invalidation_before_new_approval(self):
        from unittest import mock

        from django.utils import timezone

        from stable.services import (
            p0_horse_completion_commit as commit_module,
            p0_horse_completion_research as research_module,
        )
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate_a = self._prepare_release_candidate()
        with mock.patch.object(
            commit_module,
            "dry_run_reviewed_p0_completion_artifact",
            side_effect=P0HorseBatchError("stop A before database"),
        ):
            with self.assertRaises(P0HorseBatchError):
                commit_module.commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=candidate_a[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )

        self.profile.review_status = "hidden"
        self.profile.hidden_at = timezone.now()
        self.profile.save(update_fields=["review_status", "hidden_at"])
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate_b = self._prepare_release_candidate()
        real_append = research_module._append_approvals_ledger

        def crash_before_new_approval(batch_dir, entry):
            if entry.get("event") == "release_approved":
                raise P0HorseBatchError("crash before B approval")
            return real_append(batch_dir, entry)

        with mock.patch.object(
            research_module,
            "_append_approvals_ledger",
            side_effect=crash_before_new_approval,
        ):
            with self.assertRaisesRegex(P0HorseBatchError, "B approval"):
                commit_module.commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=candidate_b[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )

        state = BatchRunState.read(self.manifest_path.parent)
        pending_b = state.artifacts[
            "release_candidate:japan:"
            + candidate_b["release_candidate_sha256"]
        ]["pending_release"]
        b_release_sha = pending_b["sha256"]
        ledger_path = self.manifest_path.parent / "approvals_ledger.jsonl"
        ledger = [
            json.loads(line)
            for line in ledger_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        self.assertTrue(
            any(entry.get("event") == "release_superseded" for entry in ledger)
        )
        self.assertFalse(
            any(
                entry.get("event") == "release_approved"
                and entry.get("release_manifest_sha256") == b_release_sha
                for entry in ledger
            )
        )

        with mock.patch.object(
            commit_module,
            "dry_run_reviewed_p0_completion_artifact",
            side_effect=P0HorseBatchError("stop B after approval"),
        ):
            with self.assertRaisesRegex(P0HorseBatchError, "after approval"):
                commit_module.commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=candidate_b[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
        ledger = [
            json.loads(line)
            for line in ledger_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        supersede_index = next(
            index
            for index, entry in enumerate(ledger)
            if entry.get("event") == "release_superseded"
        )
        approval_indexes = [
            index
            for index, entry in enumerate(ledger)
            if entry.get("event") == "release_approved"
            and entry.get("release_manifest_sha256") == b_release_sha
        ]
        self.assertEqual(len(approval_indexes), 1)
        self.assertLess(supersede_index, approval_indexes[0])

    def test_committed_candidate_blocks_new_candidate_evidence(self):
        from django.core.management.base import CommandError
        from django.utils import timezone

        self.test_full_pipeline_via_command()
        ledger_path = self.manifest_path.parent / "approvals_ledger.jsonl"
        self.profile.hidden_at = timezone.now()
        self.profile.review_status = "hidden"
        self.profile.save(update_fields=["hidden_at", "review_status"])
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        from stable.services.p0_horse_completion_batch import BatchRunState

        state = BatchRunState.read(self.manifest_path.parent)
        current_sha = state.artifacts["release_candidate:japan"]["sha256"]
        state.artifacts.pop("release_candidate:japan")
        state.artifacts.pop(f"release_candidate:japan:{current_sha}")
        state.write()
        before = ledger_path.read_bytes()
        with self.assertRaisesRegex(CommandError, "already committed"):
            self._prepare_release_candidate()
        self.assertEqual(ledger_path.read_bytes(), before)

    def _assert_uncommitted_approved_candidate_rechecks_scope(self, mutate):
        from unittest import mock

        from stable.models import HorseProfileCompletionRun
        from stable.services import p0_horse_completion_commit as commit_module
        from stable.services.p0_horse_completion_batch import P0HorseBatchError

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate = self._prepare_release_candidate()
        with mock.patch.object(
            commit_module,
            "dry_run_reviewed_p0_completion_artifact",
            side_effect=P0HorseBatchError(
                "injected after release before database"
            ),
        ):
            with self.assertRaises(P0HorseBatchError):
                commit_module.commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=candidate[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
        self.assertFalse(
            HorseProfileCompletionRun.objects.filter(status="committed").exists()
        )
        mutate()
        self.profile.save()
        zero_report = {
            "planned_profile_creates": 0,
            "planned_profile_updates": 0,
            "planned_race_record_creates": 0,
            "planned_race_record_updates": 0,
            "planned_module_audits": 0,
        }
        with mock.patch.object(
            commit_module,
            "commit_reviewed_p0_completion_artifact",
        ) as commit_mock, mock.patch.object(
            commit_module,
            "dry_run_reviewed_p0_completion_artifact",
            return_value=zero_report,
        ):
            with self.assertRaisesRegex(P0HorseBatchError, "bindings drifted"):
                commit_module.commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=candidate[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
        commit_mock.assert_not_called()

    def test_uncommitted_approved_candidate_rechecks_hidden_drift(self):
        from django.utils import timezone

        def mutate():
            self.profile.hidden_at = timezone.now()
            self.profile.review_status = "hidden"

        self._assert_uncommitted_approved_candidate_rechecks_scope(mutate)

    def test_uncommitted_approved_candidate_rechecks_review_status_drift(self):
        def mutate():
            self.profile.review_status = "published"

        self._assert_uncommitted_approved_candidate_rechecks_scope(mutate)

    def test_uncommitted_approved_candidate_rechecks_manual_lock_drift(self):
        def mutate():
            self.profile.manual_lock_flags = {"auto_publish_blocked": True}

        self._assert_uncommitted_approved_candidate_rechecks_scope(mutate)

    def test_abandon_command_marks_manifest_abandoned(self):
        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        result = self._call(
            "--abandon",
            str(self.manifest_path),
            "--note",
            "批次构成有误",
        )
        self.assertEqual(result["status"], "abandoned")
        from stable.services.p0_horse_completion_batch import P0HorseBatchError
        from stable.services.p0_horse_completion_prepare import prepare_p0_horse_batch

        with self.assertRaises(P0HorseBatchError):
            prepare_p0_horse_batch(
                self.manifest_path,
                expected_sha256=self.approved["batch_sha256"],
                allow_network=False,
                cache_dir=self.cache_dir,
            )

    def test_prepare_blocked_payload_writes_blocker_pool(self):
        other = self._profile("无缓存马")
        self._p0_source(other)
        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        manifest_path = write_batch_manifest(manifest, state_dir=self.state_dir / "batch-pool")
        approved = approve_batch_manifest(manifest_path, reviewer="reviewer-a")
        from stable.services.p0_horse_completion_prepare import prepare_p0_horse_batch

        prepare_p0_horse_batch(
            manifest_path,
            expected_sha256=approved["batch_sha256"],
            allow_network=False,
            cache_dir=self.cache_dir,
        )
        pool_path = manifest_path.parent / "blocker_pool.jsonl"
        entries = [
            json.loads(line)
            for line in pool_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0]["horse_name"], "无缓存马")
        self.assertEqual(entries[0]["reason"], "blocked_at_prepare")

    def test_abandon_without_prepare_works(self):
        result = self._call(
            "--abandon",
            str(self.manifest_path),
            "--note",
            "选错批次",
        )
        self.assertEqual(result["status"], "abandoned")

    def test_recommit_after_rebundle_uses_candidate_snapshots(self):
        self.test_full_pipeline_via_command()
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
        )
        from stable.services.p0_horse_completion_commit import (
            commit_p0_horse_batch_region,
        )

        state = BatchRunState.read(self.manifest_path.parent)
        artifact_path = Path(state.artifacts["commit:japan"]["artifact_path"])
        original_bytes = artifact_path.read_bytes()
        # region-current bundle 可变化；已落库候选必须使用自己的不可变输入恢复。
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        state = BatchRunState.read(self.manifest_path.parent)
        recorded_sha = state.artifacts["commit:japan"]["artifact_sha256"]
        candidate_sha = state.artifacts["release_candidate:japan"]["sha256"]
        recovered = commit_p0_horse_batch_region(
            self.manifest_path,
            region="japan",
            reviewer=self.reviewer,
            approved_by="human-approver",
            release_candidate_sha256=candidate_sha,
            state_dir=self.state_dir,
            confirm_reviewed_artifact=True,
        )
        self.assertTrue(recovered["idempotent_verification"]["passed"])
        self.assertEqual(artifact_path.read_bytes(), original_bytes)
        state = BatchRunState.read(self.manifest_path.parent)
        self.assertEqual(
            state.artifacts["commit:japan"]["artifact_sha256"], recorded_sha
        )


class P0HorseBatchNetkeibaPipelineTests(P0HorseBatchTestBase):
    """End-to-end: netkeiba ID-direct path through the rolling batch pipeline
    (add-netkeiba-horse-client). The candidate has a netkeiba identity key
    and no four-field data — provider-bound identity must carry it through.
    """

    def setUp(self):
        import shutil
        import tempfile

        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        root = Path(self._tmp.name)
        self.state_dir = root / "batches"
        self.cache_dir = root / "cache"
        from django.contrib.auth import get_user_model

        self.reviewer = get_user_model().objects.create_user(
            username="p0-netkeiba-reviewer",
            password="unused",
            is_superuser=True,
            is_staff=True,
        )
        self.profile = self._profile(
            "ドラゴンウェルズ",
            source_refs={
                "horse_identity_keys": ["netkeiba:2022110137"],
                "horse_source_urls": ["https://db.netkeiba.com/horse/2022110137/"],
            },
        )
        self._p0_source(self.profile)
        from stable.services.p0_horse_completion_batch import (
            approve_batch_manifest,
            select_p0_horse_batch,
            write_batch_manifest,
        )

        manifest = select_p0_horse_batch(regions=[RacingRegion.JAPAN])
        horses = manifest.get("horses") or []
        self.assertEqual(len(horses), 1)
        self.assertEqual(horses[0]["source_namespace"], "netkeiba")
        self.manifest_path = write_batch_manifest(manifest, state_dir=self.state_dir)
        self.approved = approve_batch_manifest(self.manifest_path, reviewer="reviewer-a")
        from stable.services.p0_horse_completion_adapters import (
            p0_horse_completion_cache_path,
        )

        fixture = (
            Path(__file__).resolve().parent
            / "fixtures"
            / "p0_horse_completion"
            / "japan_netkeiba.json"
        )
        cache_path = p0_horse_completion_cache_path(
            self.cache_dir, f"profile:{self.profile.pk}"
        )
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(fixture, cache_path)

    def _call(self, *command_args) -> dict:
        from io import StringIO

        from django.core.management import call_command
        from django.test import override_settings

        out = StringIO()
        with override_settings(
            HORSE_PROFILE_COMPLETION_BATCH_STATE_DIR=str(self.state_dir),
            HORSE_PROFILE_COMPLETION_CACHE_DIR=str(self.cache_dir),
            HORSE_PROFILE_COMPLETION_REVIEW_OUTPUT_DIR=str(
                Path(self._tmp.name) / "review"
            ),
        ):
            call_command("p0_horse_completion_batch", *command_args, "--json", stdout=out)
        text = out.getvalue()
        marker = text.rfind("\n{")
        payload = text[marker + 1 :] if marker != -1 else text
        return json.loads(payload)

    def test_netkeiba_full_pipeline_with_auto_publish(self):
        prepared = self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self.assertEqual(prepared["totals"]["succeeded"], 1)
        bundled = self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        self.assertEqual(bundled["horse_count"], 1)
        candidate = self._call(
            "--prepare-release",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        committed = self._call(
            "--commit",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
            "--approved-by",
            "human-approver",
            "--release-candidate-sha256",
            candidate["release_candidate_sha256"],
            "--confirm-reviewed-artifact",
        )
        self.assertTrue(committed["idempotent_verification"]["passed"])
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.sire_text, "Frosted")
        self.assertEqual(self.profile.dam_text, "Little Dipper")
        self.assertEqual(str(self.profile.birth_date), "2022-03-26")
        self.assertEqual(self.profile.completeness_status, "complete_profile_full")
        # netkeiba key marked verified by the reviewed batch commit
        self.assertEqual(
            self.profile.source_refs.get("horse_identity_verified_keys"),
            ["netkeiba:2022110137"],
        )
        # auto first publish fired through the BASIC gate
        self.assertEqual(committed["auto_first_publish"]["published"], 1)
        self.assertEqual(self.profile.review_status, "published")


class P0HorseBatchAutoPublishTests(P0HorseBatchCommandPipelineTests):
    """Auto first publish hook after a region commit's verification passes."""

    def _commit_args(self):
        return (
            "--commit",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
            "--approved-by",
            "human-approver",
            "--release-candidate-sha256",
            self.release_candidate["release_candidate_sha256"],
            "--confirm-reviewed-artifact",
        )

    def _run_pipeline(self):
        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        self.release_candidate = self._prepare_release_candidate()
        return self._call(*self._commit_args())

    def test_auto_first_publish_after_verified_commit(self):
        from stable.models import HorseProfileCompletionRun, OperationLog
        from stable.services.p0_horse_completion_batch import BatchRunState

        committed = self._run_pipeline()
        publish = committed["auto_first_publish"]
        self.assertEqual(publish["published"], 1)
        self.assertEqual(publish["errors"], [])
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.review_status, "published")
        self.assertEqual(self.profile.published_by, self.reviewer)
        self.assertIsNotNone(self.profile.published_at)
        self.assertTrue(
            OperationLog.objects.filter(
                action_type="horse_profile_status_changed",
                target_id=str(self.profile.pk),
            ).exists()
        )
        ledger = (self.manifest_path.parent / "approvals_ledger.jsonl").read_text(
            encoding="utf-8"
        )
        self.assertIn('"auto_first_publish"', ledger)
        publish_event = next(
            entry
            for entry in (
                json.loads(line) for line in ledger.splitlines() if line.strip()
            )
            if entry.get("event") == "auto_first_publish"
        )
        self.assertEqual(
            publish_event["event_schema"],
            "p0_horse_auto_first_publish.v2",
        )
        self.assertIn("frozen_exclusions", publish_event)
        self.assertIn("frozen_exclusion_counts", publish_event)
        state = BatchRunState.read(self.manifest_path.parent)
        self.assertIn("publish:japan", state.artifacts)
        self.assertIn("publish:japan", state.completed_stages)
        run = HorseProfileCompletionRun.objects.get(id=committed["completion_run_id"])
        self.assertEqual(run.summary["auto_first_publish"]["published"], 1)

    def test_repeated_commit_reuses_completed_publish_after_manual_downgrade(self):
        from unittest import mock

        from stable.models import (
            HorseP0Source,
            HorseProfileCompletionRun,
            HorseProfileDataCandidate,
            HorseRaceRecord,
            OperationLog,
            TaskExecutionLog,
        )
        from stable.services import (
            horse_profile_publish,
            p0_horse_completion_commit as commit_module,
        )
        from stable.services.p0_horse_completion_batch import BatchRunState

        self._run_pipeline()
        batch_dir = self.manifest_path.parent
        state_before = BatchRunState.read(batch_dir)
        publish_before = json.loads(
            json.dumps(state_before.artifacts["publish:japan"])
        )
        ledger_path = batch_dir / "approvals_ledger.jsonl"
        ledger_before = ledger_path.read_bytes()

        self.profile.review_status = "ready"
        self.profile.published_at = None
        self.profile.published_by = None
        self.profile.save(
            update_fields=[
                "review_status",
                "published_at",
                "published_by",
            ]
        )
        counted_models = (
            HorseProfile,
            HorseP0Source,
            HorseRaceRecord,
            HorseProfileDataCandidate,
            HorseProfileCompletionRun,
            OperationLog,
            TaskExecutionLog,
        )
        database_counts_before = {
            model: model.objects.count() for model in counted_models
        }
        completion_runs_before = list(
            HorseProfileCompletionRun.objects.order_by("id").values(
                "id",
                "status",
                "parameters",
                "summary",
                "artifact_path",
                "updated_at",
            )
        )
        sources_before = list(
            HorseP0Source.objects.order_by("id").values()
        )
        with mock.patch.object(
            horse_profile_publish,
            "auto_publish_profiles",
            return_value={
                "published": 0,
                "skipped_already_published": 1,
                "blocked": 0,
                "blocked_reasons": {},
                "published_profile_ids": [],
                "errors": [],
            },
        ) as publish_mock, mock.patch.object(
            commit_module,
            "dry_run_reviewed_p0_completion_artifact",
            wraps=commit_module.dry_run_reviewed_p0_completion_artifact,
        ) as dry_run_mock, mock.patch.object(
            commit_module,
            "commit_reviewed_p0_completion_artifact",
            wraps=commit_module.commit_reviewed_p0_completion_artifact,
        ) as database_apply_mock, mock.patch.object(
            commit_module,
            "_run_region_publish",
            wraps=commit_module._run_region_publish,
        ) as region_publish_mock:
            repeated = commit_module.commit_p0_horse_batch_region(
                self.manifest_path,
                region="japan",
                reviewer=self.reviewer,
                approved_by="human-approver",
                release_candidate_sha256=self.release_candidate[
                    "release_candidate_sha256"
                ],
                state_dir=self.state_dir,
                confirm_reviewed_artifact=True,
            )
        publish_mock.assert_not_called()
        dry_run_mock.assert_not_called()
        database_apply_mock.assert_not_called()
        region_publish_mock.assert_not_called()
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.review_status, "ready")
        state_after = BatchRunState.read(batch_dir)
        self.assertEqual(
            state_after.artifacts["publish:japan"],
            publish_before,
        )
        self.assertEqual(
            state_after.completed_stages.count("publish:japan"),
            1,
        )
        self.assertEqual(ledger_path.read_bytes(), ledger_before)
        self.assertEqual(repeated["auto_first_publish"], publish_before)
        self.assertEqual(
            {model: model.objects.count() for model in counted_models},
            database_counts_before,
        )
        self.assertEqual(
            list(
                HorseProfileCompletionRun.objects.order_by("id").values(
                    "id",
                    "status",
                    "parameters",
                    "summary",
                    "artifact_path",
                    "updated_at",
                )
            ),
            completion_runs_before,
        )
        self.assertEqual(
            list(HorseP0Source.objects.order_by("id").values()),
            sources_before,
        )

    def test_completed_replay_requires_exact_v2_publish_ledger_evidence(self):
        from unittest import mock

        from stable.models import (
            HorseP0Source,
            HorseProfileCompletionRun,
            HorseProfileDataCandidate,
            HorseRaceRecord,
            OperationLog,
            TaskExecutionLog,
        )
        from stable.services import p0_horse_completion_commit as commit_module
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )

        self._run_pipeline()
        batch_dir = self.manifest_path.parent
        state_before = BatchRunState.read(batch_dir).to_dict()
        ledger_path = batch_dir / "approvals_ledger.jsonl"
        original_entries = [
            json.loads(line)
            for line in ledger_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        publish_index = next(
            index
            for index, entry in enumerate(original_entries)
            if entry.get("event") == "auto_first_publish"
        )
        counted_models = (
            HorseProfile,
            HorseP0Source,
            HorseRaceRecord,
            HorseProfileDataCandidate,
            HorseProfileCompletionRun,
            OperationLog,
            TaskExecutionLog,
        )

        crash_shapes = {
            "missing": [
                entry
                for index, entry in enumerate(original_entries)
                if index != publish_index
            ],
            "mismatched_count": [
                (
                    {
                        **entry,
                        "published": int(entry["published"]) + 1,
                    }
                    if index == publish_index
                    else entry
                )
                for index, entry in enumerate(original_entries)
            ],
        }
        for name, entries in crash_shapes.items():
            with self.subTest(crash_shape=name):
                ledger_path.write_text(
                    "".join(
                        json.dumps(
                            entry,
                            ensure_ascii=False,
                            sort_keys=True,
                            separators=(",", ":"),
                        )
                        + "\n"
                        for entry in entries
                    ),
                    encoding="utf-8",
                )
                ledger_before = ledger_path.read_bytes()
                database_counts_before = {
                    model: model.objects.count() for model in counted_models
                }
                with mock.patch.object(
                    commit_module,
                    "dry_run_reviewed_p0_completion_artifact",
                    wraps=commit_module.dry_run_reviewed_p0_completion_artifact,
                ) as dry_run_mock, mock.patch.object(
                    commit_module,
                    "commit_reviewed_p0_completion_artifact",
                    wraps=commit_module.commit_reviewed_p0_completion_artifact,
                ) as database_apply_mock, mock.patch.object(
                    commit_module,
                    "_run_region_publish",
                    wraps=commit_module._run_region_publish,
                ) as publish_mock:
                    with self.assertRaisesRegex(
                        P0HorseBatchError,
                        "manual audit",
                    ):
                        commit_module.commit_p0_horse_batch_region(
                            self.manifest_path,
                            region="japan",
                            reviewer=self.reviewer,
                            approved_by="human-approver",
                            release_candidate_sha256=self.release_candidate[
                                "release_candidate_sha256"
                            ],
                            state_dir=self.state_dir,
                            confirm_reviewed_artifact=True,
                        )
                dry_run_mock.assert_not_called()
                database_apply_mock.assert_not_called()
                publish_mock.assert_not_called()
                self.assertEqual(
                    BatchRunState.read(batch_dir).to_dict(),
                    state_before,
                )
                self.assertEqual(ledger_path.read_bytes(), ledger_before)
                self.assertEqual(
                    {model: model.objects.count() for model in counted_models},
                    database_counts_before,
                )
        ledger_path.write_text(
            "".join(
                json.dumps(
                    entry,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                )
                + "\n"
                for entry in original_entries
            ),
            encoding="utf-8",
        )

    def test_repeated_commit_does_not_expand_completed_frozen_exclusion(self):
        from unittest import mock

        from stable.services import horse_profile_publish
        from stable.services.p0_horse_completion_batch import BatchRunState
        from stable.services.p0_horse_completion_commit import (
            commit_p0_horse_batch_region,
        )

        self.profile.manual_lock_flags = {"auto_publish_blocked": True}
        self.profile.save(update_fields=["manual_lock_flags"])
        committed = self._run_pipeline()
        self.assertEqual(
            committed["auto_first_publish"]["frozen_exclusion_counts"],
            {"block_manual_lock": 1},
        )
        batch_dir = self.manifest_path.parent
        state_before = BatchRunState.read(batch_dir)
        publish_before = json.loads(
            json.dumps(state_before.artifacts["publish:japan"])
        )
        ledger_path = batch_dir / "approvals_ledger.jsonl"
        ledger_before = ledger_path.read_bytes()

        self.profile.manual_lock_flags = {}
        self.profile.save(update_fields=["manual_lock_flags"])
        with mock.patch.object(
            horse_profile_publish,
            "auto_publish_profiles",
            return_value={
                "published": 0,
                "skipped_already_published": 0,
                "blocked": 0,
                "blocked_reasons": {},
                "published_profile_ids": [],
                "errors": [],
            },
        ) as publish_mock:
            repeated = commit_p0_horse_batch_region(
                self.manifest_path,
                region="japan",
                reviewer=self.reviewer,
                approved_by="human-approver",
                release_candidate_sha256=self.release_candidate[
                    "release_candidate_sha256"
                ],
                state_dir=self.state_dir,
                confirm_reviewed_artifact=True,
            )
        publish_mock.assert_not_called()
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.review_status, "ready")
        state_after = BatchRunState.read(batch_dir)
        self.assertEqual(
            state_after.artifacts["publish:japan"],
            publish_before,
        )
        self.assertEqual(
            state_after.completed_stages.count("publish:japan"),
            1,
        )
        self.assertEqual(ledger_path.read_bytes(), ledger_before)
        self.assertEqual(repeated["auto_first_publish"], publish_before)

    def test_repeated_commit_rejects_incomplete_publish_and_requires_retry(self):
        from unittest import mock

        from stable.services import horse_profile_publish
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )
        from stable.services.p0_horse_completion_commit import (
            commit_p0_horse_batch_region,
        )

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        self.release_candidate = self._prepare_release_candidate()
        failed_report = {
            "published": 0,
            "skipped_already_published": 0,
            "blocked": 0,
            "blocked_reasons": {},
            "published_profile_ids": [],
            "errors": [{"profile_id": self.profile.pk, "error": "boom"}],
        }
        with mock.patch.object(
            horse_profile_publish,
            "auto_publish_profiles",
            return_value=failed_report,
        ):
            with self.assertRaises(P0HorseBatchError):
                commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=self.release_candidate[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )

        batch_dir = self.manifest_path.parent
        state_before = BatchRunState.read(batch_dir)
        publish_before = json.loads(
            json.dumps(state_before.artifacts["publish:japan"])
        )
        ledger_path = batch_dir / "approvals_ledger.jsonl"
        ledger_before = ledger_path.read_bytes()
        successful_report = {
            "published": 1,
            "skipped_already_published": 0,
            "blocked": 0,
            "blocked_reasons": {},
            "published_profile_ids": [self.profile.pk],
            "errors": [],
        }
        with mock.patch.object(
            horse_profile_publish,
            "auto_publish_profiles",
            return_value=successful_report,
        ) as publish_mock:
            with self.assertRaisesRegex(
                P0HorseBatchError,
                "--retry-publish",
            ):
                commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=self.release_candidate[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
        publish_mock.assert_not_called()
        state_after = BatchRunState.read(batch_dir)
        self.assertEqual(
            state_after.artifacts["publish:japan"],
            publish_before,
        )
        self.assertNotIn("publish:japan", state_after.completed_stages)
        self.assertEqual(ledger_path.read_bytes(), ledger_before)

    def test_third_publish_attempt_preserves_all_cumulative_ids(self):
        from unittest import mock

        from stable.services import horse_profile_publish
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            load_batch_manifest,
        )
        from stable.services.p0_horse_completion_commit import (
            _run_region_publish,
        )

        BatchRunState.create(
            batch_id=self.approved["batch_id"],
            run_dir=self.manifest_path.parent,
        )
        reports = [
            {
                "published": 1,
                "skipped_already_published": 0,
                "blocked": 0,
                "blocked_reasons": {},
                "published_profile_ids": [profile_id],
                "errors": [],
            }
            for profile_id in (101, 102, 103)
        ]
        scope = {
            "existing_profiles": [
                {
                    "profile_id": self.profile.pk,
                    "review_status": "ready",
                    "hidden": False,
                    "manual_lock": False,
                    "disposition": "attempt_publish_after_commit",
                }
            ],
            "create_new_identities": [],
        }
        with mock.patch.object(
            horse_profile_publish,
            "auto_publish_profiles",
            side_effect=reports,
        ):
            for _ in range(3):
                _run_region_publish(
                    load_batch_manifest(self.manifest_path),
                    batch_dir=self.manifest_path.parent,
                    state_dir=self.state_dir,
                    region="japan",
                    artifact_sha="a" * 64,
                    reviewer=self.reviewer,
                    completion_run=None,
                    publish_scope=scope,
                )
        state = BatchRunState.read(self.manifest_path.parent)
        self.assertEqual(
            state.artifacts["publish:japan"][
                "cumulative_published_profile_ids"
            ],
            [101, 102, 103],
        )

    def test_batch_apply_marks_identity_keys_verified(self):
        self._run_pipeline()
        self.profile.refresh_from_db()
        verified = self.profile.source_refs.get("horse_identity_verified_keys") or []
        flat = self.profile.source_refs.get("horse_identity_keys") or []
        self.assertTrue(flat)
        self.assertEqual(sorted(verified), sorted(flat))

    def test_retry_publish_recovers_from_publish_error(self):
        from unittest import mock

        from stable.services import horse_profile_publish
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )
        from stable.services.p0_horse_completion_commit import (
            commit_p0_horse_batch_region,
        )

        real_publish = horse_profile_publish.auto_publish_profiles
        calls = {"count": 0}

        def flaky_publish(profiles, *, user, note):
            calls["count"] += 1
            if calls["count"] == 1:
                return {
                    "published": 0,
                    "skipped_already_published": 0,
                    "blocked": 0,
                    "blocked_reasons": {},
                    "published_profile_ids": [],
                    "errors": [{"profile_id": self.profile.pk, "error": "boom"}],
                }
            return real_publish(profiles, user=user, note=note)

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate = self._prepare_release_candidate()
        with mock.patch.object(
            horse_profile_publish,
            "auto_publish_profiles",
            side_effect=flaky_publish,
        ):
            with self.assertRaises(P0HorseBatchError):
                commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=candidate[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
        state = BatchRunState.read(self.manifest_path.parent)
        self.assertIn("publish:japan", state.artifacts)
        self.assertNotIn("publish:japan", state.completed_stages)
        self.assertTrue(
            any(entry.get("stage") == "publish:japan" for entry in state.errors)
        )
        self.profile.refresh_from_db()
        self.assertNotEqual(self.profile.review_status, "published")
        # manifest must not reach committed terminal state
        import json as jsonlib

        manifest = jsonlib.loads(self.manifest_path.read_text(encoding="utf-8"))
        self.assertNotEqual(manifest.get("status"), "committed")

        retried = self._call(
            "--retry-publish",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        publish = retried["auto_first_publish"]
        self.assertEqual(publish["published"], 1)
        self.assertEqual(publish["errors"], [])
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.review_status, "published")
        state = BatchRunState.read(self.manifest_path.parent)
        self.assertIn("publish:japan", state.completed_stages)
        self.assertFalse(
            any(entry.get("stage") == "publish:japan" for entry in state.errors)
        )
        manifest = jsonlib.loads(self.manifest_path.read_text(encoding="utf-8"))
        self.assertEqual(manifest.get("status"), "committed")

    def test_retry_publish_after_success_fails_closed(self):
        from stable.services.p0_horse_completion_batch import P0HorseBatchError
        from stable.services.p0_horse_completion_commit import retry_region_publish

        self._run_pipeline()
        with self.assertRaises(P0HorseBatchError):
            retry_region_publish(
                self.manifest_path,
                region="japan",
                reviewer=self.reviewer,
            )

    def test_retry_publish_old_state_without_scope_fails_before_publish(self):
        from unittest import mock

        from django.core.management.base import CommandError

        from stable.services import horse_profile_publish
        from stable.services.p0_horse_completion_batch import BatchRunState

        self._run_pipeline()
        batch_dir = self.manifest_path.parent
        state = BatchRunState.read(batch_dir)
        state.artifacts["commit:japan"].pop("publish_scope")
        state.artifacts.pop("publish:japan", None)
        state.completed_stages.remove("publish:japan")
        state.write()
        ledger_path = batch_dir / "approvals_ledger.jsonl"
        ledger_before = ledger_path.read_bytes()
        with mock.patch.object(
            horse_profile_publish, "auto_publish_profiles"
        ) as publish_mock:
            with self.assertRaisesRegex(CommandError, "publish_scope"):
                self._call(
                    "--retry-publish",
                    str(self.manifest_path),
                    "--region",
                    "japan",
                    "--reviewer-id",
                    str(self.reviewer.id),
                )
        publish_mock.assert_not_called()
        state = BatchRunState.read(batch_dir)
        self.assertNotIn("publish:japan", state.completed_stages)
        self.assertEqual(ledger_path.read_bytes(), ledger_before)

    def test_committed_candidate_recovers_from_input_snapshots_after_rebundle(self):
        from unittest import mock

        from django.core.management.base import CommandError

        from stable.services import horse_profile_publish
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )
        from stable.services.p0_horse_completion_commit import (
            commit_p0_horse_batch_region,
        )

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate_a = self._prepare_release_candidate()
        state = BatchRunState.read(self.manifest_path.parent)
        history_key = (
            "release_candidate:japan:"
            + candidate_a["release_candidate_sha256"]
        )
        history_a = state.artifacts[history_key]
        snapshot_bundle = history_a["snapshot_bundle"]
        artifact = json.loads(
            Path(history_a["artifact_path"]).read_text(encoding="utf-8")
        )
        for input_name, bundle_field in (
            ("research_v3", "research"),
            ("authority_manifest", "authority"),
            ("profile_mapping_decisions", "mapping"),
        ):
            snapshot = snapshot_bundle[bundle_field]
            self.assertEqual(
                artifact["inputs"][input_name]["path"], snapshot["path"]
            )
            self.assertIn(snapshot["sha256"], Path(snapshot["path"]).name)

        with mock.patch.object(
            horse_profile_publish,
            "auto_publish_profiles",
            return_value={
                "published": 0,
                "skipped_already_published": 0,
                "blocked": 0,
                "blocked_reasons": {},
                "published_profile_ids": [],
                "errors": [{"profile_id": self.profile.pk, "error": "crash"}],
            },
        ):
            with self.assertRaises(P0HorseBatchError):
                commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=candidate_a[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
        state = BatchRunState.read(self.manifest_path.parent)
        self.assertEqual(
            HorseProfileCompletionRun.objects.get(
                artifact_path=history_a["artifact_path"]
            ).status,
            "committed",
        )
        state.artifacts.pop("commit:japan")
        state.completed_stages.remove("commit:japan")
        state.artifacts.pop("publish:japan", None)
        state.write()
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        current_bundle = BatchRunState.read(
            self.manifest_path.parent
        ).artifacts["bundle:japan"]
        self.assertNotEqual(
            current_bundle["mapping_path"],
            snapshot_bundle["mapping"]["path"],
        )
        recovered = commit_p0_horse_batch_region(
            self.manifest_path,
            region="japan",
            reviewer=self.reviewer,
            approved_by="human-approver",
            release_candidate_sha256=candidate_a[
                "release_candidate_sha256"
            ],
            state_dir=self.state_dir,
            confirm_reviewed_artifact=True,
        )
        self.assertTrue(recovered["idempotent_verification"]["passed"])
        self.assertEqual(recovered["auto_first_publish"]["errors"], [])

        self.profile.refresh_from_db()
        self.profile.manual_lock_flags = {"auto_publish_blocked": True}
        self.profile.save(update_fields=["manual_lock_flags"])
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        with self.assertRaisesRegex(CommandError, "already committed"):
            self._prepare_release_candidate()

    def test_hidden_profile_not_auto_published(self):
        from django.utils import timezone

        self.profile.review_status = "hidden"
        self.profile.hidden_at = timezone.now()
        self.profile.save(update_fields=["review_status", "hidden_at"])
        committed = self._run_pipeline()
        publish = committed["auto_first_publish"]
        self.assertEqual(publish["published"], 0)
        self.assertEqual(publish["blocked"], 0)
        self.assertEqual(
            publish["frozen_exclusions"],
            [
                {
                    "target_type": "existing_profile",
                    "profile_id": self.profile.pk,
                    "disposition": "block_hidden",
                }
            ],
        )
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.review_status, "hidden")

    def test_inline_publish_never_expands_frozen_hidden_exclusion(self):
        from unittest import mock

        from django.utils import timezone

        from stable.services import p0_horse_completion_commit as commit_module

        self.profile.review_status = "hidden"
        self.profile.hidden_at = timezone.now()
        self.profile.save(update_fields=["review_status", "hidden_at"])
        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        self.release_candidate = self._prepare_release_candidate()
        real_commit = commit_module.commit_reviewed_p0_completion_artifact

        def commit_then_unhide(**kwargs):
            report = real_commit(**kwargs)
            self.profile.refresh_from_db()
            self.profile.review_status = "ready"
            self.profile.hidden_at = None
            self.profile.save(update_fields=["review_status", "hidden_at"])
            return report

        with mock.patch.object(
            commit_module,
            "commit_reviewed_p0_completion_artifact",
            side_effect=commit_then_unhide,
        ):
            committed = self._call(*self._commit_args())
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.review_status, "ready")
        self.assertEqual(committed["auto_first_publish"]["profile_ids"], [])
        self.assertEqual(
            committed["auto_first_publish"]["frozen_exclusion_counts"],
            {"block_hidden": 1},
        )

    def test_retry_publish_never_expands_frozen_manual_lock_exclusion(self):
        from unittest import mock

        from stable.services import p0_horse_completion_commit as commit_module
        from stable.services.p0_horse_completion_batch import P0HorseBatchError

        self.profile.manual_lock_flags = {"auto_publish_blocked": True}
        self.profile.save(update_fields=["manual_lock_flags"])
        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        self.release_candidate = self._prepare_release_candidate()
        with mock.patch.object(
            commit_module,
            "_run_region_publish",
            side_effect=P0HorseBatchError("crash before publish"),
        ):
            with self.assertRaises(P0HorseBatchError):
                commit_module.commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=self.release_candidate[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
        self.profile.manual_lock_flags = {}
        self.profile.save(update_fields=["manual_lock_flags"])
        retried = self._call(
            "--retry-publish",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.review_status, "ready")
        self.assertEqual(retried["auto_first_publish"]["profile_ids"], [])
        self.assertEqual(
            retried["auto_first_publish"]["frozen_exclusion_counts"],
            {"block_manual_lock": 1},
        )

    def test_failed_verification_blocks_publish(self):
        from unittest import mock

        from stable.services import p0_horse_completion_commit as commit_module
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )
        from stable.services.p0_horse_completion_commit import (
            commit_p0_horse_batch_region,
        )

        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate = self._prepare_release_candidate()
        real_dry_run = commit_module.dry_run_reviewed_p0_completion_artifact
        calls = {"count": 0}

        def fake_dry_run(**kwargs):
            calls["count"] += 1
            report = real_dry_run(**kwargs)
            if calls["count"] >= 2:
                report = {**report, "planned_profile_updates": 1}
            return report

        with mock.patch.object(
            commit_module,
            "dry_run_reviewed_p0_completion_artifact",
            side_effect=fake_dry_run,
        ):
            with self.assertRaises(P0HorseBatchError):
                commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=candidate[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
        self.profile.refresh_from_db()
        # apply promotes draft -> ready, but nothing may be published
        self.assertNotEqual(self.profile.review_status, "published")
        state = BatchRunState.read(self.manifest_path.parent)
        self.assertNotIn("publish:japan", state.artifacts)
        # retry-publish must also refuse: verification never passed
        from stable.services.p0_horse_completion_commit import retry_region_publish

        with self.assertRaises(P0HorseBatchError):
            retry_region_publish(
                self.manifest_path,
                region="japan",
                reviewer=self.reviewer,
            )

    def test_publish_retry_excludes_run_profiles_outside_candidate_scope(self):
        """仅发布候选冻结范围，不能因同一 run 关联而扩大发布集合。"""
        from unittest import mock

        from stable.models import (
            HorseP0Source,
            HorseP0SourceStatus,
            HorseP0SourceType,
            HorseProfileCompletionRun,
        )
        from stable.services import horse_profile_publish
        from stable.services.p0_horse_completion_batch import (
            BatchRunState,
            P0HorseBatchError,
        )
        from stable.services.p0_horse_completion_commit import (
            commit_p0_horse_batch_region,
        )

        real_publish = horse_profile_publish.auto_publish_profiles
        calls = {"count": 0}

        def flaky_publish(profiles, *, user, note):
            calls["count"] += 1
            if calls["count"] == 1:
                return {
                    "published": 0,
                    "skipped_already_published": 0,
                    "blocked": 0,
                    "blocked_reasons": {},
                    "published_profile_ids": [],
                    "errors": [{"profile_id": self.profile.pk, "error": "boom"}],
                }
            return real_publish(profiles, user=user, note=note)

        created = self._profile(
            "新建马",
            source_refs={"horse_identity_verified_keys": ["netkeiba:2020100001"]},
        )
        self._call(
            "--prepare",
            str(self.manifest_path),
            "--expected-sha256",
            self.approved["batch_sha256"],
        )
        self._call(
            "--bundle",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        candidate = self._prepare_release_candidate()
        with mock.patch.object(
            horse_profile_publish,
            "auto_publish_profiles",
            side_effect=flaky_publish,
        ):
            with self.assertRaises(P0HorseBatchError):
                commit_p0_horse_batch_region(
                    self.manifest_path,
                    region="japan",
                    reviewer=self.reviewer,
                    approved_by="human-approver",
                    release_candidate_sha256=candidate[
                        "release_candidate_sha256"
                    ],
                    state_dir=self.state_dir,
                    confirm_reviewed_artifact=True,
                )
        run = HorseProfileCompletionRun.objects.order_by("-id").first()
        self.assertIsNotNone(run)
        HorseP0Source.objects.create(
            profile=created,
            source_type=HorseP0SourceType.MAJOR_RACE_PARTICIPANT,
            status=HorseP0SourceStatus.ACTIVE,
            racing_region="japan",
            horse_name=created.original_name,
            participant_key=f"test-created:{created.pk}",
            source_url="https://example.test/race/created",
            completion_run=run,
        )
        retried = self._call(
            "--retry-publish",
            str(self.manifest_path),
            "--region",
            "japan",
            "--reviewer-id",
            str(self.reviewer.id),
        )
        publish = retried["auto_first_publish"]
        self.assertEqual(publish["published"], 1)
        self.assertEqual(publish["errors"], [])
        created.refresh_from_db()
        self.assertNotEqual(created.review_status, "published")

    def test_locked_profile_not_auto_published(self):
        self.profile.manual_lock_flags = {"auto_publish_blocked": True}
        self.profile.save(update_fields=["manual_lock_flags"])
        committed = self._run_pipeline()
        publish = committed["auto_first_publish"]
        self.assertEqual(publish["published"], 0)
        self.assertEqual(publish["blocked"], 0)
        self.assertEqual(
            publish["frozen_exclusion_counts"],
            {"block_manual_lock": 1},
        )
        self.profile.refresh_from_db()
        self.assertNotEqual(self.profile.review_status, "published")

    def test_regions_pending_commit_or_publish_helper(self):
        from stable.services.p0_horse_completion_batch import BatchRunState
        from stable.services.p0_horse_completion_commit import (
            _regions_pending_commit_or_publish,
        )

        state = BatchRunState(
            batch_id="test-batch",
            run_dir=self.state_dir,
            stage="commit",
            completed_stages=["commit:japan", "publish:japan", "commit:hong_kong"],
            artifacts={},
        )
        manifest = {"regions": ["japan", "hong_kong"]}
        self.assertEqual(
            _regions_pending_commit_or_publish(manifest, state), ["hong_kong"]
        )
        state.completed_stages.append("publish:hong_kong")
        self.assertEqual(_regions_pending_commit_or_publish(manifest, state), [])
