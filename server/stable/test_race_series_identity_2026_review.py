from __future__ import annotations

import hashlib
import importlib
import importlib.util
import json
import tempfile
import unittest
from copy import deepcopy
from datetime import date
from pathlib import Path

from django.core.management import call_command
from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from django.test import SimpleTestCase
from openpyxl import load_workbook

from stable.models import (
    HistoricalRaceEventTarget,
    RaceEvent,
    RaceEventAlias,
    RaceEventStatus,
    RaceEventSurface,
    RaceSeries,
    RaceSeriesName,
    RacingRegion,
)
from stable.services.race_series_identity_review import prepare_race_series_identity_review


MODULE_NAME = "stable.services.race_series_identity_2026_review"
MODULE_SPEC = importlib.util.find_spec(MODULE_NAME)


def _service():
    return importlib.import_module(MODULE_NAME)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _classification(
    target_id: int,
    *,
    classification: str,
    reason: str = "",
    candidate_event_ids: list[int] | None = None,
    country_region: str = "japan",
    event_id: int | None = None,
) -> dict:
    return {
        "target_id": target_id,
        "target_identity": {
            "payload": {"id": target_id, "race_series_id": 1000 + target_id},
            "sha256": f"{target_id:064x}",
        },
        "country_region": country_region,
        "year": 2026,
        "series_id": 1000 + target_id,
        "series_key": f"historical-{target_id}",
        "expectation_status": "held",
        "resolution_status": "pending",
        "before_event_id": None,
        "event_id": event_id,
        "candidate_event_identity": (
            {
                "payload": {
                    "id": event_id,
                    "race_series_id": 2000 + target_id,
                    "original_name": f"Race {target_id}",
                },
                "sha256": f"{event_id:064x}",
            }
            if event_id
            else None
        ),
        "candidate_event_ids": list(candidate_event_ids or ([] if event_id is None else [event_id])),
        "classification": classification,
        "reason": reason,
        "target_original_name": f"Race {target_id}",
        "target_chinese_name": f"赛事 {target_id}",
        "public_source_urls": [f"https://example.test/races/{target_id}"],
    }


def _review_fixture() -> list[dict]:
    return [
        _classification(1, classification="already_linked", event_id=101),
        _classification(2, classification="identity_conflict", reason="series_mismatch", event_id=102),
        _classification(
            3,
            classification="identity_conflict",
            reason="ambiguous_name_match",
            candidate_event_ids=[103, 203],
        ),
        _classification(4, classification="missing_event", reason="no_series_year_event"),
        _classification(5, classification="status_conflict", reason="not_held_target"),
    ]


def _compatible_facts() -> dict[int, dict]:
    return {
        2: {
            "source_series_id": 2002,
            "destination_series_id": 1002,
            "event_id": 102,
            "source_annual_event_ids": [102],
            "source_target_ids": [],
            "source_name_ids": [],
            "source_relation_ids": [],
            "destination_year_event_ids": [],
            "event_owner_target_id": None,
            "do_not_merge": False,
            "region_matches": True,
            "year_matches": True,
            "status_compatible": True,
            "detail_consistent": True,
            "source_series_identity_sha256": "b" * 64,
            "destination_series_identity_sha256": "c" * 64,
        }
    }


@unittest.skipUnless(MODULE_SPEC is not None, "target review adapter is not implemented yet")
class RaceSeriesIdentity2026ClassificationTests(SimpleTestCase):
    def test_partition_is_exhaustive_and_conserves_every_target_once(self):
        snapshot = _service().build_review_snapshot(
            classifications=_review_fixture(),
            alias_suggestions_by_target={},
            dependency_facts=_compatible_facts(),
        )

        self.assertEqual(snapshot["counts"]["total_targets"], 5)
        self.assertEqual(snapshot["counts"]["already_linked"], 1)
        self.assertEqual(snapshot["counts"]["unique_series_mismatch"], 1)
        self.assertEqual(snapshot["counts"]["ambiguous_name_match"], 1)
        self.assertEqual(snapshot["counts"]["no_name_match"], 1)
        self.assertEqual(snapshot["counts"]["not_held"], 1)
        all_ids = [row["target_id"] for row in snapshot["all_rows"]]
        self.assertEqual(sorted(all_ids), [1, 2, 3, 4, 5])
        self.assertEqual(len(all_ids), len(set(all_ids)))
        self.assertEqual(snapshot["anomalies"], [])

    def test_alias_suggestions_do_not_change_the_base_bucket(self):
        classifications = [
            _classification(4, classification="missing_event", reason="no_series_year_event")
        ]
        snapshot = _service().build_review_snapshot(
            classifications=classifications,
            alias_suggestions_by_target={4: [{"event_id": 404, "source": "event_alias"}]},
            dependency_facts={},
        )

        row = snapshot["sheets"]["无名称匹配"][0]
        self.assertEqual(row["review_bucket"], "no_name_match")
        self.assertEqual(row["supplemental_suggestions"][0]["event_id"], 404)
        self.assertEqual(snapshot["sheets"]["唯一名称匹配"], [])

    def test_unknown_reason_is_preserved_as_a_blocking_anomaly(self):
        snapshot = _service().build_review_snapshot(
            classifications=[
                _classification(9, classification="identity_conflict", reason="future_reason")
            ],
            alias_suggestions_by_target={},
            dependency_facts={},
        )

        self.assertEqual(snapshot["counts"]["total_targets"], 1)
        self.assertEqual(snapshot["anomalies"][0]["target_id"], 9)
        self.assertTrue(snapshot["blocks_decisions"])

    def test_cross_region_name_match_cannot_be_a_positive_candidate(self):
        row = _classification(
            10,
            classification="identity_conflict",
            reason="series_mismatch",
            event_id=110,
            country_region="japan",
        )
        facts = _compatible_facts()[2] | {
            "source_series_id": 2010,
            "destination_series_id": 1010,
            "event_id": 110,
            "region_matches": False,
        }
        snapshot = _service().build_review_snapshot(
            classifications=[row],
            alias_suggestions_by_target={},
            dependency_facts={10: facts},
        )

        candidate = snapshot["sheets"]["唯一名称匹配"][0]
        self.assertFalse(candidate["engine_compatible"])
        self.assertIn("region_matches", candidate["compatibility_failures"])


@unittest.skipUnless(MODULE_SPEC is not None, "target review adapter is not implemented yet")
class RaceSeriesIdentity2026CompatibilityTests(SimpleTestCase):
    def test_compatibility_requires_all_dependencies_and_disjoint_batch_identity(self):
        classifications = [
            _classification(2, classification="identity_conflict", reason="series_mismatch", event_id=102),
            _classification(6, classification="identity_conflict", reason="series_mismatch", event_id=106),
        ]
        shared = _compatible_facts()[2]
        facts = {
            2: shared,
            6: shared
            | {
                "event_id": 106,
                "source_annual_event_ids": [106, 99],
                "source_target_ids": [88],
                "do_not_merge": True,
            },
        }
        snapshot = _service().build_review_snapshot(
            classifications=classifications,
            alias_suggestions_by_target={},
            dependency_facts=facts,
        )

        rows = {row["target_id"]: row for row in snapshot["sheets"]["唯一名称匹配"]}
        self.assertFalse(rows[2]["engine_compatible"])
        self.assertIn("duplicate_source_series", rows[2]["compatibility_failures"])
        self.assertFalse(rows[6]["engine_compatible"])
        self.assertGreaterEqual(
            set(rows[6]["compatibility_failures"]),
            {"source_has_other_events", "source_has_targets", "do_not_merge", "duplicate_source_series"},
        )

    def test_compatible_row_is_not_an_action_without_an_explicit_decision(self):
        snapshot = _service().build_review_snapshot(
            classifications=[
                _classification(2, classification="identity_conflict", reason="series_mismatch", event_id=102)
            ],
            alias_suggestions_by_target={},
            dependency_facts=_compatible_facts(),
        )

        row = snapshot["sheets"]["唯一名称匹配"][0]
        self.assertTrue(row["engine_compatible"])
        self.assertEqual(row["decision"], "defer")


@unittest.skipUnless(MODULE_SPEC is not None, "target review adapter is not implemented yet")
class RaceSeriesIdentity2026PackageTests(SimpleTestCase):
    def _write_package(
        self,
        root: Path,
        *,
        target_id: int = 2,
        engine_compatible: bool = True,
        locked_series_identity: bool = True,
    ) -> tuple[Path, dict]:
        classification = _classification(
            target_id,
            classification="identity_conflict",
            reason="series_mismatch",
            event_id=100 + target_id,
        )
        facts = _compatible_facts()[2] | {
            "source_series_id": 2000 + target_id,
            "destination_series_id": 1000 + target_id,
            "event_id": 100 + target_id,
            "source_annual_event_ids": [100 + target_id],
        }
        if not engine_compatible:
            facts["source_name_ids"] = [9001]
        if not locked_series_identity:
            facts.pop("source_series_identity_sha256", None)
        snapshot = _service().build_review_snapshot(
            classifications=[classification],
            alias_suggestions_by_target={},
            dependency_facts={target_id: facts},
        )
        package = root / f"package-{target_id}"
        result = _service().write_review_package(
            snapshot=snapshot,
            output_dir=package,
            production_head="f" * 40,
            as_of="2026-07-23T08:00:00+00:00",
        )
        return package, result

    def test_package_has_canonical_files_and_exact_six_sheet_workbook(self):
        with tempfile.TemporaryDirectory() as temporary:
            package, result = self._write_package(Path(temporary))

            self.assertEqual(
                {path.name for path in package.iterdir()},
                {"snapshot.json", "review.json", "review.csv", "review.xlsx", "manifest.json"},
            )
            workbook = load_workbook(package / "review.xlsx", data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "审核说明",
                    "唯一名称匹配",
                    "同名多候选",
                    "无名称匹配",
                    "未举办",
                    "异常清单",
                ],
            )
            unique_sheet = workbook["唯一名称匹配"]
            headers = [cell.value for cell in unique_sheet[1]]
            self.assertEqual(
                headers[:8],
                [
                    "decision",
                    "review_note",
                    "engine_compatible",
                    "compatibility_failures",
                    "target_chinese_name",
                    "target_original_name",
                    "event_chinese_name",
                    "event_original_name",
                ],
            )
            self.assertEqual(unique_sheet.freeze_panes, "C2")
            self.assertEqual(unique_sheet.auto_filter.ref, unique_sheet.dimensions)
            self.assertGreaterEqual(unique_sheet.column_dimensions["A"].width, 20)
            self.assertGreaterEqual(unique_sheet.column_dimensions["B"].width, 35)
            self.assertGreaterEqual(unique_sheet.column_dimensions["E"].width, 20)
            self.assertGreaterEqual(unique_sheet.row_dimensions[1].height, 36)
            self.assertEqual(unique_sheet["A1"].font.color.rgb, "00FFFFFF")
            self.assertNotEqual(unique_sheet["A1"].fill.fgColor.rgb, unique_sheet["C1"].fill.fgColor.rgb)
            validations = list(unique_sheet.data_validations.dataValidation)
            self.assertEqual(len(validations), 1)
            self.assertEqual(validations[0].type, "list")
            self.assertIn("merge_and_link", validations[0].formula1)
            self.assertIn("A2", str(validations[0].sqref))
            for sheet_name in ("同名多候选", "无名称匹配", "未举办", "异常清单"):
                sheet = workbook[sheet_name]
                self.assertEqual(sheet.freeze_panes, "C2")
                self.assertEqual(len(sheet.data_validations.dataValidation), 0)
            self.assertEqual(result["manifest_sha256"], _sha256(package / "manifest.json"))

    def test_exported_payload_recursively_excludes_sensitive_keys_and_values(self):
        snapshot = _service().build_review_snapshot(
            classifications=[
                _classification(2, classification="identity_conflict", reason="series_mismatch", event_id=102)
                | {
                    "source_refs": {"official": "https://example.test/race", "token": "secret-value"},
                    "module_statuses": {"private": True},
                    "notes": "password=hunter2",
                    "manual_lock_flags": {"results": True},
                }
            ],
            alias_suggestions_by_target={},
            dependency_facts=_compatible_facts(),
        )

        encoded = json.dumps(snapshot, ensure_ascii=False).casefold()
        for forbidden in (
            "source_refs",
            "module_statuses",
            "manual_lock_flags",
            "notes",
            "secret-value",
            "hunter2",
            "password=",
        ):
            self.assertNotIn(forbidden, encoded)

    def test_original_manifest_and_cross_package_workbook_are_fail_closed(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            package_a, result_a = self._write_package(root, target_id=2)
            package_b, result_b = self._write_package(root, target_id=7)
            reviewed = root / "reviewed.xlsx"
            reviewed.write_bytes((package_a / "review.xlsx").read_bytes())

            with self.assertRaisesRegex(_service().RaceSeriesIdentity2026ReviewError, "manifest"):
                _service().build_decisions_from_reviewed_workbook(
                    original_package_dir=package_a,
                    expected_manifest_sha256="0" * 64,
                    reviewed_workbook=reviewed,
                    expected_workbook_sha256=_sha256(reviewed),
                )
            with self.assertRaises(_service().RaceSeriesIdentity2026ReviewError):
                _service().build_decisions_from_reviewed_workbook(
                    original_package_dir=package_b,
                    expected_manifest_sha256=result_b["manifest_sha256"],
                    reviewed_workbook=reviewed,
                    expected_workbook_sha256=_sha256(reviewed),
                )
            self.assertNotEqual(result_a["manifest_sha256"], result_b["manifest_sha256"])

    def test_machine_column_or_non_unique_sheet_edit_is_rejected(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            package, result = self._write_package(root)
            reviewed = root / "reviewed.xlsx"
            workbook = load_workbook(package / "review.xlsx", data_only=False)
            sheet = workbook["唯一名称匹配"]
            headers = {cell.value: cell.column for cell in sheet[1]}
            sheet.cell(2, headers["target_id"], 999)
            workbook.save(reviewed)

            with self.assertRaises(_service().RaceSeriesIdentity2026ReviewError):
                _service().build_decisions_from_reviewed_workbook(
                    original_package_dir=package,
                    expected_manifest_sha256=result["manifest_sha256"],
                    reviewed_workbook=reviewed,
                    expected_workbook_sha256=_sha256(reviewed),
                )

    def test_reviewed_positive_and_negative_rows_convert_to_existing_decisions_contract(self):
        for action in ("merge_and_link", "keep_independent", "ignore_false_match"):
            with self.subTest(action=action), tempfile.TemporaryDirectory() as temporary:
                root = Path(temporary)
                package, result = self._write_package(root)
                reviewed = root / "reviewed.xlsx"
                workbook = load_workbook(package / "review.xlsx", data_only=False)
                sheet = workbook["唯一名称匹配"]
                headers = {cell.value: cell.column for cell in sheet[1]}
                sheet.cell(2, headers["decision"], action)
                sheet.cell(2, headers["review_note"], "人工确认审核结论")
                workbook.save(reviewed)

                decisions = _service().build_decisions_from_reviewed_workbook(
                    original_package_dir=package,
                    expected_manifest_sha256=result["manifest_sha256"],
                    reviewed_workbook=reviewed,
                    expected_workbook_sha256=_sha256(reviewed),
                )

                self.assertEqual(
                    decisions["field_repairs"],
                    {"schema_version": "1.0", "repairs": []},
                )
                self.assertEqual(len(decisions["decisions"]["decisions"]), 1)
                row = decisions["decisions"]["decisions"][0]
                self.assertEqual(row["decision"], action)
                self.assertEqual(row["target_identity_sha256"], f"{2:064x}")
                self.assertEqual(row["event_identity_sha256"], f"{102:064x}")
                self.assertEqual(row["source_series_identity_sha256"], "b" * 64)
                self.assertEqual(
                    row["destination_series_identity_sha256"], "c" * 64
                )
                self.assertEqual(row["evidence"]["summary"], "人工确认审核结论")
                self.assertEqual(
                    row["evidence"]["source_urls"],
                    ["https://example.test/races/2"],
                )

    def test_non_defer_requires_note_public_url_and_technical_compatibility(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            package, result = self._write_package(root)
            reviewed = root / "reviewed.xlsx"
            workbook = load_workbook(package / "review.xlsx", data_only=False)
            sheet = workbook["唯一名称匹配"]
            headers = {cell.value: cell.column for cell in sheet[1]}
            sheet.cell(2, headers["decision"], "merge_and_link")
            sheet.cell(2, headers["review_note"], "")
            workbook.save(reviewed)

            with self.assertRaisesRegex(_service().RaceSeriesIdentity2026ReviewError, "review_note"):
                _service().build_decisions_from_reviewed_workbook(
                    original_package_dir=package,
                    expected_manifest_sha256=result["manifest_sha256"],
                    reviewed_workbook=reviewed,
                    expected_workbook_sha256=_sha256(reviewed),
                )

    def test_only_positive_merge_requires_engine_compatibility(self):
        for action, should_succeed in (
            ("merge_and_link", False),
            ("keep_independent", True),
            ("ignore_false_match", True),
        ):
            with self.subTest(action=action), tempfile.TemporaryDirectory() as temporary:
                root = Path(temporary)
                package, result = self._write_package(
                    root, engine_compatible=False
                )
                reviewed = root / "reviewed.xlsx"
                workbook = load_workbook(package / "review.xlsx", data_only=False)
                sheet = workbook["唯一名称匹配"]
                headers = {cell.value: cell.column for cell in sheet[1]}
                sheet.cell(2, headers["decision"], action)
                sheet.cell(2, headers["review_note"], "人工确认系列对结论")
                workbook.save(reviewed)
                call = lambda: _service().build_decisions_from_reviewed_workbook(
                    original_package_dir=package,
                    expected_manifest_sha256=result["manifest_sha256"],
                    reviewed_workbook=reviewed,
                    expected_workbook_sha256=_sha256(reviewed),
                )
                if should_succeed:
                    output = call()
                    self.assertEqual(
                        output["decisions"]["decisions"][0]["decision"], action
                    )
                else:
                    with self.assertRaisesRegex(
                        _service().RaceSeriesIdentity2026ReviewError,
                        "not engine compatible",
                    ):
                        call()

    def test_non_defer_rejects_missing_locked_series_identity_sha(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            package, result = self._write_package(
                root, locked_series_identity=False
            )
            reviewed = root / "reviewed.xlsx"
            workbook = load_workbook(package / "review.xlsx", data_only=False)
            sheet = workbook["唯一名称匹配"]
            headers = {cell.value: cell.column for cell in sheet[1]}
            sheet.cell(2, headers["decision"], "keep_independent")
            sheet.cell(2, headers["review_note"], "人工确认保持独立")
            workbook.save(reviewed)

            with self.assertRaisesRegex(
                _service().RaceSeriesIdentity2026ReviewError,
                "target/event/series identity SHA",
            ):
                _service().build_decisions_from_reviewed_workbook(
                    original_package_dir=package,
                    expected_manifest_sha256=result["manifest_sha256"],
                    reviewed_workbook=reviewed,
                    expected_workbook_sha256=_sha256(reviewed),
                )


class RaceSeriesIdentity2026RedGateTests(SimpleTestCase):
    def test_target_review_adapter_module_exists(self):
        self.assertIsNotNone(
            MODULE_SPEC,
            "stable.services.race_series_identity_2026_review must be implemented",
        )


@unittest.skipUnless(MODULE_SPEC is not None, "target review adapter is not implemented yet")
class RaceSeriesIdentity2026OrmAndCommandTests(TestCase):
    def setUp(self):
        self.destination = RaceSeries.objects.create(
            key="shared-race-history",
            country_region=RacingRegion.JAPAN,
            canonical_name_original="Shared Race",
            source_refs={"official": "https://history.example.test/shared"},
        )
        self.source = RaceSeries.objects.create(
            key="shared-race-2026",
            country_region=RacingRegion.JAPAN,
            canonical_name_original="Shared Race 2026",
        )
        self.target = HistoricalRaceEventTarget.objects.create(
            race_series=self.destination,
            year=2026,
            country_region=RacingRegion.JAPAN,
            original_name="Shared Race",
            chinese_name="共同赛事",
            racecourse="Tokyo",
            grade_text="G3",
            normalized_grade="G3",
            surface=RaceEventSurface.TURF,
            distance_text="1600m",
            local_date=date(2026, 5, 1),
            source_refs={"result_url": "https://result.example.test/shared"},
        )
        self.event = RaceEvent.objects.create(
            race_series=self.source,
            series_key=self.source.key,
            year=2026,
            slug="shared-race-2026",
            original_name="Shared Race",
            chinese_name="共同赛事",
            country_region=RacingRegion.JAPAN,
            racecourse="Tokyo",
            grade_text="G3",
            normalized_grade="G3",
            surface=RaceEventSurface.TURF,
            distance_text="1600m",
            local_date=date(2026, 5, 1),
            status=RaceEventStatus.SCHEDULED,
            source_refs={
                "official": "https://event.example.test/shared",
                "url": "https://user:password@example.test/private",
                "source_url": "https://example.test/race?token=secret",
            },
        )
        unmatched_series = RaceSeries.objects.create(
            key="unmatched-history",
            country_region=RacingRegion.JAPAN,
            canonical_name_original="Unmatched Historical Race",
        )
        self.unmatched_target = HistoricalRaceEventTarget.objects.create(
            race_series=unmatched_series,
            year=2026,
            country_region=RacingRegion.JAPAN,
            original_name="Unmatched Historical Race",
        )
        RaceSeriesName.objects.create(series=unmatched_series, text="Shared Sponsored Race")
        RaceEventAlias.objects.create(event=self.event, text="Shared Sponsored Race")

    def _prepare_review_action(self, root: Path, *, action: str) -> dict:
        snapshot = _service().export_2026_review_snapshot()
        snapshot["blocks_decisions"] = False
        package = root / "package"
        package_result = _service().write_review_package(
            snapshot=snapshot,
            output_dir=package,
            production_head="f" * 40,
            as_of=snapshot["as_of"],
        )
        reviewed = root / "reviewed.xlsx"
        workbook = load_workbook(package / "review.xlsx", data_only=False)
        sheet = workbook["唯一名称匹配"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        sheet.cell(2, headers["decision"], action)
        sheet.cell(2, headers["review_note"], "人工核对系列对后确认")
        workbook.save(reviewed)
        built = _service().build_decisions_from_reviewed_workbook(
            original_package_dir=package,
            expected_manifest_sha256=package_result["manifest_sha256"],
            reviewed_workbook=reviewed,
            expected_workbook_sha256=_sha256(reviewed),
        )
        decisions_path = root / "decisions.json"
        repairs_path = root / "field-repairs.json"
        decisions_path.write_text(
            json.dumps(built["decisions"], ensure_ascii=False), encoding="utf-8"
        )
        repairs_path.write_text(
            json.dumps(built["field_repairs"], ensure_ascii=False), encoding="utf-8"
        )
        return prepare_race_series_identity_review(
            decisions_path=decisions_path,
            field_repairs_path=repairs_path,
            output_dir=root / "prepared",
        )

    def test_orm_snapshot_is_batched_and_alias_is_only_supplemental(self):
        with CaptureQueriesContext(connection) as queries:
            snapshot = _service().export_2026_review_snapshot()

        self.assertLessEqual(len(queries), 30)
        self.assertEqual(snapshot["counts"]["total_targets"], 2)
        self.assertEqual(snapshot["counts"]["unique_series_mismatch"], 1)
        self.assertEqual(snapshot["counts"]["no_name_match"], 1)
        row = snapshot["sheets"]["唯一名称匹配"][0]
        self.assertTrue(row["engine_compatible"])
        self.assertEqual(row["event_identity_sha256"], _service().event_identity(self.event)["sha256"])
        self.assertEqual(
            row["public_source_urls"],
            [
                "https://event.example.test/shared",
                "https://history.example.test/shared",
                "https://result.example.test/shared",
            ],
        )
        self.assertEqual(row["supplemental_suggestions"], [])
        unmatched = snapshot["sheets"]["无名称匹配"][0]
        self.assertEqual(unmatched["target_id"], self.unmatched_target.pk)
        self.assertEqual(
            unmatched["supplemental_suggestions"][0]["event_id"],
            self.event.pk,
        )
        self.assertTrue(snapshot["baseline_drift"])
        self.assertTrue(snapshot["blocks_decisions"])

    def test_reviewed_output_is_accepted_by_existing_prepare_and_command_has_no_commit_mode(self):
        snapshot = _service().export_2026_review_snapshot()
        snapshot["blocks_decisions"] = False
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            package = root / "package"
            package_result = _service().write_review_package(
                snapshot=snapshot,
                output_dir=package,
                production_head="f" * 40,
                as_of=snapshot["as_of"],
            )
            reviewed = root / "reviewed.xlsx"
            workbook = load_workbook(package / "review.xlsx", data_only=False)
            sheet = workbook["唯一名称匹配"]
            headers = {cell.value: cell.column for cell in sheet[1]}
            sheet.cell(2, headers["decision"], "merge_and_link")
            sheet.cell(2, headers["review_note"], "人工核对官方赛历后确认")
            workbook.save(reviewed)
            decisions_dir = root / "decisions"
            call_command(
                "review_2026_race_series_identities",
                "--build-decisions",
                "--output-dir",
                str(decisions_dir),
                "--original-package-dir",
                str(package),
                "--expected-manifest-sha256",
                package_result["manifest_sha256"],
                "--reviewed-workbook",
                str(reviewed),
                "--expected-workbook-sha256",
                _sha256(reviewed),
            )
            prepared = prepare_race_series_identity_review(
                decisions_path=decisions_dir / "decisions.json",
                field_repairs_path=decisions_dir / "field-repairs.json",
                output_dir=root / "prepared",
            )

            self.assertEqual(prepared["positive_action_count"], 1)
            from stable.management.commands.review_2026_race_series_identities import Command

            parser = Command().create_parser("manage.py", "review_2026_race_series_identities")
            self.assertNotIn("--commit", parser.format_help())

    def test_default_management_command_only_exports_the_five_file_review_package(self):
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "export"
            call_command(
                "review_2026_race_series_identities",
                "--output-dir",
                str(output),
                "--production-head",
                "a" * 40,
            )

            self.assertEqual(
                {path.name for path in output.iterdir()},
                {
                    "snapshot.json",
                    "review.json",
                    "review.csv",
                    "review.xlsx",
                    "manifest.json",
                },
            )
            manifest = json.loads((output / "manifest.json").read_text(encoding="utf-8"))
            self.assertTrue(manifest["blocks_decisions"])

    def test_positive_incompatible_same_region_pair_can_prepare_negative_lock(self):
        RaceSeriesName.objects.create(series=self.source, text="Source dependency")
        with tempfile.TemporaryDirectory() as temporary:
            prepared = self._prepare_review_action(
                Path(temporary), action="keep_independent"
            )

        self.assertEqual(prepared["positive_action_count"], 0)
        self.assertEqual(prepared["negative_pair_count"], 1)

    def test_cross_region_pair_can_prepare_ignore_false_match(self):
        self.source.country_region = RacingRegion.UNITED_KINGDOM
        self.source.save(update_fields={"country_region"})
        self.event.country_region = RacingRegion.UNITED_KINGDOM
        self.event.save(update_fields={"country_region"})
        with tempfile.TemporaryDirectory() as temporary:
            prepared = self._prepare_review_action(
                Path(temporary), action="ignore_false_match"
            )

        self.assertEqual(prepared["positive_action_count"], 0)
        self.assertEqual(prepared["negative_pair_count"], 1)
